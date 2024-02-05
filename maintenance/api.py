import json
import sys

from django.contrib.auth.models import User
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from openpyxl.pivot.table import Location

from admin_panel.anton import new_sms
from admin_panel.models import Locations, UserAddOns
from maintenance.models import Maintenancex, MaintenanceHistory, MaintenanceAssetGroup, MaintenanceAssetSubGroup, \
    MaintenanceAsset, Maintenance, WordOrderTransactions, WorkOrder, WorkOrderMaterial


@csrf_exempt
def interface(request):
    response = {
        'status_code': 0,
        'message': ""
    }

    success_response = {
        'status_code': 200,
        'message': "Procedure Completed Successfully"
    }

    # get method
    method = request.method

    try:

        body = json.loads(request.body)
        module = body.get('module')
        data = body.get('data')
        print(body)

        if method == 'PUT':
            if module == 'maintenance_asset_group':
                name = data.get('name')
                user_pk = data.get('mypk')
                owner = User.objects.get(pk=user_pk)
                MaintenanceAssetGroup(name=name, owner=owner).save()
                success_response['message'] = "Asset Group Created Successfully"

            elif module == 'maintenance_asset_sub_group':
                name = data.get('sub_group')
                user_pk = data.get('mypk')
                group_pk = data.get('group')
                group = MaintenanceAssetGroup.objects.get(name=group_pk)
                owner = User.objects.get(pk=user_pk)
                MaintenanceAssetSubGroup(group=group, owner=owner, name=name).save()
                success_response['message'] = "Asset SubGroup Created Successfully"

            elif module == 'maintenance_asset':
                sku = data.get('sku')
                group_pk = data.get('group')
                sub_group_pk = data.get('sub_group')
                group = MaintenanceAssetGroup.objects.get(pk=group_pk)
                subgroup = MaintenanceAssetSubGroup.objects.get(pk=sub_group_pk)
                user_pk = data.get('mypk')
                brand = data.get('brand')
                color = data.get('color')
                location_pk = data.get('location')
                origin = data.get('origin')
                location = Locations.objects.get(pk=location_pk)
                image = data.get('image')
                name = data.get('name')

                owner = User.objects.get(pk=user_pk)

                MaintenanceAsset(sku=sku, group=group, subgroup=subgroup, owner=owner, brand=brand, color=color,
                                 origin=origin, location=location, image=image,name=name).save()
                asset = MaintenanceAsset.objects.get(sku=sku)
                success_response['message'] = asset.pk


            elif module == 'maintenance_request':
                doc_ini = "MT24-"
                entry_no = f"{doc_ini}{Maintenance.objects.all().count() + 1}"
                title = data.get('title')
                asset_pk = data.get('asset')
                asset = MaintenanceAsset.objects.get(pk=asset_pk)
                mypk = data.get('mypk')
                owner = User.objects.get(pk=mypk)
                detail = data.get('detail')
                analyse = data.get('analyse')

                maintenance = Maintenance(title=title, description=detail, owner=owner,asset=asset,entry_no=entry_no,analyse=analyse)
                maintenance.save()
                sms_message = (f"NEW MAINTENANCE REQUEST\n\nTitle: {title}\n\nDescription: Sent Via Email or in "
                               f"ocean\n\nOwner: {owner.get_full_name()}")
                new_sms('0546310011', sms_message)
                success_response['message'] = entry_no

            elif module == 'maintenance_log':
                maintenance = data.get('maintenance')
                mypk = data.get('mypk')
                title = data.get('title')
                description = data.get('description')

                # get objects
                main = Maintenance.objects.get(pk=maintenance)
                owner = User.objects.get(pk=mypk)

                # add to history
                history = MaintenanceHistory(maintenance=main, owner=owner, description=description, title=title)
                history.save()
                success_response['message'] = "Record Updated"

            elif module == 'generate_wo':
                mypk = data.get('mypk')
                tech_pk = data.get('technician')
                analyse = data.get('analysis')
                entry_no = data.get('entry_no')

                wr = Maintenance.objects.get(entry_no=entry_no)
                owner = User.objects.get(pk=mypk)
                technician = User.objects.get(pk=tech_pk)
                wo_no = f"WO24{WorkOrder.objects.all().count()}"

                # generate work order
                WorkOrder(
                    wo_no=wo_no,
                    w_request=wr,
                    technician=technician,
                    owner=owner,
                    analyse=analyse
                ).save()

                # get just created entry
                wo = WorkOrder.objects.get(
                    wo_no=wo_no,
                    w_request=wr,
                    technician=technician,
                    owner=owner,
                    analyse=analyse
                )

                # add work order transaction
                title = "Work Order Generated"
                description = "Work order has been generated and assigned"
                WordOrderTransactions(wo=wo,title=title,description=description,owner=owner)

                # send sms to technician
                tech = UserAddOns.objects.get(user=technician)
                tech_phone = tech.phone
                message = f"TITLE : New Work Order\n"
                message += f"A new Work Order with entry number {wo_no} has been created and assigned to you.\n"
                message += f"http://ocean.snedaghana.loc/maintainance/wo-transactions/{wo_no}/"

                new_sms(tech_phone,message)

                wr.is_open = False


                # print(maintainance)
                # print(maintainance.pk)

                success_response['message'] = {
                    'wr_pk':wr.pk,
                    'msg':"Work Order Generated"
                }

            elif module == 'wo_material':
                wo_no = data.get('wo_no')
                mat_name = data.get('mat_name')
                mat_descr = data.get('mat_descr')
                price = data.get('price')
                quantity = data.get('quantity')
                mypk = data.get('mypk')

                wo = WorkOrder.objects.get(wo_no=wo_no)
                owner = User.objects.get(pk=mypk)
                WorkOrderMaterial(wo=wo,name=mat_name,description=mat_descr,unit_price=price,quantity=quantity).save()
                success_response['message'] = wo.pk


            else:
                success_response['status_code'] = 404
                success_response['message'] = "Module Not Found"

        elif method == 'VIEW':
            arr = []
            if module == 'maintenance_asset_group':
                key = data.get('key', '*')
                if key == '*':
                    asset_groups = MaintenanceAssetGroup.objects.all()
                else:
                    asset_groups = MaintenanceAssetGroup.objects.filter(pk=key)

                # loop through assets objects
                for asset_group in asset_groups:
                    subgroups = asset_group.subgroups()
                    sg = []
                    for subgroup in subgroups:
                        sg.append({
                            'name': subgroup.name,
                            'pk': subgroup.pk
                        })
                    arr.append({
                        'pk': asset_group.pk,
                        'name': asset_group.name,
                        'is_active': asset_group.is_active,
                        'owner': {
                            'name': asset_group.owner.get_full_name(),
                            'username': asset_group.owner.username,
                            'pk': asset_group.owner.pk
                        },
                        'timestamp': {
                            'created_on': asset_group.created_on,
                            'modified_on': asset_group.modified_on
                        },
                        'subgroups': sg
                    })

                success_response['message'] = arr

            elif module == 'maintenance':
                print(data)
                key = data.get('key','*')
                if key == '*':
                    maints = Maintenance.objects.all()
                else:
                    maints = Maintenance.objects.filter(pk=key)

                for maint in maints:
                    arr.append({
                        'entry_no':maint.entry_no,
                        'asset':maint.asset_obj(),
                        'title':maint.title,
                        'description':maint.description,
                        'evidence':maint.evidence.url,
                        'analyse':maint.analyse,
                        'owner':maint.owner_obj(),
                        'is_open':maint.is_open,
                        'date':maint.date,
                        'time':maint.time,
                        'wo':maint.wo(),
                        'next':maint.next_entry(),
                        'previous':maint.previous_entry()
                    })

                # print(arr)
                success_response['message'] = arr

            elif module == 'maintenance_asset_sub_group':
                key = data.get('key', '*')
                if key == '*':
                    asset_sub_groups = MaintenanceAssetGroup.objects.all()
                else:
                    asset_sub_groups = MaintenanceAssetGroup.objects.filter(pk=key)

                for subgroup in asset_sub_groups:
                    arr.append(
                        {
                            'name': subgroup.name,
                            'pk': subgroup.pk
                        }
                    )

            elif module == 'maintenance_asset':
                key = data.get('key', '*')
                assets = MaintenanceAsset.objects.all()
                if key != '*':
                    assets = MaintenanceAsset.objects.filter(pk=key)

                for asset in assets:
                    arr.append({
                        'sku': asset.sku,
                        'pk': assets.pk,
                        'name': asset.name,
                        'group': {
                            'name': asset.group.name,
                            'pk': asset.group.pk
                        },
                        'subgroup': {
                            'pk': asset.subgroup.pk,
                            'name': asset.subgroup.name
                        },
                        'owner': asset.owner.get_fullname(),
                        'created_on': asset.created_on,
                        'image': asset.image_link()
                    })

            elif module == 'maint_asset':
                pk = data.get('key')
                asset = MaintenanceAsset.objects.get(pk=pk)
                next_pk = 0
                prev = 0
                next_query = MaintenanceAsset.objects.filter(pk__gt=pk)
                if next_query.count() > 0:
                    next_pk = MaintenanceAsset.objects.filter(pk__gt=pk).order_by('-pk').last().pk

                if MaintenanceAsset.objects.filter(pk__lt=pk).count() > 0:
                    prev = MaintenanceAsset.objects.filter(pk__lt=pk).order_by('pk').last().pk

                obj = {
                    'sku': asset.sku,
                    'group': {
                        'name': asset.group.name,
                        'pk': asset.group.pk
                    },
                    'owner': asset.owner.get_full_name(),
                    'created_on': asset.created_on,
                    'image': asset.image.url,
                    'subgroup': {
                        'pk': asset.subgroup.pk,
                        'name': asset.subgroup.name
                    },
                    'is_active': asset.is_active,
                    'brand': asset.brand,
                    'origin': asset.origin,
                    'color': asset.color,
                    'location': {
                        'code': asset.location.code,
                        'descr': asset.location.descr
                    },
                    'next': next_pk,
                    'previous': prev,
                    'barcode': asset.sku,
                    'name': asset.name
                }

                arr = obj
                print(arr)

                success_response['message'] = arr

            elif module == 'work_order':
                key = data.get('key','*')
                if key == '*':
                    wos = WorkOrder.objects.all()
                else:
                    wos = WorkOrder.objects.filter(pk=key)

                for wo in wos:
                    arr.append({
                        'pk':wo.pk,
                        'wr':wo.wr(),
                        'technician':wo.tech(),
                        'analyse':wo.analyse,
                        'meta':wo.metah(),
                        'is_open':wo.is_open,
                        'entry_no':wo.wo_no,
                        'owner':wo.owner.get_full_name(),
                        'materials':wo.materials(),
                        'attachments':wo.attachments(),
                        'transactions':wo.transactions(),
                        'next':wo.next_entry(),
                        'prev':wo.previous_entry()
                    })

                success_response['message'] = arr

            elif module == 'generate_wo_material_request':
                from fpdf import FPDF
                class PDF(FPDF):
                    def header(self):
                        self.set_font('Arial', 'B', 12)
                        self.cell(0, 10, 'Material Request', 0, 1, 'C')

                    def footer(self):
                        self.set_y(-15)
                        self.set_font('Arial', 'I', 8)
                        self.cell(0, 10, 'Page %s' % self.page_no(), 0, 0, 'C')
                
                pdf = PDF('L',"mm",'A4')
                pdf.add_page()
                
                

                
                
                wo_no = data.get('wo')
                wo = WorkOrder.objects.get(wo_no=wo_no)
                mypk = data.get('mypk')
                owner = User.objects.get(pk=mypk)
                myname = owner.get_full_name()
                wr = wo.wr()['record']
                materials = wo.materials()

                # work order
                pdf.set_font('Arial', 'B', 10)
                pdf.cell(40,5,"Work Order : ",0,0,'L')
                pdf.set_font('Arial', '', 10)
                pdf.cell(40,5,wo.wo_no,0,1,'L')
                
                # location
                pdf.set_font('Arial', 'B', 10)
                pdf.cell(40,5,"LOCATION : ",0,0,'L')
                pdf.set_font('Arial', '', 10)
                pdf.cell(40,5,wr['location'],0,1,'L')
                
                # asset
                asset = wo.wr()['asset']
                pdf.set_font('Arial', 'B', 10)
                pdf.cell(40,5,"ASSET NAME : ",0,0,'L')
                pdf.set_font('Arial', '', 10)
                pdf.cell(100,5,f"{asset['name']}",0,1,'L')
                pdf.set_font('Arial', 'B', 10)
                pdf.cell(40,5,"ASSET SKU : ",0,0,'L')
                pdf.set_font('Arial', '', 10)
                pdf.cell(100,5,f"{asset['sku']}",0,1,'L')
                # issue
                pdf.set_font('Arial', 'B', 10)
                pdf.cell(40,5,"ISS.TITLE : ",0,0,'L')
                pdf.set_font('Arial', '', 10)
                pdf.cell(100,5,f"{wr['title']}",0,1,'L')

                # issue details
                pdf.set_font('Arial', 'B', 10)
                pdf.cell(40,5,"ISS.Description : ",0,1,'L')
                pdf.set_font('Arial', '', 10)
                pdf.multi_cell(250,5,wr['description'],0,'L',False,False)
                pdf.ln(10)
                
                if materials['available']:
                    pdf.set_font('Arial', 'B', 16)
                    pdf.cell(50,10,"NAME",1,0,'L',False)
                    pdf.cell(100,10,"DESCRIPTION",1,0,'L',False)
                    pdf.cell(40,10,"PRICE",1,0,'L',False)
                    pdf.cell(40,10,"QUANTITY",1,0,'L',False)
                    pdf.cell(40,10,"TOTAL",1,1,'L',False)

                    
                    mats = materials['materials']
                    pdf.set_font('Arial', '', 10)
                    total = 0
                    for material in mats:
                        total += material['total']
                        pdf.cell(50,10,material['name'],1,0,'L',False)
                        pdf.cell(100,10,material['description'],1,0,'L',False)
                        pdf.cell(40,10,str(material['unit_price']),1,0,'L',False)
                        pdf.cell(40,10,str(material['quantity']),1,0,'L',False)
                        pdf.cell(40,10,str(material['total']),1,1,'L',False)

                    pdf.ln(10)
                    pdf.cell(50,10,f"{myname}",1,0,'C')
                    pdf.cell(170,10,f"TOTAL : {total}",0,0,'C')
                    pdf.cell(50,10,f"",1,1,'C')

                    pdf.cell(50,5,f"Prepared By",0,0,'C')
                    pdf.cell(170,5,"")
                    pdf.cell(50,5,f"Approved By.",0,1,'C')

                    pdf.output('static/general/tmp/pdf.pdf')
                    success_response['message'] = 'static/general/tmp/pdf.pdf'

                else:
                    success_response['status_code'] = 404
                    success_response['message'] = "No materials to request for"

            elif module == 'gen_wo_report':
                wo_no = data.get('wo_no')
                wo = WorkOrder.objects.get(wo_no = wo_no)
                materials = wo.materials()
                attachments = wo.attachments()
                wr = wo.w_request
                wr_no = wr.entry_no
                title = wr.title
                reporter = wr.owner.get_full_name()
                date_reported = wr.date
                technician = wo.technician.get_full_name()
                issue = wr.description
                analysis = wo.analyse
                closed_by = wo.closed_by.get_full_name()
                transactions = WordOrderTransactions.objects.filter(wo=wo)

                from fpdf import FPDF
                class PDF(FPDF):
                    def header(self):
                        self.set_font('Arial', 'B', 20)
                        self.cell(0, 10, 'Work Order Report', 0, 1, 'C')

                    def footer(self):
                        self.set_y(-15)
                        self.set_font('Arial', 'I', 8)
                        self.cell(0, 10, 'Page %s' % self.page_no(), 0, 0, 'C')
                
                pdf = PDF('P',"mm",'A4')
                pdf.add_page()
                pdf.set_font('Arial', 'B', 12)
                pdf.cell(95,10,"WORK ORDER DETAILS",0,0)
                pdf.cell(95,10,"WORK REQUEST DETAILS",0,1,'R')

                pdf.set_font('Arial', '', 10)
                pdf.cell(95,5, f"WO N0.: {wo_no}",0,0)
                pdf.cell(95,5, f"WR N0.: {wr_no}",0,1,'R')

                pdf.set_font('Arial', '', 10)
                pdf.cell(95,5, f"Technician : {technician}",0,0)
                pdf.cell(95,5, f"Reported By: {reporter}",0,1,'R')

                pdf.cell(95,5, f"Closed By : {closed_by}",0,0)
                pdf.cell(95,5, f"Date Reported: {date_reported}",0,1,'R')
                pdf.ln(5)

                # issue
                pdf.set_font('Arial', 'B', 20)
                pdf.cell(0, 10, 'ISSUE', 0, 1, 'L')

                
                pdf.set_font('Arial', 'B', 12)
                pdf.cell(100,5,f"Title",0,1)
                pdf.set_font('Arial', '', 10)
                pdf.multi_cell(199,5,f" {title}",0,1)
                pdf.ln(5)

                pdf.set_font('Arial', 'B', 12)
                pdf.cell(100,5,f"Description",0,1)
                pdf.set_font('Arial', '', 10)
                pdf.multi_cell(199,5,f"{issue}",0,'L')
                pdf.ln(5)

                pdf.set_font('Arial', 'B', 12)
                pdf.cell(100,5,f"Analysis",0,1)
                pdf.set_font('Arial', '', 10)
                pdf.multi_cell(199,5,f"{analysis}",0,'L')
                pdf.ln(5)

                pdf.set_font('Arial', 'B', 20)
                pdf.cell(0, 10, 'Transactions', 0, 1, 'L')
                for transaction in transactions:
                    pdf.set_font('Arial', 'B', 12)
                    pdf.cell(100,5,f"{transaction.title}",0,1)
                    pdf.set_font('Arial', '', 10)
                    pdf.multi_cell(199,5,f"{transaction.description}",0,'L')
                    pdf.ln(5)
                
                



                pdf.output('static/general/tmp/wo_report.pdf')
                success_response['message'] = 'static/general/tmp/wo_report.pdf'

                



            else:
                success_response['status_code'] = 404
                success_response['message'] = "Unknown View Module"

        elif method == 'PATCH':
            if module == 'close':
                maintenance = data.get('maintenance')
                mypk = data.get('mypk')
                message = data.get('message')
                main = Maintenance.objects.get(pk=maintenance)

                owner = User.objects.get(pk=mypk)
                history = MaintenanceHistory(title='CLOSING', description=message, owner=owner, maintenance=main)

                main.is_open = 2
                history.save()
                main.save()

                success_response['message'] = "Work Order Closed"

            elif module == 'close_wo':
                wo_no = data.get('wo_no')
                clossing_message = data.get('clossing_message')
                mypk = data.get('mypk')
                owner = User.objects.get(pk=mypk)

                wo = WorkOrder.objects.get(wo_no=wo_no)
                WordOrderTransactions(wo=wo,title='Closing',description=clossing_message,owner=owner).save()
                wo.is_open = False
                wo.closed_by = owner
                wo.save()

                success_response['message'] = "Record Closed"

        response = success_response

    except Exception as e:
        error_type, error_instance, traceback = sys.exc_info()
        tb_path = traceback.tb_frame.f_code.co_filename
        line_number = traceback.tb_lineno
        response["status_code"] = 500
        response[
            "message"] = f"An error of type {error_type} occurred on line {line_number} in file {tb_path}. Details: {e}"

    return JsonResponse(response, safe=False)
