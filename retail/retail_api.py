import json
import sys
from decimal import Decimal

import pyodbc
from django.contrib.auth.models import User
from django.db.models import Q, Sum
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from fpdf import FPDF
from openpyxl.chart import PieChart, Reference
from sympy import Product

from admin_panel.anton import format_currency
from admin_panel.models import Emails, Locations
from ocean.settings import RET_DB_HOST, RET_DB_USER, RET_DB_PASS, RET_DB_NAME
from retail.db import ret_cursor, get_stock
from retail.models import BoltItems, BoltGroups, ProductSupplier, ProductGroup, ProductSubGroup, Products, Stock, \
    RecipeGroup, RecipeProduct, Recipe, StockHd, StockMonitor, RawStock
from retail.prodMast import ProdMaster
from retail.retail_tools import create_recipe_card


@csrf_exempt
def interface(request):
    response = {
        'status_code': 0,
        'message': ""
    }

    success_response = {
        'status_code': 200,
        'message': "Procedure Completed Successfully"
    }

    # get method
    method = request.method

    try:

        body = json.loads(request.body)
        module = body.get('module')
        data = body.get('data')

        if method == 'PUT':  # add bolt product
            if module == 'bolt_item':
                group = data.get('group')
                barcode = data.get('barcode')
                item_des = data.get('item_des')
                price = data.get('price')
                stock_nia = data.get('stock_nia')
                stock_spintex = data.get('stock_spintex')
                stock_osu = data.get('stock_osu')

                cursor = ret_cursor()
                qqq = f"SELECT retail1 from prod_mast where barcode = '{barcode}'"
                print(qqq)
                cursor.execute(qqq)
                prod = cursor.fetchone()

                if prod is None:
                    inv_price = 0.00
                else:
                    inv_price = prod[0]

                BoltItems(barcode=barcode, item_des=item_des, price=price, stock_osu=stock_osu,
                          stock_spintex=stock_spintex, stock_nia=stock_nia, group_id=group, inv_price=inv_price).save()
                success_response['message'] = "Product Added"

            elif module == 'bolt_group':  # add bolt group
                name = data.get('name')
                try:
                    BoltGroups(name=name).save()
                except Exception as e:
                    pass
                success_response['message'] = BoltGroups.objects.get(name=name).pk

            elif module == 'sync_retail_suppliers':
                query = "select supp_code as 'code',supp_name as 'name',contact as 'person',phone1 as 'phone',address1 as 'email',address2 as 'city',country_id as 'country' from supplier"
                cursor = ret_cursor()
                cursor.execute(query)
                counts = 0
                not_counted = 0
                for supplier in cursor.fetchall():
                    code = supplier[0].strip()
                    name = supplier[1].strip()
                    person = supplier[2].strip()
                    phone = '' if supplier[3] is None else supplier[3].strip()
                    email = '' if supplier[4] is None else supplier[4].strip()
                    city = '' if supplier[5] is None else supplier[5].strip()
                    country = '' if supplier[6] is None else supplier[6].strip()

                    try:
                        supp, create = ProductSupplier.objects.get_or_create(code=code, name=name, phone=phone,
                                                                             person=person,
                                                                             email=email, city=city, country=country)
                        counts = counts + 1
                    except Exception as e:
                        not_counted = not_counted + 1

                success_response['message'] = f"{counts} / {counts + not_counted} suppliers synced"
                response = success_response

            elif module == 'sync_retail_groups':
                query = f"select group_code,group_des from group_mast"
                cursor = ret_cursor()
                cursor.execute(query)
                count = 0
                not_counted = 0

                for group in cursor.fetchall():
                    code, name = group[0].strip(), group[1].strip()
                    try:
                        get, add = ProductGroup.objects.get_or_create(code=code, name=name)
                        count = count + 1
                    except Exception as e:
                        not_counted = not_counted + 1

                success_response['message'] = f"{count} / {count + not_counted} groups synced"
                response = success_response

            elif module == 'sync_retail_sub_groups':
                cursor = ret_cursor()
                query = f"SELECT group_code,(select group_des from group_mast where group_code = sub_group.group_code) as 'group', sub_group,sub_group_des from sub_group"
                cursor.execute(query)
                count = 0
                not_counted = 0
                errors = []

                for sub_group in cursor.fetchall():

                    group_code, group_des, sub_group_code, sub_group_des = sub_group[0], sub_group[1].strip(), \
                        sub_group[2].strip(), sub_group[3].strip()
                    group, add = ProductGroup.objects.get_or_create(code=group_code, name=group_des)

                    if add:
                        group = ProductGroup.objects.get(code=group_code, name=group_des)

                    try:
                        sub, create = ProductSubGroup.objects.get_or_create(group=group, code=sub_group_code,
                                                                            name=sub_group_des)
                        count = count + 1
                    except Exception as e:
                        errors.append(e)
                        not_counted = not_counted + 1

                success_response['message'] = f"{count} / {count + not_counted} sub groups synced with errors {errors}"
                response = success_response

            elif module == 'sync_retail_products':
                cursor = ret_cursor()
                query = "SELECT item_code, barcode, item_des, (SELECT group_des FROM group_mast WHERE group_mast.group_code = prod_mast.group_code) AS 'group', (SELECT sub_group_des FROM sub_group WHERE sub_group.group_code = prod_mast.group_code AND sub_group.sub_group = prod_mast.sub_group) AS 'sub_group', (SELECT supp_name FROM supplier WHERE supplier.supp_code = prod_mast.supp_code) AS 'supplier', retail1 FROM prod_mast WHERE item_type != 0"
                cursor.execute(query)
                saved = 0
                not_synced = 0
                error = []
                for product in cursor.fetchall():
                    code = product[0]
                    barcode = str(product[1]).strip()
                    item_des = product[2].strip()
                    group = product[3]
                    print(product[4])
                    try:
                        sub_group = product[4].strip()
                    except Exception as e:
                        sub_group = product[4]

                    supplier = product[5]
                    retail1 = product[6]

                    # add to products

                    if ProductSubGroup.objects.filter(name=sub_group).count() == 1:
                        subgroup = ProductSubGroup.objects.get(name=sub_group)
                        # delete product
                        if Products.objects.filter(code=code).exists():
                            # update
                            prod = Products.objects.get(code=code)
                            prod.barcode = barcode
                            prod.name = item_des
                            prod.price = retail1
                            prod.subgroup = subgroup
                            prod.save()
                        else:
                            # save new
                            Products.objects.get_or_create(subgroup=subgroup, name=item_des, barcode=barcode,
                                                           code=code, price=retail1)
                        saved = saved + 1
                    else:
                        not_synced = not_synced + 1
                        error.append(f"{barcode} - {item_des} # sub group {sub_group} / does not exist")

                success_response['message'] = f"{saved} / {saved + not_synced} products synced"
                response = success_response

            elif module == 'update_stock':
                products = Products.objects.all()
                for product in products:
                    barcode = product.barcode.strip()
                    item_code = product.code

                    stock = get_stock(item_code)
                    print(product.name, stock)
                    spintex = stock.get('spintex')
                    nia = stock.get('nia')
                    osu = stock.get('osu')
                    warehouse = stock.get('warehouse')
                    kitchen = stock.get('kitchen')

                    if Stock.objects.filter(product=product, location='001').exists():
                        sp_stock = Stock.objects.get(product=product, location='001')
                        sp_stock.quantity = spintex
                        sp_stock.save()
                    else:
                        Stock.objects.create(product=product, quantity=spintex, location='001')

                    if Stock.objects.filter(product=product, location='202').exists():
                        stock202 = Stock.objects.get(product=product, location='202')
                        stock202.quantity = spintex
                        stock202.save()
                    else:
                        Stock.objects.create(product=product, quantity=nia, location='202')

                    if Stock.objects.filter(product=product, location='205').exists():
                        stock205 = Stock.objects.get(product=product, location='205')
                        stock205.quantity = spintex
                        stock205.save()
                    else:
                        Stock.objects.create(product=product, quantity=osu, location='205')

                    if Stock.objects.filter(product=product, location='999').exists():
                        stock999 = Stock.objects.get(product=product, location='999')
                        stock999.quantity = spintex
                        stock999.save()
                    else:
                        Stock.objects.create(product=product, quantity=warehouse, location='999')

                    if Stock.objects.filter(product=product, location='201').exists():
                        stock201 = Stock.objects.get(product=product, location='201')
                        stock201.quantity = spintex
                        stock201.save()
                    else:
                        Stock.objects.create(product=product, quantity=kitchen, location='201')
                success_response['message'] = "Stock Updated"

            elif module == 'recipe_group':
                name = data.get('name')
                us = data.get('mypk')
                owner = User.objects.get(pk=us)
                RecipeGroup(name=name, owner=owner).save()

                success_response['message'] = "Group Added!!"
            elif module == 'recipe_product':
                gk = data.get('gk')
                group = RecipeGroup.objects.get(pk=gk)
                us = data.get('mypk')
                owner = User.objects.get(pk=us)
                name = data.get('name')
                barcode = data.get('barcode')
                si_unit = data.get('si_unit')

                RecipeProduct(group=group, name=name, barcode=barcode, si_unit=si_unit, owner=owner).save()
                recipe = RecipeProduct.objects.get(group=group, name=name, barcode=barcode, si_unit=si_unit,
                                                   owner=owner)
                success_response['message'] = recipe.pk

            elif module == 'recipe_items':
                pro_key = data.get('product')
                product = RecipeProduct.objects.get(pk=pro_key)

                us = data.get('mypk')
                owner = User.objects.get(pk=us)
                items = data.get('items')

                # delete all recipes
                Recipe.objects.filter(product=product).delete()

                for item in items:
                    print(item)
                    name = item['name']
                    qty = item['qty']
                    si_unit = item['si_unit']

                    Recipe(product=product, name=name, owner=owner, quantity=qty, si_unit=si_unit).save()

                success_response['message'] = "Recipe Saved Successfully"

            # retrieve frozen stock
            elif module == 'retrieve_frozen_stock':
                mypk = data.get('mypk')
                entry = data.get('')

                # db query
                cursor = ret_cursor()
                query = f"select loc_id,ref_no,ld_date,remarks,grp,grp_from,grp_to from stock_keep_hd where ref_no = '{entry}'"
                cursor.execute(query)
                hd = cursor.fetchone()

                if hd:
                    loc_id, ref_no, date_kept, remarks, is_group, st_grp, end_grp = hd
                    owner = User.objects.get(pk=mypk)
                    loc = Locations.objects.get(loc_id=loc_id)

                    StockHd(
                        loc=loc_id, ref_no=ref_no, date_kept=date_kept,
                        remarks=remarks, is_group=is_group, st_grp=st_grp, end_grp=end_grp, owner=owner
                    ).save()

                    stock_hd = StockHd.objects.get(
                        loc=loc_id, ref_no=ref_no, date_kept=date_kept,
                        remarks=remarks, is_group=is_group, st_grp=st_grp, end_grp=end_grp, owner=owner
                    )

                    # save transactions

                    success_response['status_code'] = 200
                    success_response['message'] = loc_id
                else:
                    success_response['status_code'] = 404
                    success_response['message'] = f"No stock keep for entry number {entry}"

            elif module == 'sync_locations':
                cursor = ret_cursor()
                cursor.execute("SELECT br_code,RTRIM(br_name) from branch")

                for br in cursor.fetchall():
                    code, name = br
                    if Locations.objects.filter(code=code).count() == 0:
                        # create location
                        Locations(code=code, descr=name, owner_id=1).save()

            elif module == 'bulk_monitor':
                barcodes = data.get('barcodes')
                enable = data.get('enable')
                yes = 0
                no = 0
                print(enable)
                for barcode in barcodes:

                    # get product
                    if Products.objects.filter(barcode=barcode).count() == 1:
                        product = Products.objects.get(barcode=barcode)
                        product.stock_monitor = enable
                        product.save()
                        yes += 1
                    else:
                        no += 1

                success_response['message'] = f"Bulk Monitor Flag set to {enable}  for ({yes}/{yes+no}) products"


            else:
                success_response['status_code'] = 404
                success_response['message'] = f"no {method} method with module called {module}"

        elif method == 'VIEW':
            arr = []
            if module == 'location_master':
                locations = Locations.objects.all()
                for location in locations:
                    arr.append(location.obj())

                success_response['message'] = arr
                response = success_response
            elif module == 'bolt_products':
                pk = data.get('key') or '*'

                if pk == '*':

                    items = BoltItems.objects.all()
                else:
                    items = BoltItems.objects.filter(Q(pk=pk) | Q(barcode=pk))

                item_list = []
                for item in items:
                    item_list.append({
                        'pk': item.pk,
                        'barcode': item.barcode,
                        'item_des': item.item_des,
                        'price': item.price,
                        'price_diff': item.price_diff,
                        'stock_nia': item.stock_nia,
                        'stock_spintex': item.stock_spintex,
                        'stock_osu': item.stock_osu,
                        'group': {
                            'pk': item.group.pk,
                            'name': item.group.name
                        }
                    })

                success_response['message'] = item_list

            elif module == 'bolt_group':
                pk = data.get('key') or '*'
                grps = []
                if pk == '*':
                    groups = BoltGroups.objects.all()
                else:
                    groups = BoltGroups.objects.filter(pk=pk)

                for group in groups:
                    grps.append({
                        'pk': group.pk,
                        'name': group.name
                    })

                success_response['message'] = grps

            elif module == 'price_change':
                send = data.get('send_mail') or 'no'
                rate_inc = data.get('rate_at', 20)
                cursor = ret_cursor()
                import openpyxl
                worksheet = openpyxl.Workbook()
                sheet = worksheet.active
                sheet['A1'] = "SKU"
                sheet['B1'] = "NAME"
                sheet['C1'] = "PRICE"
                sheet_row = 2

                spintex_stock_book = openpyxl.Workbook()
                spintex_stock_sheet = spintex_stock_book.active

                spintex_stock_sheet['A1'] = "SKU"
                spintex_stock_sheet['B1'] = "NAME"
                spintex_stock_sheet['C1'] = "STOCK STATUS"
                spintex_stock_row = 2

                osu_stock_book = openpyxl.Workbook()
                osu_stock_sheet = osu_stock_book.active

                osu_stock_sheet['A1'] = "SKU"
                osu_stock_sheet['B1'] = "NAME"
                osu_stock_sheet['C1'] = "STOCK STATUS"
                osu_stock_row = 2

                nia_stock_book = openpyxl.Workbook()
                nia_stock_sheet = nia_stock_book.active
                nia_stock_sheet['A1'] = "SKU"
                nia_stock_sheet['B1'] = "NAME"
                nia_stock_sheet['C1'] = "STOCK STATUS"
                nia_stock_row = 2

                for item in BoltItems.objects.all():
                    barcode = item.barcode.replace('.0', '')

                    # get product detail

                    product = cursor.execute(
                        f"SELECT retail1,item_code,item_type from prod_mast where barcode = '{barcode}'").fetchone()
                    if product is not None:
                        price, item_code, item_type = product
                        if price != item.price:
                            # print("CHANGE")
                            sheet[f"A{sheet_row}"] = barcode
                            sheet[f"B{sheet_row}"] = item.item_des
                            sheet[f"C{sheet_row}"] = Decimal(product[0]) / Decimal((1 - rate_inc / 100))
                            sheet_row += 1
                            item.price = price

                        # check stock and discontinues
                        stock = get_stock(item_code)
                        spintex = stock['spintex']
                        nia = stock['nia']
                        osu = stock['osu']

                        stock_spintex = 1 if spintex > 0 or item_type == '2' else 0
                        stock_osu = 1 if osu > 0 or item_type == '2' else 0
                        stock_nia = 1 if nia > 0 or item_type == '2' else 0

                        if stock_spintex != item.stock_spintex:
                            # change stock standing

                            spintex_stock_sheet[f"A{spintex_stock_row}"] = barcode
                            spintex_stock_sheet[f"B{spintex_stock_row}"] = item.item_des
                            spintex_stock_sheet[f"C{spintex_stock_row}"] = stock_spintex
                            item.stock_spintex = spintex

                            spintex_stock_row += 1

                        if stock_osu != item.stock_osu:
                            # change stock standing

                            osu_stock_sheet[f"A{osu_stock_row}"] = barcode
                            osu_stock_sheet[f"B{osu_stock_row}"] = item.item_des
                            osu_stock_sheet[f"C{osu_stock_row}"] = stock_osu
                            item.stock_osu = stock_osu

                            osu_stock_row += 1

                        if stock_nia != item.stock_nia:
                            # change stock standing

                            nia_stock_sheet[f"A{nia_stock_row}"] = barcode
                            nia_stock_sheet[f"B{nia_stock_row}"] = item.item_des
                            nia_stock_sheet[f"C{nia_stock_row}"] = stock_nia
                            item.stock_nia = stock_nia

                            nia_stock_row += 1

                price_count = sheet_row - 2
                nia_count = nia_stock_row - 2
                spintex_count = spintex_stock_row - 2
                osu_count = osu_stock_row - 2

                tr = ""
                if price_count > 0:
                    tr += f"<tr><td><strong>PRICE CHANGE</strong></td><td>{price_count}</td></tr>"
                if spintex_count > 0:
                    tr += f"<tr><td><strong>SPINTEX STOCK</strong></td><td>{spintex_count}</td></tr>"
                if nia_count > 0:
                    tr += f"<tr><td><strong>NIA</strong></td><td>{nia_count}</td></tr>"
                if osu_count > 0:
                    tr += f"<tr><td><strong>OSU STOCK</strong></td><td>{osu_count}</td></tr>"

                html = f"<table>{tr}</table>"

                print(html)

                from datetime import datetime
                current_datetime = datetime.now()
                formatted_datetime = current_datetime.strftime("%Y_%m_%d_%H_%M_%S")

                price_change_file = f"static/general/bolt_price_changes/price_changes_as_of_{formatted_datetime}.xlsx"
                spintex_stock_change_file = f"static/general/bolt_price_changes/spintex_stock_change_as_of_{formatted_datetime}.xlsx"
                osu_stock_change_file = f"static/general/bolt_price_changes/osu_stock_change_as_of_{formatted_datetime}.xlsx"
                nia_stock_change_file = f"static/general/bolt_price_changes/nia_stock_change_as_of_{formatted_datetime}.xlsx"

                worksheet.save(price_change_file)
                spintex_stock_book.save(spintex_stock_change_file)
                osu_stock_book.save(osu_stock_change_file)
                nia_stock_book.save(nia_stock_change_file)

                if send == 'yes':
                    if price_count > 0 or spintex_count > 0 or nia_count > 0 or osu_count > 0:
                        Emails(sent_to="solomon@snedaghana.com", subject='BOLT UPDATE', body=html,
                               attachments=f"{price_change_file},{spintex_stock_change_file},"
                                           f"{osu_stock_change_file},{nia_stock_change_file}").save()

                success_response['message'] = {
                    'price_change': {
                        'file': price_change_file,
                        'count': sheet_row - 2
                    },
                    'spintex_stock_change': {
                        'file': spintex_stock_change_file,
                        'count': spintex_stock_row - 2
                    },
                    'osu_stock_change_file': {
                        'file': osu_stock_change_file,
                        'count': osu_stock_row - 2
                    },
                    'nia_stock_change_file': {
                        'file': nia_stock_change_file,
                        'count': nia_stock_row - 2
                    }
                }

            elif module == 'export_items':
                key = data.get('key')
                format = data.get('format')
                if key == '*':
                    items = BoltItems.objects.all()
                    file_name = f"static/general/bolt/bolt-items.xlsx"
                else:
                    items = BoltItems.objects.filter(group_id=key)
                    group = BoltGroups.objects.get(pk=key)
                    file_name = f"static/general/bolt/{group.name}.xlsx"

                import openpyxl
                workbook = openpyxl.Workbook()
                sheet = workbook.active

                if format == 'excel':
                    sheet['A1'] = "CATEGORY"
                    sheet['B1'] = "BARCODE"
                    sheet['C1'] = "NAME"
                    sheet['D1'] = "PRICE"
                    sheet['E1'] = "SPINTEX"
                    sheet['F1'] = "NIA"
                    sheet['G1'] = "OSU"
                    row = 2

                    for item in items:
                        sheet[f'A{row}'] = item.group.name
                        sheet[f'B{row}'] = item.barcode
                        sheet[f'C{row}'] = item.item_des
                        sheet[f'D{row}'] = item.price
                        sheet[f'E{row}'] = item.stock_spintex
                        sheet[f'F{row}'] = item.stock_nia
                        sheet[f'G{row}'] = item.stock_osu
                        row += 1

                    workbook.save(file_name)
                    success_response['message'] = file_name

            elif module == 'slow_moving_items':
                loc = data.get('loc') or ''
                days = data.get('days')
                export = data.get('export')
                if export == 'excel':
                    import openpyxl
                    book = openpyxl.Workbook()
                    sheet = book.active
                    sheet['A1'] = "SLOW MOVING ITEMS"
                    sheet['A2'] = "BARCODE"
                    sheet["B3"] = "NAME"
                    sheet['C3'] = "AVAILABLE QUANTITY"
                    sheet['D2'] = "SOLD QUANTITY"
                    sh_row = 3

                query = f"exec dbo.Sp_slow_moving_rept N'%',N'%',N'%',N'',N'Zade',N'AA001',N'SOO216',N'',N'YWS195','2023-11-13 00:00:00',N'30',N'%',N'%',N'%','2023-08-15 00:00:00','2023-11-13 00:00:00',N'ALL'"
                print(query)
                server = f"{RET_DB_HOST}"
                database = RET_DB_NAME
                username = RET_DB_USER
                password = RET_DB_PASS
                driver = '{ODBC Driver 17 for SQL Server}'  # Change this to the driver you're using
                connection_string = f"DRIVER={driver};SERVER={server};DATABASE={database};UID={username};PWD={password}"
                connection = pyodbc.connect(connection_string)

                cursor = ret_cursor()
                cursor.execute("""
                EXEC dbo.Sp_slow_moving_rept ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
                """,
                               '001', '%', '%', '', 'ZWAN', 'AA001', 'SOO216', '', 'YWS195', '2024-06-22 00:00:00',
                               '30', '%', '%', '%', '2024-05-23 00:00:00', '2024-06-22 00:00:00', 'ALL')
                # Fetch all rows from the executed stored procedure
                rows = cursor.fetchall()
                print(rows)
                arr = []
                for row in rows:
                    print(row)
                    barcode = row[2]
                    item_code = row[0]
                    item_ref = row[1]
                    name = row[3]
                    av_stock = row[17]
                    sol_stock = row[21]
                    if export == 'json':
                        obj = {
                            'barcode': barcode,
                            'item_code': item_code,
                            'item_ref': item_ref,
                            'name': name,
                            'available': av_stock,
                            'sold': sol_stock
                        }
                    elif export == 'excel':
                        sheet[f"A{sh_row}"] = barcode
                        sheet[f"B{sh_row}"] = name
                        sheet[f"C{sh_row}"] = av_stock
                        sheet[f"D{sh_row}"] = sol_stock

                if export == 'json':
                    success_response['message'] = arr
                elif export == 'excel':
                    file_name = f"static/general/tmp/SLOW_MOVING_ITEMS.xlsx"
                    book.save(file_name)
                    success_response['message'] = file_name

                response = success_response

            elif module == 'retail_categories':
                import openpyxl
                document = data.get('doc')
                group_id = data.get('group_id') or '*'

                if group_id == '*':
                    groups = ProductGroup.objects.all()
                else:
                    groups = ProductGroup.objects.filter(pk=group_id)

                arr = []

                book = openpyxl.Workbook()
                sheet = book.active
                sheet.title = 'CATEGORIES'
                sheet['A1'] = 'CODE'
                sheet['B1'] = 'NAME'
                sheet['C1'] = 'SUBS'
                sheet_count = 2

                for group in groups:
                    if document == 'json':
                        arr.append({
                            'code': group.code,
                            'name': group.name,
                            'subs': group.subgroups().count()
                        })
                    elif document == 'excel':
                        sheet[f"A{sheet_count}"] = group.code
                        sheet[f"B{sheet_count}"] = group.name
                        sheet[f"C{sheet_count}"] = group.subgroups().count()
                        sheet_count += 1

                if document == 'json':
                    success_response['message'] = arr
                elif document == 'excel':
                    file_name = 'static/general/tmp/categories.xlsx'
                    book.save(file_name)
                    success_response['message'] = file_name

            elif module == 'retail_sub_categories':
                import openpyxl
                document = data.get('doc')
                sub_group_id = data.get('sub_group_id') or '*'
                if sub_group_id == '*':
                    sub_groups = ProductSubGroup.objects.all()
                else:
                    sub_groups = ProductSubGroup.objects.filter(pk=sub_group_id)

                arr = []
                book = openpyxl.Workbook()
                sheet = book.active
                sheet.title = "SUB CATEGORIES"
                sheet['A1'] = "GROUP"
                sheet['B1'] = "CODE"
                sheet['C1'] = "NAME"
                sheet['D1'] = "PRODUCTS"

                sheet_count = 2

                for sub_group in sub_groups:
                    group = sub_group.group
                    code = sub_group.code
                    name = sub_group.name
                    products = sub_group.products().count()

                    if document == 'json':
                        arr.append(
                            {'pk': sub_group.pk, 'group': group.name, 'code': code, 'name': name, 'products': products})
                    elif document == 'excel':
                        sheet[f"A{sheet_count}"] = group.name
                        sheet[f"B{sheet_count}"] = code
                        sheet[f"C{sheet_count}"] = name
                        sheet[f"D{sheet_count}"] = products

                if document == 'json':
                    success_response['message'] = arr
                elif document == 'excel':
                    file_name = 'static/general/tmp/sub_categories.xlsx'
                    book.save(file_name)
                    success_response['message'] = file_name

            elif module == 'retail_products':
                import openpyxl
                product = data.get('product') or '*'
                document = data.get('doc') or 'json'

                if product == '*':
                    products = Products.objects.all()
                else:
                    products = Products.objects.filter(pk=product)

                book = openpyxl.Workbook()
                sheet = book.active
                sheet.title = 'Products'

                sheet['A1'] = 'Group'
                sheet['B1'] = 'Sub Group'
                sheet['C1'] = 'Code'
                sheet['D1'] = 'Barcode'
                sheet['E1'] = 'Name'
                sheet['F1'] = 'Price'
                sheet['G1'] = 'IS_ON_BOLT'
                sheet_count = 2
                arr = []

                for product in products:
                    sub_group = product.subgroup
                    group = sub_group.group.name
                    sub_name = sub_group.name
                    code = product.code
                    barcode = product.barcode.strip()
                    name = product.name
                    price = product.price
                    pk = product.pk
                    bolt = product.is_on_bolt()

                    # check if on bolt

                    if document == 'json':
                        arr.append({
                            'pk': pk,
                            'group': group,
                            'sub_group': sub_name,
                            'code': code,
                            'barcode': barcode,
                            'name': name,
                            'price': price,
                            'is_on_bolt': bolt
                        })
                    elif document == 'excel':
                        sheet['A' + str(sheet_count)] = group
                        sheet['B' + str(sheet_count)] = sub_name
                        sheet['C' + str(sheet_count)] = code
                        sheet['D' + str(sheet_count)] = barcode
                        sheet['E' + str(sheet_count)] = name
                        sheet['F' + str(sheet_count)] = price
                        sheet['G' + str(sheet_count)] = bolt
                        sheet_count += 1

                if document == 'json':
                    success_response['message'] = arr
                elif document == 'excel':
                    file_name = 'static/general/tmp/products.xlsx'
                    book.save(file_name)
                    success_response['message'] = file_name

            elif module == 'recipe_group':
                key = data.get('key', '*')

                if key == '*':
                    groups = RecipeGroup.objects.all()
                else:
                    groups = RecipeGroup.objects.filter(pk=key)

                for group in groups:
                    prod = []
                    products = group.products()
                    p_count = products.count()
                    for product in products:
                        prod.append({
                            'pk': product.pk,
                            'name': product.name,
                            'barcode': product.barcode,
                            'si_unit': product.si_unit,
                            'recipe_items': product.recipe_items(),
                            'image': product.image.url or 'none'
                        })

                    arr.append({
                        'name': group.name,
                        'pk': group.pk,
                        'is_open': group.is_open,
                        'products': {
                            'counts': p_count,
                            'list': prod
                        }
                    })

                success_response['message'] = arr

            elif module == 'recipe_product':
                key = data.get('key', '*')
                search_type = data.get('s_type', 'normal')

                if search_type == 'by_name':
                    products = RecipeProduct.objects.filter(
                        name__icontains=key) if key != '*' else RecipeProduct.objects.all()
                else:
                    products = RecipeProduct.objects.filter(pk=key)

                if products.count() > 0:
                    for product in products:
                        # print(product)
                        rec_list = []
                        for r in product.recipe():
                            rec_list.append({
                                'pk': r.pk,
                                'name': r.name,
                                'quantity': r.quantity,
                                'si_unit': r.si_unit
                            })

                        arr.append({
                            'pk': product.pk,
                            'name': product.name,
                            'barcode': product.barcode,
                            'si_unit': product.si_unit,
                            'recipe_items': {
                                'count': product.recipe_items(),
                                'list': rec_list
                            },
                            'is_open': product.is_open,
                            'image': product.img_url(),
                            'next': product.next(),
                            'previous': product.prev()
                        })

                        success_response['message'] = arr
                else:
                    success_response['status_code'] = 404
                    success_response['message'] = "Product Not Found"
            elif module == 'recipe_item':
                key = data.get('key', '*')

                if key == '*':
                    recipe_item = Recipe.objects.all()
                else:
                    recipe_item = Recipe.objects.filter(product_id=key)

                for item in recipe_item:
                    arr.append({
                        'name': item.name,
                        'quantity': item.quantity,
                        'si_unit': item.si_unit
                    })

                success_response['message'] = arr

            elif module == 'export_recipe_group':
                group_pk = data.get('group')
                group = RecipeGroup.objects.get(pk=group_pk)
                file_name = f"static/general/tmp/{group.name}.xlsx"
                products = group.products()
                import openpyxl

                book = openpyxl.Workbook()

                for product in products:

                    name = product.name
                    sheet = book.create_sheet(title=name.replace('/', ''))
                    sheet['A1'] = 'NAME'
                    sheet['B1'] = "UNIT"
                    sheet['C1'] = "QUANTITY"

                    sheet_row = 2
                    l_qty = 0
                    recipes = product.recipe()
                    print(name)
                    for recipe in recipes:
                        r_name = recipe.name
                        si_unit = recipe.si_unit
                        quantity = recipe.quantity
                        sheet[f"A{sheet_row}"] = r_name
                        sheet[f"B{sheet_row}"] = si_unit
                        sheet[f"C{sheet_row}"] = quantity
                        sheet_row += 1
                        l_qty += Decimal(quantity)
                        print(r_name)

                    sheet[f"A{sheet_row}"] = 'TOTAL'
                    sheet[f"C{sheet_row}"] = l_qty

                    print()

                book.save(file_name)
                success_response['message'] = f'/{file_name}'

            elif module == 'stock_monitor':
                # get enabled stocks
                products = Products.objects.filter(stock_monitor=True)
                for product in products:
                    pk = product.pk
                    code = product.code
                    pm = ProdMaster()
                    stk = pm.get_stock(code)

                    nia = stk['nia']
                    osu = stk['osu']
                    spintex = stk['spintex']
                    warehouse = stk['warehouse']
                    vegetables = stk['vegetables']
                    kitchen = stk['kitchen']
                    total = stk['total']
                    sales = stk['sales']

                    # spintex
                    sp = False
                    if spintex >= 0:
                        sp = True

                    if StockMonitor.objects.filter(location='001', product=product).exists():
                        stp = StockMonitor.objects.get(location='001', product=product)
                        stp.stock_qty = spintex
                        stp.valid = sp
                        stp.save()
                    else:
                        StockMonitor(location='001', product=product, stock_qty=spintex, valid=sp).save()

                    ## NIA
                    ni = False
                    if nia >= 0:
                        ni = True

                    if StockMonitor.objects.filter(location='202', product=product).exists():
                        stp = StockMonitor.objects.get(location='202', product=product)
                        stp.stock_qty = nia
                        stp.valid = ni
                        stp.save()
                    else:
                        StockMonitor(location='202', product=product, stock_qty=nia, valid=ni).save()

                    os = False
                    if osu >= 0:
                        os = True

                    if StockMonitor.objects.filter(location='205', product=product).exists():
                        stp = StockMonitor.objects.get(location='205', product=product)
                        stp.stock_qty = os
                        stp.valid = sp
                        stp.save()
                    else:
                        StockMonitor(location='205', product=product, stock_qty=osu, valid=os).save()

                    # kitchen
                    kt = False
                    if kitchen >= 0:
                        kt = True

                    if StockMonitor.objects.filter(location='201', product=product).exists():
                        stp = StockMonitor.objects.get(location='201', product=product)
                        stp.stock_qty = kitchen
                        stp.valid = kt
                        stp.save()
                    else:
                        StockMonitor(location='201', product=product, stock_qty=kitchen, valid=kt).save()

                    # warehouse
                    wh = False
                    if warehouse >= 0:
                        wh = True

                    if StockMonitor.objects.filter(location='999', product=product).exists():
                        stp = StockMonitor.objects.get(location='999', product=product)
                        stp.stock_qty = warehouse
                        stp.valid = wh
                        stp.save()
                    else:
                        StockMonitor(location='999', product=product, stock_qty=warehouse, valid=wh).save()

            elif module == 'see_stock_monitor':
                arr = []
                filter = data.get('filter')
                loc_code = data.get('loc_code','*')
                print(filter)
                doc = data.get('doc', 'json')
                if doc == 'excel':
                    import openpyxl
                    book = openpyxl.Workbook()
                    sheet = book.active
                    sheet.title = "Stock Check"
                    sheet.append(['LOCATION', 'BARCODE', 'NAME', 'STOCK'])
                # if filter == 'not_accurate':
                #     all_m = StockMonitor.objects.filter(valid=False)
                # elif filter == 'accurate':
                #     all_m = StockMonitor.objects.filter(valid=True)
                # else:
                #     all_m = StockMonitor.objects.all()
                all_m = Products.objects.all()
                locations = Locations.objects.filter(type='retail')
                if loc_code != '*':
                    locations = locations.filter(code=loc_code)
                for m in all_m:
                    for location in locations:
                        code = location.code
                        loc_stock = \
                        RawStock.objects.filter(prod_id=m.code, loc_id=location.code).aggregate(sum=Sum('qty'))[
                            'sum'] or Decimal('0.00')
                        row = [location.descr, m.name, m.name, loc_stock]
                        mark = 0
                        if loc_stock >= mark:
                            print(f"{loc_stock} is positive for ",filter,location.code,m.code)
                        elif loc_stock < mark:
                            print(mark, "is negative for",filter,location.code,m.code)
                        else:
                            print(mark, "Is neither negative or positive",location.code,m.code)

                        if filter == 'negative' and loc_stock < mark:
                            print("YES NEGATIVE")

                            if doc == 'excel':
                                sheet.append(row)
                            arr.append({'loc_code': code, 'loc_name': location.descr, 'barcode': m.barcode,
                                        'item_name': m.name,
                                        'stock': loc_stock})

                        elif filter == 'positive' and loc_stock > mark:
                            arr.append({'loc_code': code, 'loc_name': location.descr, 'barcode': m.barcode,
                                        'item_name': m.name,
                                        'stock': loc_stock})
                            if doc == 'excel':
                                sheet.append(row)
                                print("YES POSITIVE")

                        elif filter == 'neutral' and loc_stock == mark:
                            arr.append({'loc_code': code, 'loc_name': location.descr, 'barcode': m.barcode,
                                        'item_name': m.name,
                                        'stock': loc_stock})
                            if doc == 'excel':
                                sheet.append(row)


                    # arr.append(m.stock())

                if doc == 'excel':
                    filename = 'static/general/tmp/stock.xlsx'
                    book.save(filename)
                    arr = filename
                success_response['message'] = arr
                response = success_response

            elif module == 'mr_check':
                mr_no = data.get('mr_no')
                doc_type = data.get('doc', "JSON")
                healthy = 0
                not_healthy = 0

                cursor = ret_cursor()
                cursor.execute(
                    f"SELECT entry_no, entry_date, remark, (SELECT RTRIM(user_name) FROM user_file WHERE user_id = "
                    f"request_hd.user_id) AS 'requested_by', (SELECT RTRIM(br_name) FROM branch WHERE br_code = request_hd.loc_id) "
                    f"AS 'loc_requested', (SELECT RTRIM(br_name) FROM branch WHERE br_code = request_hd.loc_from) "
                    f"AS 'loc_to_send',loc_id FROM request_hd where entry_no = '{mr_no}';")
                mr_hd = cursor.fetchone()
                if mr_hd is not None:
                    if doc_type == 'excel':
                        import openpyxl
                        book = openpyxl.Workbook()
                        sheet = book.active
                        sheet.title = "RECORDS"
                        sheet[f'A1'] = "ITEM CODE"
                        sheet[f"B1"] = "NAME"
                        sheet[f"C1"] = "REQ. QTY"
                        sheet[f'D1'] = "LAST TRAN. ENTRY"
                        sheet[f"E1"] = "LAST TRAN. DATE"
                        sheet[f'F1'] = "LAST TRAN. QTY"
                        sheet[f"G1"] = "SOLD BETWEEN"
                        sheet[f"H1"] = "SOLD PERENTAGE"
                        sheet[f'I1'] = "HEALTH"

                        sheet_row = 2

                    arr = []
                    entry_no, entry_date, remark, requested_by, loc_requested, loc_to_send, loc_id = mr_hd
                    header = {
                        "entry_no": entry_no.strip(),
                        "entry_date": entry_date,
                        'remark': remark.strip() if remark is not None else None,
                        'requested_by': requested_by,
                        'loc_requested': loc_requested,
                        'loc_to_send': loc_to_send,
                        'loc_id': loc_id
                    }

                    # connect to branch and get sold
                    loc = Locations.objects.get(code=loc_id)
                    pos_cursor = ret_cursor(loc.ip_address, '', loc.db, loc.db_user, loc.db_password)

                    # get transactions
                    trans_query = cursor.execute(
                        f"select item_code,RTRIM(item_des) as 'name',total_units from request_tran where entry_no = '{mr_no}'")
                    tr_arr = []
                    for tran in trans_query.fetchall():
                        item_code, item_name, request_qty = tran
                        cursor.execute(
                            f"select top(1) th.entry_no,th.entry_date,tr.total_units from tran_tr tr  right join tran_hd th on th.entry_no = tr.entry_no where tr.item_code = '{item_code}' order by th.entry_date desc")
                        last_tran_row = cursor.fetchone()

                        last_tr_entry, last_tran_date, last_tran_qty, sold_qty = ['none', 'none', 0, 0]

                        if last_tran_row is not None:
                            last_tr_entry, last_tran_date, last_tran_qty = last_tran_row

                            pos_cursor.execute(
                                f"SELECT sum(tran_qty) as 'sold_qty' FROM history_tran where bill_date between '{last_tran_date}' and '{entry_date}' and prod_id = '{item_code}'")
                            sold_resp = pos_cursor.fetchone()
                            if sold_resp is not None:
                                if sold_resp[0] is not None:
                                    sold_qty = sold_resp[0]

                        health = False
                        if sold_qty > Decimal(0.5) * last_tran_qty:
                            health = True
                            healthy += 1
                        else:
                            not_healthy += 1

                        percentage_sold = 0
                        if last_tran_qty != 0 and sold_qty != 0:  # Ensure we don't divide by zero
                            percentage_sold = (Decimal(sold_qty) / Decimal(last_tran_qty)) * 100
                        if doc_type == 'JSON':
                            tr_arr.append({
                                'item_code': tran[0],
                                'name': tran[1],
                                'request_qty': tran[2],
                                'last_tran_entry': last_tr_entry,
                                'last_tran_date': last_tran_date,
                                "last_tran_qty": last_tran_qty,
                                'sold_qty': sold_qty,
                                "percentage_sold": percentage_sold,
                                'health': health
                            })
                        elif doc_type == 'excel':
                            sheet[f'A{sheet_row}'] = tran[0]
                            sheet[f"B{sheet_row}"] = item_name
                            sheet[f"C{sheet_row}"] = tran[2]
                            sheet[f'D{sheet_row}'] = last_tr_entry
                            sheet[f"E{sheet_row}"] = last_tran_date
                            sheet[f'F{sheet_row}'] = last_tran_qty
                            sheet[f"G{sheet_row}"] = sold_qty
                            sheet[f"H{sheet_row}"] = percentage_sold
                            sheet[f'I{sheet_row}'] = health

                            sheet_row += 1

                    if doc_type == "JSON":
                        resp = {
                            'header': header,
                            'transactions': tr_arr
                        }

                        arr.append(resp)
                    elif doc_type == 'excel':
                        # Write "yes" and "no" to cells
                        sheet['K1'] = "HEALTHY"
                        sheet['L1'] = "UNHEALTHY"
                        sheet['K2'] = healthy  # Example value for "yes"
                        sheet['L2'] = not_healthy  # Example value for "no"

                        # Create a pie chart
                        chart = PieChart()
                        data = Reference(sheet, min_col=11, min_row=2, max_col=12,
                                         max_row=2)  # Range of data values (yes and no counts)
                        labels = Reference(sheet, min_col=11, min_row=1, max_col=12,
                                           max_row=1)  # Range of category labels (yes and no)
                        chart.add_data(data, titles_from_data=True)
                        # chart.title = "STOCK HEALTH"
                        chart.set_categories(labels)

                        # Set the title of the chart
                        chart.title = "Yes vs No"

                        # Add the chart to the worksheet
                        sheet.add_chart(chart, "K5")

                        # Set the dimensions of the chart
                        chart.width = 15  # Adjust width
                        chart.height = 10  # Adjust height

                        file_name = f"static/general/tmp/mr_check{mr_no}.xlsx"
                        book.save(file_name)
                        resp = file_name
                    else:
                        resp = f"Unknown Document To Render {doc_type}"
                    success_response['message'] = resp
                    response = success_response
                else:
                    success_response['message'] = f"no request with entry {mr_no}"
                    success_response['status_code'] = 404
                    response = success_response

            elif module == 'documents':
                start_date = data.get('start_date')
                end_date = data.get('end_date')
                print(start_date,end_date)
                query = f"""
                    DECLARE @start_date varchar(20) = '{start_date}'
                    DECLARE @end_date varchar(20) = '{end_date}'
                    
                    SELECT 'GRN' AS document, (
                        SELECT 
                            (SELECT COUNT(*) FROM grn_hd where grn_date between @start_date and @end_date) AS total_entries,
                            (SELECT COUNT(*) FROM grn_hd WHERE grn_date between @start_date and @end_date and valid = 1 AND posted = 1) AS posted,
                            (SELECT COUNT(*) FROM grn_hd WHERE grn_date between @start_date and @end_date and valid = 1 AND posted = 0) AS not_posted,
                            (SELECT COUNT(*) FROM grn_hd WHERE grn_date between @start_date and @end_date and valid = 0) AS deleted,
                            (SELECT SUM(inv_amt) FROM grn_hd where grn_date between @start_date and @end_date) AS total_value,
                            (SELECT SUM(inv_amt) FROM grn_hd where grn_date between @start_date and @end_date and valid = 1 and posted = 1) AS posted_value,
                            (SELECT SUM(inv_amt) FROM grn_hd where grn_date between @start_date and @end_date and valid = 1 and posted = 0) AS pending_value,
                            (SELECT SUM(inv_amt) FROM grn_hd where grn_date between @start_date and @end_date and valid = 0) AS deleted_value
                        FOR JSON PATH, WITHOUT_ARRAY_WRAPPER
                    ) AS Data
                    
                    
                    
                    -- INVOICE
                    UNION ALL
                    SELECT 'BACK OFFICE SALES' AS document, (
                        SELECT 
                            (SELECT COUNT(*) FROM inv_hd where entry_date between @start_date and @end_date) AS total_entries,
                            (SELECT COUNT(*) FROM inv_hd where entry_date between @start_date and @end_date and valid = 1 AND posted = 1) AS posted,
                            (SELECT COUNT(*) FROM inv_hd where entry_date between @start_date and @end_date and valid = 1 AND posted = 0) AS not_posted,
                            (SELECT COUNT(*) FROM inv_hd where entry_date between @start_date and @end_date and valid = 0) AS deleted,
                            (SELECT SUM(inv_amt) FROM inv_hd where entry_date between @start_date and @end_date) AS total_value,
                            (SELECT SUM(inv_amt) FROM inv_hd where entry_date between @start_date and @end_date and valid = 1 and posted = 1) AS posted_value,
                            (SELECT SUM(inv_amt) FROM inv_hd where entry_date between @start_date and @end_date and valid = 1 and posted = 0) AS pending_value,
                            (SELECT SUM(inv_amt) FROM inv_hd where entry_date between @start_date and @end_date and valid = 0) AS deleted_value
                        FOR JSON PATH, WITHOUT_ARRAY_WRAPPER
                    ) AS Data
                    
                    -- pos sales
                    UNION ALL
                    SELECT 'POS SALES' AS document, (
                        SELECT 
                            (SELECT COUNT(*) FROM pos_tran_hd where entry_date between @start_date and @end_date) AS total_entries,
                            (SELECT COUNT(*) FROM pos_tran_hd where entry_date between @start_date and @end_date and inv_valid = 1 AND inv_upd = 1) AS posted,
                            (SELECT COUNT(*) FROM pos_tran_hd where entry_date between @start_date and @end_date and inv_valid = 1 AND inv_upd = 0) AS not_posted,
                            (SELECT COUNT(*) FROM pos_tran_hd where entry_date between @start_date and @end_date and inv_valid = 0) AS deleted,
                            (SELECT SUM(inv_amt) FROM pos_tran_hd where entry_date between @start_date and @end_date) AS total_value,
                            (SELECT SUM(inv_amt) FROM pos_tran_hd where entry_date between @start_date and @end_date and inv_valid = 1 and inv_upd = 1) AS posted_value,
                            (SELECT SUM(inv_amt) FROM pos_tran_hd where entry_date between @start_date and @end_date and inv_valid = 1 and inv_upd = 0) AS pending_value,
                            (SELECT SUM(inv_amt) FROM pos_tran_hd where entry_date between @start_date and @end_date and inv_valid = 0) AS deleted_value
                        FOR JSON PATH, WITHOUT_ARRAY_WRAPPER
                    ) AS Data
                    
                    -- ADJUSTMENT
                    UNION ALL
                    SELECT 'ADJUSTMENT' AS document, (
                        SELECT 
                            (SELECT COUNT(*) FROM adj_hd where entry_date between @start_date and @end_date) AS total_entries,
                            (SELECT COUNT(*) FROM adj_hd where entry_date between @start_date and @end_date and valid = 1 AND posted = 1) AS posted,
                            (SELECT COUNT(*) FROM adj_hd where entry_date between @start_date and @end_date and valid = 1 AND posted = 0) AS not_posted,
                            (SELECT COUNT(*) FROM adj_hd where entry_date between @start_date and @end_date and valid = 0) AS deleted,
                            (SELECT SUM(tot_retail) FROM adj_hd where entry_date between @start_date and @end_date) AS total_value,
                            (SELECT SUM(tot_retail) FROM adj_hd where entry_date between @start_date and @end_date and valid = 1 and posted = 1) AS posted_value,
                            (SELECT SUM(tot_retail) FROM adj_hd where entry_date between @start_date and @end_date and valid = 1 and posted = 0) AS pending_value,
                            (SELECT SUM(tot_retail) FROM adj_hd where entry_date between @start_date and @end_date and valid = 0) AS deleted_value
                        FOR JSON PATH, WITHOUT_ARRAY_WRAPPER
                    ) AS Data

                    

                """
                arr = []
                cursor = ret_cursor()
                cursor.execute(query)
                for document in cursor.fetchall():
                    #print(document)
                    doc,daita = document
                    #print(daita)
                    obj = {}
                    json_data = json.loads(daita)
                    obj['document'] = doc
                    obj['total_entries'] = json_data["total_entries"]
                    obj['posted'] = json_data["posted"]
                    obj['not_posted'] = json_data["not_posted"]
                    obj['deleted'] = json_data["deleted"]
                    obj['total_value'] = format_currency(json_data.get('total_value'))
                    obj['posted_value'] = format_currency(json_data.get('posted_value'))
                    obj['pending_value'] = format_currency(json_data.get("pending_value"))


                    arr.append(obj)
                    # print(doc)
                    # print(json_data)
                    #print()

                success_response['message'] = arr
                response = success_response

            elif module == 'sales_graph_week':
                from django.utils.timezone import now
                loc_q = "SELECT br_code,RTRIM(br_name) from branch"
                cursor = ret_cursor()
                cursor.execute(loc_q)
                from datetime import timedelta
                # Calculate the start date (7 days ago)
                start_date = now().date() - timedelta(days=7)
                end_date = now().date()

                date_range = [start_date + timedelta(days=x) for x in range(0, (end_date - start_date).days + 1)]
                labels = []
                sales_data = {}
                for dr in date_range:
                    labels.append(dr)

                for location in cursor.fetchall():
                    name = location[1]
                    code = location[0]
                    # locations data set

                    # get sales
                    sales_arr = []
                    for d in labels:
                        s_q = f"SELECT inv_amt FROM pos_tran_hd where location_id = '{code}' and entry_date = '{d}'"
                        print(s_q)
                        cursor.execute(s_q)
                        row = cursor.fetchone()
                        if row is None:
                            sales = 0.00
                        else:
                            sales = row[0]
                        sales_arr.append(sales)

                    obj = {
                        "branch": name,
                        "sales": sales_arr
                    }

                    sales_data.update({name: sales_arr})
                arr = {
                    "labels": labels,
                    "dataset": sales_data
                }

                success_response['message'] = arr
                response = success_response

            else:
                raise Exception("No View Module")


        elif method == 'PATCH':
            if module == 'price_update':
                items = BoltItems.objects.all()

                for item in items:
                    item.price = item.inv_price
                    item.save()

            elif module == 'close_recipe':
                prod_key = data.get('prod_key')
                product = RecipeProduct.objects.get(pk=prod_key)
                items = product.recipe()

                if items.count() > 0:
                    # make recipe card
                    item_arr = []
                    for item in items:
                        item_arr.append([item.name, item.si_unit, item.quantity])

                    create_recipe_card(product.name.replace('/', ' '), item_arr)

                    success_response['message'] = "Recipe Closed"
                    product.is_open = False
                    product.save()

                else:
                    success_response['status_code'] = 404
                    success_response['message'] = "No Recipe Items"

            # flag stock monitoring
            elif module == 'flag_stock_monitoring':
                prod_pk = data.get('prod_pk')
                flag = data.get('flag')

                product = Products.objects.get(pk=prod_pk)
                name = product.name
                current = product.stock_monitor
                product.stock_monitor = flag
                product.save()

                success_response['message'] = f"Minitory flag changed for product {name} from {current} to {flag} "

        response = success_response

    except Exception as e:
        error_type, error_instance, traceback = sys.exc_info()
        tb_path = traceback.tb_frame.f_code.co_filename
        line_number = traceback.tb_lineno
        response["status_code"] = 500
        response[
            "message"] = f"An error occurred on line {line_number} in file {tb_path}. Details: {e}"

    return JsonResponse(response, safe=False)
