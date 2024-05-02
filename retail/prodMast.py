import openpyxl


class ProdMaster:

    def cursor(self):
        import pyodbc

        # connect to database
        server = '192.168.2.4'
        database = 'SMSEXPV17'
        username = 'sa'
        password = 'sa@123456'
        driver = '{ODBC Driver 17 for SQL Server}'

        # Create a connection string
        conn_str = f'DRIVER={driver};SERVER={server};DATABASE={database};UID={username};PWD={password}'

        connection = pyodbc.connect(conn_str)
        cursor = connection.cursor()
        return cursor

    def get_stock(self, item_code):
        stock_cursor = self.cursor()

        query = f"SELECT stock.loc_id, (select br_name from branch where br_code = stock.loc_id ) as loc_name , isnull(sum(qty),0) as qty, ('2') as trtype, stock_price.avg_cost, stock_price.last_net_cost, stock_price.last_rec_supp, stock_price.last_rec_price, stock_price.local_supp_curr, stock_price.last_rec_date, stock_price.last_cost2, stock_price.last_cost3, isnull(sum(stock.item_wt),0) as tot_wt, stock_price.last_rec_um FROM stock LEFT OUTER JOIN stock_price ON stock.item_code = stock_price.item_code AND stock.loc_id = stock_price.loc_id, user_loc_access ,prod_mast WHERE ( user_loc_access.loc_id = stock.loc_id ) and ( stock.item_code = prod_mast.item_code ) and( ( stock.item_code = '{item_code}' ) AND ( user_loc_access.loc_access = '1' ) AND ( user_loc_access.user_id = '411' ) AND( prod_mast.item_type in ('1','3','5','7')) ) GROUP BY stock.loc_id, stock_price.avg_cost, stock_price.last_net_cost, stock_price.last_rec_price, stock_price.last_rec_um, stock_price.local_supp_curr, stock_price.last_rec_supp, stock_price.last_rec_date, stock_price.last_cost2, stock_price.last_cost3 UNION SELECT stock_chk.loc_id, (select br_name from branch where br_code = stock_chk.loc_id ) as loc_name , isnull(sum(qty),0) as qty, '3', stock_price.avg_cost, stock_price.last_net_cost, stock_price.last_rec_supp, stock_price.last_rec_price, \
        stock_price.local_supp_curr, stock_price.last_rec_date, stock_price.last_cost2,stock_price.last_cost3, isnull(sum(stock_chk.item_wt),0) as tot_wt, stock_price.last_rec_um FROM stock_chk LEFT OUTER JOIN stock_price ON stock_chk.item_code = stock_price.item_code AND stock_chk.loc_id = stock_price.loc_id ,user_loc_access ,prod_mast \
        WHERE  ( user_loc_access.loc_id = stock_chk.loc_id ) and ( stock_chk.item_code = prod_mast.item_code ) and ( ( stock_chk.item_code = '{item_code}' ) AND ( user_loc_access.loc_access = '1' ) AND ( user_loc_access.user_id = '411' ) AND( prod_mast.item_type in ('1','3','5','7')))  GROUP BY stock_chk.loc_id, stock_price.avg_cost, stock_price.last_net_cost, stock_price.last_rec_supp, stock_price.last_rec_price, stock_price.last_rec_um, stock_price.local_supp_curr, stock_price.last_rec_date, stock_price.last_cost2, stock_price.last_cost3 ORDER BY 1 ASC "

        stock_cursor.execute(query)
        nia = 0
        osu = 0
        spintex = 0
        warehouse = 0
        kitchen = 0
        vegetable = 0

        for row in stock_cursor.fetchall():

            loc_id = row[0]
            qty = row[2]

            if loc_id == '001':
                spintex += qty

            if loc_id == '201':
                kitchen += qty

            if loc_id == '202':
                nia += qty

            if loc_id == '203':
                vegetable += qty

            if loc_id == '205':
                osu += qty

            if loc_id == '999':
                warehouse += qty

        # print('nia: ' + str(nia))
        # print('osu: ' + str(osu))
        # print('spintex: ' + str(spintex))
        # print('warehouse: ' + str(warehouse))
        # print('kitchen: ' + str(kitchen))
        # print('vegetable: ' + str(vegetable))

        stock_cursor.execute("select br_code as 'loc_id',RTRIM(br_name) as 'loc_name' from branch")
        locations = stock_cursor.fetchall()

        obj = {}

        for location in locations:
            x_loc_id = location[0]
            loc_name = location[1]
            sold_years = [
                '2018', '2019', '2020', '2021', '2022', '2023', '2024'
            ]
            sold_query = (
                f"SELECT SUM(tran_qty) as 'SOLD',sum(gross_amt) as 'SOLD AMOUNT'  FROM pos_tran where entry_no in "
                f"(select entry_no from pos_tran_hd where entry_date between '2024-01-01' and '2024-12-31' and location_id = '{x_loc_id}' ) and "
                f"prod_id = '{item_code}'")
            stock_cursor.execute(sold_query)

            row = stock_cursor.fetchone()

            if row is not None:
                obj[x_loc_id] = {
                    'sold': row[0],
                    'sold_amt': row[1],
                }

        return {
            'nia': nia,
            'osu': osu,
            'spintex': spintex,
            'warehouse': warehouse,
            'vegetables': vegetable,
            'kitchen': kitchen,
            'total': sum([nia, osu, spintex, warehouse, vegetable, kitchen]),
            'sales': obj
        }

    def groupExport(self):
        import openpyxl
        cursor = self.cursor()
        cursor.execute("select group_code,group_des from group_mast ")
        for row in cursor.fetchall():
            group_code, group_des = row
            print(f"Processing Group: {group_code} - {group_des}")

            # Create a new workbook for each group
            book = openpyxl.Workbook()

            # Get sub groups for the current group
            cursor.execute(f"SELECT sub_group, sub_group_des FROM sub_group WHERE group_code = '{group_code}'")
            sub_groups = cursor.fetchall()

            for sub_group in sub_groups:
                sub_group_code, sub_group_des = sub_group
                ss = sub_group_des.replace('/', '')
                sheet = book.create_sheet(title=ss.strip())  # Use create_sheet to add a new sheet
                sheet.append(
                    ['BARCODE', "NAME", 'SUPPLIER', 'PRICE', 'LAST SOLD', 'SOLD QTY', 'SOLD AMT', 'STOCK_TYPE'])
                # print(f"  Sub-Group: {sub_group_code} - {ss.strip()}")

                # Fetch and insert data into the sheet
                query = (f"SELECT barcode, item_code,(select supp_name from supplier where supplier.supp_code = "
                         f"prod_mast.supp_code) as 'supplier',group_code,sub_group,item_des,retail1,item_type FROm "
                         f"prod_mast where group_code = '{group_code}' and sub_group = '{sub_group_code}'")
                print(query)
                exit()
                for product in cursor.fetchall():
                    barcode, item_code, supplier, group_code, sub_group, item_des, retail1, item_type = product
                    stock_type = "SELLING"
                    if item_type == 0:
                        stock_type = "DISCONTINUED"

                    cursor.execute(
                        f"SELECT TOP(1) prod_id, tran_qty, tran_amt, (SELECT entry_date FROM pos_tran_hd WHERE pos_tran_hd.entry_no = pos_tran.entry_no) AS 'sold_date' FROM pos_tran where prod_id = '{item_code}' ORDER BY sold_date DESC;")
                    last = cursor.fetchone()
                    if last is None:
                        prod_id, tran_qty, tran_amt, sold_date = "NONE"
                    else:
                        prod_id, tran_qty, tran_amt, sold_date = last

                    sheet.append([barcode, item_des, supplier, retail1, sold_date, tran_qty, tran_amt, stock_type])

                    # print()
                    # print(barcode, item_code, supplier, group_code, sub_group, item_des, retail1)
                    # print(last)
                    # print()

            # Remove the default sheet created with the workbook
            del book[book.sheetnames[0]]

            # Save the workbook with a filename based on the group description
            file_name = group_des.replace('/', '').strip()
            book.save(fr"groups/{file_name}.xlsx")
        #
        # for product in cursor:
        #     barcode, item_code, supp_code, group_code, sub_group, item_des, retail1 = product
        #     print(product)

    def groupDiscontinued(self):
        book = openpyxl.Workbook()
        db = self.cursor()
        db.execute("SELECT RTRIM(group_code),RTRIM(group_des) from group_mast")
        for group in db.fetchall():

            grp_code, group_name = group
            print(grp_code, group_name)
            sheet = book.create_sheet(title=group_name.replace('/', ' '))
            sheet.append(['BARCODE', 'NAME', 'SUPPLIER', "STOCK"])
            disc_items = db.execute(
                f"SELECT rtrim(barcode),rtrim(item_des), (select rtrim(supp_name) from supplier where supp_code = prod_mast.supp_code),item_code FROM prod_mast where group_code = '{grp_code}' and item_type = 0")
            disc_items = disc_items.fetchall()
            for item in disc_items:
                barcode, name, suppler, item_code = item
                stock = self.get_stock(item_code)['total']
                sheet.append([barcode, name, suppler, stock])

        book.save("discontinued.xlsx")

    def stockBySupplier(self, supp_code='*'):
        if supp_code == '*':
            query = f"SELECT RTRIM(supp_code),RTRIM(supp_name) from supplier order by supp_name"
        else:
            query = f"SELECT RTRIM(supp_code),RTRIM(supp_name) from supplier where supp_code = '{supp_code}' order by supp_name"

        db = self.cursor()
        db.execute(query)
        rows = db.fetchall()
        # Create a new workbook for each group
        book = openpyxl.Workbook()

        for row in rows:
            supp_code, supp_name = row
            ss = supp_name.replace('/', '')[:30]
            sheet = book.create_sheet(title=ss.strip())  # Use create_sheet to add a new sheet
            sheet.append(
                ['BARCODE', "NAME", 'SPINTEX', 'NIA', 'OSU', 'WAREHOUSE', 'KITCHEN', 'VEGETABLES', 'TOTAL'])

            print("Processing", supp_name, '....')
            db.execute(
                f"SELECT item_code,RTRIM(barcode),RTRIM(item_des) from prod_mast where supp_code = '{supp_code}' order by item_des")
            items = db.fetchall()
            for item in items:
                itemcode, barcode, name = item
                stock = self.get_stock(itemcode)
                nia = stock['nia']
                spintex = stock['spintex']
                osu = stock['osu']
                warehouse = stock['warehouse']
                kitchen = stock['kitchen']
                vegetables = stock['vegetables']
                total = stock['total']

                li = [barcode, name, spintex, nia, osu, warehouse, kitchen, vegetables, total]
                sheet.append(li)

        del book[book.sheetnames[0]]
        file_name = "stock_by_supplier.xlsx"
        book.save(file_name)
        print("Done")

    def stockByGroup(self, group_code='*'):
        db = self.cursor()
        if group_code == '*':
            groups = "SELECT group_code,RTRIM(group_des) FROM group_mast order by group_des"
        else:
            groups = f"SELECT  group_code,TRIM(group_des) FROM group_mast where = group_code = '{group_code}' order by group_des"

        # Create a new workbook for each group
        book = openpyxl.Workbook()
        # get products
        db.execute(groups)
        for group in db.fetchall():
            group_code, group_des = group
            ss = group_des.replace('/', '')[:30]
            sheet = book.create_sheet(title=ss.strip())  # Use create_sheet to add a new sheet
            sheet.append(
                ['BARCODE', "NAME", 'SPINTEX', 'NIA', 'OSU', 'WAREHOUSE', 'KITCHEN', 'VEGETABLES', 'TOTAL'])

            print("Processing", group_des, '....')
            db.execute(
                f"SELECT item_code,RTRIM(barcode),RTRIM(item_des) from prod_mast where group_code = '{group_code}' order by item_des")
            items = db.fetchall()
            for item in items:
                itemcode, barcode, name = item
                stock = self.get_stock(itemcode)
                nia = stock['nia']
                spintex = stock['spintex']
                osu = stock['osu']
                warehouse = stock['warehouse']
                kitchen = stock['kitchen']
                vegetables = stock['vegetables']
                total = stock['total']

                li = [barcode, name, spintex, nia, osu, warehouse, kitchen, vegetables, total]
                sheet.append(li)

        del book[book.sheetnames[0]]
        file_name = "stock_by_group.xlsx"
        book.save(file_name)
        print("Done")

    def getProduct(self, barcode):
        db = self.cursor()
        obj = {
            'barcode': "",
            'itemcode': '',
            'name': '',
            'stock': {},
            'price': 0
        }

        db.execute(
            f"SELECT top(1) item_code,barcode,item_des,retail1 from prod_mast where barcode like '%{barcode}%' or item_code like '%{barcode}%'")
        row = db.fetchone()
        if row is not None:
            item_code, barcode, name, price = row
            obj['barcode'] = barcode.strip()
            obj['itemcode'] = item_code
            obj['name'] = name.strip()
            obj['retail'] = price
            obj['stock'] = self.get_stock(item_code)

        return obj

    def isDiscontinued(self, item_code):
        cur = self.cursor()
        cur.execute(f"Select item_type from prod_mast where item_code = '{item_code}'")
        row = cur.fetchone()
        if row is None:
            return False
        else:
            if row[0] == 1:
                return True
            else:
                return False




