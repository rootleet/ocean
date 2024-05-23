import csv
import json
import traceback
from datetime import date
from django.db import IntegrityError
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from fpdf import FPDF

from admin_panel.anton import make_md5_hash, format_currency
from admin_panel.models import MailQueues, MailSenders, MailAttachments
from cmms.forms import NewSalesCustomer, NewSaleTransactions
from crm.models import Logs
from ocean.settings import DB_SERVER, DB_NAME, DB_USER, DB_PORT, DB_PASSWORD, CMMS_PROF_APPROVER
from django.contrib.auth import get_user_model
import pyodbc
from cmms.models import *
from decimal import Decimal
from django.contrib import messages

from cmms.extra import db


@csrf_exempt
def api(request):
    global header
    method = request.method
    response = {"status_code": "", "message": ""}
    ce = ""
    try:
        body = json.loads(request.body)
        module = body.get('module')
        data = body.get('data')

        if method == 'VIEW':
            try:
                arr = []
                if module == 'product':
                    scope = data.get('range')
                    item_uni = data.get('item_uni')
                    db_col = data.get('db_col') or 'barcode'

                    if scope == 'single':
                        q = f"SELECT barcode,item_ref,item_des1 FROM product_master WHERE {db_col} = '{item_uni}'"
                    else:
                        q = "SELECT barcode,item_ref,item_des1  FROM product_master"

                    server = f"{DB_SERVER},{DB_PORT}"
                    database = DB_NAME
                    username = DB_USER
                    password = DB_PASSWORD
                    driver = '{ODBC Driver 17 for SQL Server}'  # Change this to the driver you're using
                    connection_string = f"DRIVER={driver};SERVER={server};DATABASE={database};UID={username};PWD={password}"
                    connection = pyodbc.connect(connection_string)
                    cursor = connection.cursor()

                    cursor.execute(q)
                    count = 0
                    arr = []
                    for part in cursor.fetchall():
                        count += 1
                        obj = {'barcode': part[0].strip(), 'item_ref': part[1].strip(), 'name': part[2].strip()}
                        arr.append(obj)

                    resp = {'count': count, 'trans': arr}
                    response['status_code'] = 200
                    response['status'] = 'success'
                    response['message'] = resp
                    cursor.close()
                    connection.close()

                elif module == 'groups':
                    cursor = db()
                    query = "select Group_code,group_des from group_master order by group_des"
                    cursor.execute(query)
                    rows = cursor.fetchall()
                    arr = []
                    grp_counts = 0
                    for row in rows:
                        grp_counts += 1
                        group_code = row[0].strip()
                        group_name = row[1].strip()
                        arr.append({'code': group_code, 'name': group_name})

                    response['message'] = {
                        'counts': grp_counts, 'groups': arr
                    }
                    response['status_code'] = 200
                    response['status'] = 'success'

                elif module == 'stock':
                    stage = data.get('stage')
                    if stage == 'active':
                        # check for active
                        if StockCountHD.objects.filter(status=1).exists():
                            active = StockCountHD.objects.get(status=1)
                            trans = StockCountTrans.objects.filter(stock_count_hd=active).order_by('-pk')
                            arr = []
                            for tran in trans:
                                obj = {
                                    'item_ref': tran.item_ref,
                                    'barcode': tran.barcode,
                                    'name': tran.name,
                                    'quantity': tran.quantity,
                                    'price': tran.sell_price,
                                    'value': tran.quantity * tran.sell_price,
                                    'owner': tran.owner,
                                    'comment': tran.comment
                                }
                                arr.append(obj)
                            response['message'] = {
                                'count': trans.count(), 'trans': arr
                            }
                            response['status_code'] = 200
                            response['status'] = 'success'
                        else:
                            response['message'] = 'NO DATA'
                            response['status_code'] = 404
                            response['status'] = 'error'

                    elif stage == 'preview':
                        style = data.get('compare')
                        if style == 'final_compare':
                            doc = data.get('doc')
                            tot_phy = Decimal(0.00)
                            tot_sys = Decimal(0.00)
                            tot_qdif = Decimal(0.00)
                            tot_vdif = Decimal(0.00)
                            if doc == 'excel':
                                import openpyxl
                                workbook = openpyxl.Workbook()
                                sheet = workbook.active
                                # make sheet head
                                sheet[f'A1'] = "ITEM REFERENCE"
                                sheet[f'B1'] = "BARCODED"
                                sheet[f'C1'] = "DESCRIPTION"
                                sheet[f'D1'] = "PHYSICAL QUANTITY"
                                sheet[f'E1'] = "SYSTEM QUANTITY"
                                sheet[f'F1'] = "QUANTITY DIFFERENCE"
                                sheet[f'G1'] = "SELLING PRICE"
                                sheet[f'H1'] = "VALUE DIFFERENCE"
                                sheet[f'I1'] = "COMMENT"

                            # check if there is open stock
                            pk = data.get('pk')
                            open_st = StockCountHD.objects.filter(pk=pk).count()
                            is_open = False
                            if open_st == 1:

                                stock_hd = StockCountHD.objects.get(pk=pk)
                                trans = StockCountTrans.objects.filter(stock_count_hd=stock_hd)
                                frozen = stock_hd.frozen
                                header = {
                                    'frozen': {
                                        'pk': frozen.pk,
                                        'loc': frozen.loc_id,
                                        'remarks': frozen.remarks,
                                        'status': frozen.status,
                                        'owner': frozen.owner.username,
                                        'created': f"{frozen.created_date} {frozen.created_time}",
                                        'entry': frozen.ref,
                                        'posted': frozen.posted()

                                    },
                                    'count': {
                                        'pk': stock_hd.pk,
                                        'entry': stock_hd.entry_no(),
                                        'created': stock_hd.created_at,
                                        'comment': stock_hd.comment,
                                        'status': stock_hd.status,
                                        'owner': stock_hd.owner.username,
                                        'next': stock_hd.next(),
                                        'prev': stock_hd.prev(),
                                        'approved': stock_hd.approve
                                    }
                                }
                                arr = []
                                not_arr = []
                                entries = {
                                    'wr_entry': 0,
                                    'lost': 0,
                                    'sys_error': 0,
                                    'unknown': 0,
                                    'not_counted': 0
                                }
                                values = {
                                    'wr_entry': 0,
                                    'lost': 0,
                                    'sys_error': 0,
                                    'unknown': 0,
                                    'not_counted': 0
                                }
                                row_c = 2
                                for tran in trans:
                                    comment = "NO COMMENT"
                                    if len(tran.comment) > 0:
                                        comment = tran.comment
                                    arr.append({
                                        'item_ref': tran.item_ref,
                                        'barcode': tran.barcode,
                                        'name': tran.name,
                                        'froze_qty': tran.froze_qty,
                                        'counted_qty': tran.counted_qty,
                                        'diff_qty': tran.diff_qty,
                                        'comment': comment,
                                        'issue': tran.issue
                                    })

                                if doc == 'preview':
                                    header['entries'] = entries
                                    header['values'] = values
                                    response['message'] = {
                                        'header': header, 'trans': arr
                                    }
                                elif doc == 'excel':
                                    sheet[f'A2'] = "SUMMARY"
                                    sheet['D2'] = tot_phy
                                    sheet['E2'] = tot_sys
                                    sheet['F2'] = tot_qdif
                                    sheet['G2'] = "-"
                                    sheet['H2'] = tot_vdif
                                    from datetime import datetime
                                    current_datetime = datetime.now()
                                    formatted_datetime = current_datetime.strftime("%Y%m%d%H%M%S")
                                    file = f"static/general/tmp/{stock_hd.loc}_{g_name.strip().replace(' ', '_')}_AS_OF_{as_of}_{formatted_datetime}.xlsx"
                                    workbook.save(file)
                                    response['message'] = file

                                response['status_code'] = 200
                                response['status'] = 'success'
                            else:
                                response['message'] = "NO OPEN STOCK"
                                response['status_code'] = 404
                                response['status'] = 'limit'

                        elif style == 'count_sheet':
                            print(data)
                            doc = data.get('doc')
                            key = data.get('key')
                            print(doc)

                            if doc == 'fr':
                                # frozen
                                if StockFreezeHd.objects.filter(pk=key).count() == 1:

                                    f_hd = StockFreezeHd.objects.get(pk=key)
                                    trans = f_hd.trans_only()

                                    pdf = FPDF('P', 'mm', 'A4')
                                    pdf.add_page()
                                    pdf.set_font('Arial', 'BU', 10)
                                    pdf.cell(180, 5, "STOCK COUNT SHEET", 0, 1, 'C')
                                    pdf.ln(10)

                                    pdf.set_font('Arial', 'B', 10)
                                    pdf.cell(25, 5, "LOCATION : ", 0, 0, 'L')
                                    pdf.set_font('Arial', '', 7)
                                    pdf.cell(80, 5, f" {f_hd.loc_id}", 0, 1, 'L')

                                    pdf.set_font('Arial', 'B', 10)
                                    pdf.cell(25, 5, "REMARKS : ", 0, 0, 'L')
                                    pdf.set_font('Arial', '', 7)
                                    pdf.cell(80, 5, f" {f_hd.remarks}", 0, 1, 'L')
                                    pdf.ln(10)
                                    # header
                                    pdf.set_font('Arial', 'B', 10)
                                    pdf.cell(10, 5, f"LN", 1, 0, 'L')
                                    pdf.cell(10, 5, f"LN", 1, 0, 'L')
                                    pdf.cell(25, 5, f"ITEM REF", 1, 0, 'L')
                                    pdf.cell(30, 5, f"BARCODE", 1, 0, 'L')
                                    pdf.cell(100, 5, f"NAME", 1, 0, 'L')
                                    # pdf.cell(20, 5, f"FROZEN", 1, 0, 'L')
                                    pdf.cell(20, 5, f"COUNTED", 1, 1, 'L')

                                    pdf.set_font('Arial', '', 7)

                                    line = 0

                                    for tran in trans:
                                        line += 1
                                        pdf.cell(10, 5, f"{line}", 1, 0, 'L')
                                        pdf.cell(25, 5, f"{tran.item_ref}", 1, 0, 'L')
                                        pdf.cell(30, 5, f"{tran.barcode}", 1, 0, 'L')
                                        pdf.cell(100, 5, f"{tran.name}", 1, 0, 'L')
                                        # pdf.cell(20, 5, f"{tran.qty}", 1, 0, 'L')
                                        pdf.cell(20, 5, f"", 1, 1, 'L')

                                    file = f'static/general/tmp/{f_hd.loc_id}.pdf'
                                    pdf.output(file, 'F')

                                    response = {
                                        'status_code': 200, 'message': file, 'status': 'success'
                                    }
                                else:
                                    response = {
                                        'status_code': 404, 'message': "ENTRY NOT FOUND", 'status': 'not_found'
                                    }
                            elif doc == 'stc':
                                if StockCountHD.objects.filter(pk=key).count() == 1:
                                    hd = StockCountHD.objects.get(pk=key)
                                    f_hd = hd.frozen
                                    trans = hd.trans()
                                    trans_only = trans['trans']

                                    pdf = FPDF('P', 'mm', 'A4')
                                    pdf.add_page()
                                    pdf.set_font('Arial', 'BU', 10)
                                    pdf.cell(180, 5, "COUNT REPORT", 0, 1, 'C')
                                    pdf.ln(10)

                                    pdf.set_font('Arial', 'B', 10)
                                    pdf.cell(25, 5, "LOCATION : ", 0, 0, 'L')
                                    pdf.set_font('Arial', '', 7)
                                    pdf.cell(80, 5, f" {f_hd.loc_id}", 0, 1, 'L')

                                    pdf.set_font('Arial', 'B', 10)
                                    pdf.cell(25, 5, "REMARKS : ", 0, 0, 'L')
                                    pdf.set_font('Arial', '', 7)
                                    pdf.cell(80, 5, f" {f_hd.remarks}", 0, 1, 'L')
                                    pdf.ln(10)
                                    # header
                                    pdf.set_font('Arial', 'B', 8)
                                    pdf.cell(10, 5, f"LN", 1, 0, 'L')
                                    pdf.cell(20, 5, f"BARCODE", 1, 0, 'L')
                                    pdf.cell(55, 5, f"NAME", 1, 0, 'L')
                                    pdf.cell(15, 5, f"SELL", 1, 0, 'L')
                                    pdf.cell(15, 5, f"FRZN", 1, 0, 'L')
                                    pdf.cell(15, 5, f"FR VL", 1, 0, 'L')
                                    pdf.cell(15, 5, f"CT", 1, 0, 'L')
                                    pdf.cell(15, 5, f"CT VL", 1, 0, 'L')
                                    pdf.cell(15, 5, f"DIFF", 1, 0, 'L')
                                    pdf.cell(15, 5, f"DF VL", 1, 1, 'L')
                                    pdf.set_font('Arial', '', 5)
                                    line = 0
                                    summary = trans['summary']
                                    pdf.cell(100, 5, f"SUMMARY", 1, 0, 'L')
                                    pdf.cell(15, 5, f"{summary['total_frozen']}", 1, 0, 'L')
                                    pdf.cell(15, 5, f"{summary['value_frozen']}", 1, 0, 'L')
                                    pdf.cell(15, 5, f"{summary['total_counted']}", 1, 0, 'L')
                                    pdf.cell(15, 5, f"{summary['value_counted']}", 1, 0, 'L')
                                    pdf.cell(15, 5, f"{summary['qty_difference']}", 1, 0, 'L')
                                    pdf.cell(15, 5, f"{summary['value_difference']}", 1, 1, 'L')

                                    pdf.set_font('Arial', '', 5)
                                    for tran in trans_only:
                                        line += 1
                                        pdf.cell(10, 5, f"{line}", 1, 0, 'L')
                                        pdf.cell(20, 5, f"{tran.barcode}", 1, 0, 'L')
                                        pdf.cell(55, 5, f"{tran.name}", 1, 0, 'L')
                                        pdf.cell(15, 5, f"{tran.sell_price}", 1, 0, 'L')
                                        pdf.cell(15, 5, f"{tran.froze_qty}", 1, 0, 'L')
                                        pdf.cell(15, 5, f"{tran.froze_qty * tran.sell_price}", 1, 0, 'L')
                                        pdf.cell(15, 5, f"{tran.counted_qty}", 1, 0, 'L')
                                        pdf.cell(15, 5, f"{tran.counted_qty * tran.sell_price}", 1, 0, 'L')
                                        pdf.cell(15, 5, f"{tran.diff_qty}", 1, 0, 'L')
                                        pdf.cell(15, 5, f"{tran.diff_qty * tran.sell_price}", 1, 1, 'L')

                                        # summary

                                    pdf.set_font('Arial', 'B', 8)
                                    pdf.cell(100, 5, f"SUMMARY", 1, 0, 'L')
                                    pdf.cell(15, 5, f"{summary['total_frozen']}", 1, 0, 'L')
                                    pdf.cell(15, 5, f"{summary['value_frozen']}", 1, 0, 'L')
                                    pdf.cell(15, 5, f"{summary['total_counted']}", 1, 0, 'L')
                                    pdf.cell(15, 5, f"{summary['value_counted']}", 1, 0, 'L')
                                    pdf.cell(15, 5, f"{summary['qty_difference']}", 1, 0, 'L')
                                    pdf.cell(15, 5, f"{summary['value_difference']}", 1, 1, 'L')

                                    file = f'static/general/tmp/{hd.entry_no()}.pdf'
                                    pdf.output(file, 'F')

                                    response = {
                                        'status_code': 200, 'message': file, 'status': 'success'
                                    }


                                else:
                                    response = {
                                        'status_code': 404, 'message': "ENTRY NOT FOUND", 'status': 'not_found'
                                    }
                            else:
                                response['message'] = f"UNKNOWN DOCUMENT {doc}"

                    elif stage == 'frozen':
                        pk = data.get('pk')

                        # check if hd exist
                        hd_x = StockFreezeHd.objects.filter(pk=pk)

                        if hd_x.count() == 1:
                            response['status_code'] = 200

                            hd = StockFreezeHd.objects.get(pk=pk)
                            trans = hd.trans()

                            header = {
                                'pk': hd.pk,
                                'loc': hd.loc_id,
                                'remarks': hd.remarks,
                                'ref': hd.ref,
                                'time': f"{hd.created_date} {hd.created_time}",
                                'status': hd.status,
                                'owner': hd.owner.username,
                                'next': hd.next(),
                                'prev': hd.prev(),
                                'approve': hd.approve
                            }

                            tr = []

                            for t in trans['trans']:
                                ref = t.item_ref
                                barcode = t.barcode
                                name = t.name
                                qty = t.qty

                                tr.append({
                                    'item_ref': ref,
                                    'barcode': barcode,
                                    'name': name,
                                    'qty': qty,
                                    'price': t.price
                                })

                            response['message'] = {
                                'header': header,
                                'count': trans['count'],
                                'trans': tr
                            }

                        else:
                            response['status_code'] = 404
                            response['message'] = f"Expected 1 but got {hd_x.count()} counts"

                    elif stage == 'frozen_hd':

                        frozn = StockFreezeHd.objects.filter(status=1).order_by('-pk')
                        counts = frozn.count()
                        arr = []
                        for fr in frozn:
                            arr.append({
                                'pk': fr.pk,
                                'entry': f"FR{fr.loc_id}{fr.pk}",
                                'remarks': fr.remarks,
                                'location': fr.loc_id,
                                'approve': fr.approve,
                                'posted': fr.posted()
                            })

                        response['message'] = arr
                        response['status_code'] = 200

                    elif stage == 'ref_frozen':
                        try:
                            server = f"{DB_SERVER},{DB_PORT}"
                            database = DB_NAME
                            username = DB_USER
                            password = DB_PASSWORD
                            driver = '{ODBC Driver 17 for SQL Server}'  # Change this to the driver you're using
                            connection_string = f"DRIVER={driver};SERVER={server};DATABASE={database};UID={username};PWD={password}"
                            connection = pyodbc.connect(connection_string)
                            cursor = connection.cursor()

                            ref = data.get('ref')
                            query = f"SELECT item_ref,qty_avail,unit_price,(SELECT barcode from product_master where item_ref = stock_freeze_tran.item_ref ) as 'barcode',(SELECT item_des1 from product_master where item_ref = stock_freeze_tran.item_ref ) as 'name' from stock_freeze_tran where entry_no = '{ref}'"
                            trans = []
                            cursor.execute(query)
                            count = 0
                            for tran in cursor.fetchall():
                                count += 1
                                trans.append({
                                    'item_ref': str(tran[0].strip()),
                                    'qty': tran[1],
                                    'unit_price': tran[2],
                                    'barcode': str(tran[3]).strip(),
                                    'name': str(tran[4]).strip()
                                })

                            # get header
                            h_q = f"SELECT loc_id,remarks from stock_freeze_hd where entry_no = '{ref}'"
                            h_row = cursor.execute(h_q).fetchone()

                            if h_row is None:
                                loc = 'none',
                                rem = 'none'
                            else:
                                loc = str(h_row[0]).strip()
                                rem = str(h_row[1]).strip()

                            header = {
                                'loc': loc, 'remarks': rem
                            }

                            response['message'] = {
                                'header': header,
                                'count': count,
                                'trans': trans
                            }

                            response['status_code'] = 200
                        except Exception as e:
                            response['status_code'] = 500
                            response['message'] = f"{str(e)} line : 502"

                    elif stage == 'export':
                        doc = data.get('doc')
                        key = data.get('key')
                        document = data.get('document')

                        if doc == 'STC':
                            # count
                            hd = StockCountHD.objects.filter(pk=key)
                        elif doc == 'STR':
                            # sales transactions
                            hd = SalesCustomers.objects.filter(pk=key)
                        count = hd.count()
                        if count == 1:
                            response['status_code'] = 200
                            header = hd.last()
                            trans = header.trans()

                            if document == 'csv' and doc == 'STC':
                                file_name = f"static/general/tmp/{header.frozen.loc_id} - {header.frozen.remarks} - {header.frozen.ref}.csv"
                                # Header
                                header = ["ITEM_REF", "COUNTED"]

                                with open(file_name, mode="w", newline='') as file:
                                    writer = csv.writer(file)

                                    # Writing the header
                                    writer.writerow(header)

                                    for tran in trans['trans']:
                                        if tran.counted_qty > 0:
                                            counted_qty = tran.counted_qty
                                        else:
                                            counted_qty = tran.counted_qty

                                        writer.writerow(
                                            [tran.item_ref, counted_qty])

                                response['message'] = file_name

                            elif document == 'excel' and doc == 'STC':
                                import openpyxl
                                workbook = openpyxl.Workbook()
                                sheet = workbook.active
                                # make sheet head
                                sheet['A1'] = f"LOCATION : {header.frozen.loc_id}"
                                sheet['A2'] = f"REFERENCE : {header.frozen.ref}"
                                sheet['A3'] = f"REMARK : {header.frozen.remarks}"

                                sheet[f'A5'] = "ITEM REFERENCE"
                                sheet[f'B5'] = "BARCODED"
                                sheet[f'C5'] = "DESCRIPTION"
                                sheet[f'D5'] = "FROZEN QTY"
                                sheet[f'E5'] = "COUNTED QTY"
                                sheet[f'F5'] = "DIFFERENCE"

                                sheet['A6'] = "SUMMARY"
                                sheet['D6'] = trans['summary']['total_frozen']
                                sheet['E6'] = trans['summary']['total_counted']
                                sheet['F6'] = trans['summary']['qty_difference']

                                row = 7
                                for tran in trans['trans']:
                                    row += 1
                                    sheet[f"A{row}"] = tran.item_ref
                                    sheet[f"B{row}"] = tran.barcode
                                    sheet[f"C{row}"] = tran.name
                                    sheet[f"D{row}"] = tran.froze_qty
                                    sheet[f"E{row}"] = tran.counted_qty
                                    sheet[f"F{row}"] = tran.diff_qty

                                from datetime import datetime
                                current_datetime = datetime.now()
                                formatted_datetime = current_datetime.strftime("%Y%m%d%H%M%S")
                                file = f"static/general/tmp/{doc}_{header.frozen.ref}.xlsx"
                                workbook.save(file)
                                response['message'] = file

                            elif document == 'excel' and doc == "STR":
                                import openpyxl
                                workbook = openpyxl.Workbook()
                                sheet = workbook.active
                                # make sheet head
                                sheet['A1'] = f"DATE"
                                sheet['B1'] = f"BY"
                                sheet['C1'] = f"TITLE"
                                sheet['D1'] = "Description"

                                # loop through transactins
                                count = 1
                                for tran in header.trans():
                                    count += 1
                                    sheet[f'A{count}'] = f"{tran.created_date} {tran.created_time}"
                                    sheet[f'B{count}'] = f"{tran.owner.username}"
                                    sheet[f'C{count}'] = f"{tran.title}"
                                    sheet[f'D{count}'] = f"{tran.details}"

                                from datetime import datetime
                                current_datetime = datetime.now()
                                formatted_datetime = current_datetime.strftime("%Y%m%d%H%M%S")
                                file = f"static/general/tmp/{doc}_{header.name}.xlsx"
                                workbook.save(file)
                                response['message'] = file

                        else:
                            response['status_code'] = 404
                            response['message'] = f"ENTRY NOT FOUNC ( {doc} - {key} )"

                elif module == 'customer':
                    cust_type = data.get('cust_type')
                    if cust_type == 'cmms_service':
                        cursor = db()
                        cust_code = data.get('cust_code') or 'all'
                        if cust_code == 'all':
                            cursor.execute("select cust_code,cust_name,cust_contact,email1 from customer_master")
                        else:
                            cursor.execute(
                                f"select cust_code,cust_name,cust_contact,email1 from customer_master  where cust_code = '{cust_code}'")
                        cust_arr = []
                        for customer in cursor.fetchall():
                            cust_arr.append({
                                'code': str(customer[0]).strip(),
                                'name': str(customer[1]).strip(),
                                'phone': str(customer[2]).strip(),
                                'email': str(customer[3]).strip()
                            })
                        response['message'] = cust_arr
                        response['status_code'] = 200

                elif module == 'cust_asset':
                    customer = data.get('customer')

                    # check if customer exist
                    cust_count = db().execute(
                        f"SELECT COUNT(*) FROM customer_master where cust_code = '{customer}'").fetchone()[0]
                    if cust_count == 1:
                        # customer details
                        customer_details = db().execute(
                            f"select cust_code,cust_name,cust_contact,email1 from customer_master "
                            f"where cust_code = '{customer}'").fetchone()
                        header = {
                            'code': customer,
                            'name': str(customer_details[1]).strip(),
                            'phone': str(customer_details[2]).strip()
                        }
                        assets_q = f"select ASSET_CODE,asset_desc,asset_no,asset_ref_no,model_no from asset_mast where cust_code = '{customer}'"
                        asset_arr = []
                        asset_cursor = db()
                        asset_cursor.execute(assets_q)

                        for asset in asset_cursor.fetchall():
                            asset_cursor.execute(
                                f"SELECT count(*) as service_times FROM invoice_hd where asset_code = '{str(asset[0]).strip()}'")

                            obj = {
                                'code': str(asset[0]).strip(),
                                'name': str(asset[1]).strip(),
                                'number': str(asset[2]).strip(),
                                'chassis': str(asset[3]).strip(),
                                'model': str(asset[4]).strip(),
                                'service_times': asset_cursor.fetchone()[0]
                            }
                            asset_arr.append(obj)

                        response['status_code'] = 200
                        response['message'] = {
                            'customer': header, 'assets': asset_arr
                        }
                        asset_cursor.close()
                    else:
                        response['status_code'] = 404
                        response['message'] = f"NO CUSTOMER WITH CODE {customer}"

                elif module == 'service_history':
                    conn = db()
                    asset = data.get('asset')
                    conn.execute(
                        f"select Entry_no,Invoice_date,wo_date,tot_amt,tax_amt,net_amt,labor_amount,material_amount from invoice_hd where asset_code = '{asset}' order by Invoice_date desc")
                    serv_arr = []
                    for service in conn.fetchall():

                        # get materials
                        conn.execute(
                            f"select invoice_type,descr,uom,unit_qty,unit_price,tot_amt,tax_amt,net_amt from invoice_tran where Entry_no = '{service[0]}'")
                        tran_arr = []
                        for tran in conn.fetchall():
                            tobj = {
                                'type': str(tran[0]).strip(),
                                'name': str(tran[1]).strip(),
                                'packing': str(tran[2]).strip(),
                                'unit_qty': str(tran[3]).strip(),
                                'unit_price': str(tran[4]).strip(),
                                'net': str(tran[5]).strip(),
                                'tax': str(tran[6]).strip(),
                                'gross': str(tran[7]).strip()
                            }
                            tran_arr.append(tobj)

                        obj = {
                            'invoice': str(service[0]).strip(),
                            'date': str(service[1]).strip(),
                            'tot_amt': str(service[3]).strip(),
                            'tax_amt': str(service[4]).strip(),
                            'net_amt': str(service[5]).strip(),
                            'lab_amt': str(service[6]).strip(),
                            'mat_amt': str(service[7]).strip(),
                            'trans': tran_arr
                        }
                        serv_arr.append(obj)

                    conn.close()
                    response['status_code'] = 200
                    response['message'] = serv_arr

                elif module == 'invoice_detail':
                    invoice = data.get('invoice')
                    header = {}
                    assets = {}
                    trans = []

                    conn = db()
                    inv_hd = conn.execute(
                        f"select Entry_no,Invoice_date,tot_amt,tax_amt,net_amt,labor_amount,material_amount,asset_code from invoice_hd where Entry_no= '{invoice}'").fetchone()
                    header['invoice_no'] = str(inv_hd[0]).strip()
                    header['date'] = str(inv_hd[1]).strip()
                    header['net'] = str(inv_hd[2]).strip()
                    header['tax'] = str(inv_hd[3]).strip()
                    header['gross'] = str(inv_hd[4]).strip()
                    header['lab'] = str(inv_hd[5]).strip()
                    header['mat'] = str(inv_hd[6]).strip()

                    # asset details
                    asset_code = inv_hd[7]
                    asset = conn.execute(
                        f"select ASSET_CODE,asset_desc,asset_no,asset_ref_no,model_no,cust_code from asset_mast where ASSET_CODE = '{asset_code}'").fetchone()
                    assets['code'] = str(asset[0]).strip()
                    assets['name'] = str(asset[1]).strip()
                    assets['number'] = str(asset[2]).strip()
                    assets['chassis'] = str(asset[3]).strip()
                    assets['model'] = str(asset[4]).strip()
                    assets['owner'] = str(asset[5]).strip()

                    conn.execute(
                        f"select invoice_type,descr,uom,unit_qty,unit_price,tot_amt,tax_amt,net_amt from invoice_tran where Entry_no = '{invoice}'")
                    for tran in conn.fetchall():
                        tobj = {
                            'type': str(tran[0]).strip(),
                            'name': str(tran[1]).strip(),
                            'packing': str(tran[2]).strip(),
                            'unit_qty': str(tran[3]).strip(),
                            'unit_price': str(tran[4]).strip(),
                            'net': str(tran[5]).strip(),
                            'tax': str(tran[6]).strip(),
                            'gross': str(tran[7]).strip()
                        }
                        trans.append(tobj)

                    response['message'] = {
                        'header': header, 'asset': assets, 'trans': trans
                    }
                    response['status_code'] = 200

                elif module == 'just_costomer':
                    cust_code = data.get('code')
                    db().execute(
                        f"SELECT cust_name, cust_contact, email1, email2 FROM customer_master WHERE cust_code = '{cust_code}'"
                    )

                    row = db().fetchone()
                    if row:
                        response = {
                            'status_code': 200,
                            'message': {
                                'name': str(row[0]).strip(),
                                'phone': str(row[1]).strip(),
                                'email': str(row[2]).strip()
                            }
                        }
                    else:
                        response = {
                            'status_code': 404,
                            'message': 'Customer not found'
                        }

                elif module == 'cmms_sales_customer':
                    key = data.get('key')
                    arr = []

                    if key == 'all':
                        customers = SalesCustomers.objects.all()
                    else:
                        customers = SalesCustomers.objects.filter(pk=key)

                    if customers.count() > 0:
                        for customer in customers:
                            x_region = getattr(customer.region, 'name', 'UNKNOWN') if customer and hasattr(customer,
                                                                                                           'region') else 'unknown'
                            suburb = getattr(customer.suburb, 'name', 'UNKNOWN') if customer and hasattr(customer,
                                                                                                         'suburb') else 'unknown'

                            obj = {
                                'pk': customer.pk,
                                'url': customer.url,
                                'company': customer.company,
                                'name': customer.name,
                                'mobile': customer.mobile,
                                'email': customer.email,
                                'address': customer.address,
                                'type_of_client': customer.type_of_client,
                                'timestamp': {
                                    'created_date': customer.created_date,
                                    'created_time': customer.created_time,
                                    'updated_date': customer.updated_date,
                                    'updated_time': customer.updated_time
                                },
                                'status': customer.status,
                                'owner': {
                                    'pk': customer.owner.pk,
                                    'username': customer.owner.username
                                },
                                'geography': {
                                    'region': x_region,
                                    'suburb': suburb
                                }

                            }
                            arr.append(obj)

                        response['status_code'] = 200
                        response['message'] = arr
                    else:
                        response['status_code'] = 404
                        response['message'] = f"NO CUSTOMER FOUND WITH KEY {key}"

                elif module == 'cmms_sales_deals':
                    cust = data.get('sales_customer')
                    if SalesCustomers.objects.filter(pk=cust).count() == 1:
                        customer = SalesCustomers.objects.get(pk=cust)
                        deals = customer.deals()
                        deal_arr = []
                        for deal in deals:
                            obj = {
                                'sales_rep': {
                                    'pk': deal.owner.pk,
                                    'username': deal.owner.username
                                },
                                'purch_rep': {
                                    'name': deal.pur_rep_name,
                                    'email': deal.pur_rep_email,
                                    'phone': deal.pur_rep_phone,
                                    'company': deal.customer.company
                                },
                                'asset': {
                                    'name': deal.asset,
                                    'model': deal.model,
                                    'stock_type': deal.stock_type,
                                    'requirement': deal.requirement
                                },
                                'status': deal.status,
                                'date': deal.created_date,
                                'time': deal.created_time,
                                'pk': deal.pk

                            }
                            deal_arr.append(obj)
                        response['status_code'] = 200
                        response['message'] = deal_arr

                    else:
                        response['status_code'] = 404
                        response['message'] = f"CUSTOMER DOES NOT EXIST"

                elif module == 'cmms_sales_deal':
                    dealkey = data.get('deal')

                    try:
                        deal = SalesDeals.objects.get(pk=dealkey)
                        trans = deal.transactions()
                        transactions = []
                        for tran in trans:
                            obj = {
                                'title': tran.title,
                                'details': tran.details,
                                'owner': {
                                    'pk': tran.owner.pk,
                                    'username': tran.owner.username
                                },
                                'timestamp': {
                                    'd_created': tran.created_date,
                                    't_created': tran.created_time
                                }
                            }
                            transactions.append(obj)

                        header = {
                            'pk': deal.pk,
                            'sales_rep': {
                                'pk': deal.owner.pk,
                                'username': deal.owner.username
                            },
                            'purch_rep': {
                                'name': deal.pur_rep_name,
                                'email': deal.pur_rep_email,
                                'phone': deal.pur_rep_phone,
                                'company': deal.customer.company
                            },
                            'asset': {
                                'name': deal.asset,
                                'model': deal.model,
                                'stock_type': deal.stock_type,
                                'requirement': deal.requirement
                            },
                            'status': deal.status,
                            'date': deal.created_date,
                            'time': deal.created_time
                        }

                        response['status_code'] = 200
                        response['message'] = {
                            'deal': header, 'trans': transactions
                        }

                    except Exception as e:
                        response['status_code'] = 505
                        response['message'] = str(e)

                elif module == 'reqtran':
                    request_number = data.get('reqnum')
                    cursor = db()

                    cursor.execute(f"SELECT * FROM mr_request_hd where entry_no = '{request_number}'")
                    req = cursor.fetchone()

                    if req is not None:
                        arr = []
                        cursor.execute(
                            f"select item_code,barcode,item_des,total_units,entry_date from mr_request_tran where entry_no = '{request_number}'")
                        for tran in cursor.fetchall():
                            item_code, barcode, item_des, total_units, entry_date = tran

                            # get previous record

                            cursor.execute(
                                f"SELECT top(1) entry_no,doc_date,doc_ref_no1,total_units FROM stock_chk where doc_type in ('TR','GR','AD') and item_code = '{item_code}' and doc_date <= '{entry_date}' and doc_ref_no1 != '{request_number}' order by doc_date desc")
                            last_transfer = cursor.fetchone()
                            if last_transfer is not None:
                                entry_no, doc_date, doc_ref_no1, total_tran_units = last_transfer

                                cursor.execute(
                                    f"select sum(total_units) from stock_chk where doc_type = 'IS' and item_code = '{item_code}' and doc_date <= '{entry_date}'")
                                iss_t = cursor.fetchone()[0]

                                arr.append({
                                    'itemcode': item_code.strip(),
                                    'barcode': barcode.strip(),
                                    'name': item_des.strip(),
                                    'unit': total_units,
                                    'transfer': {
                                        'entry_no': entry_no,
                                        'qty': total_tran_units,
                                        'tranfer_date': doc_date
                                    },
                                    'usage': iss_t
                                })

                        response['status_code'] = 202
                        response['message'] = arr
                    else:
                        response['status_code'] = 404
                        response['message'] = f"NO RECORD FOR {request_number}"

                elif module == 'asset_group':
                    target = data.get('target', '*')
                    if target == '*':
                        groups = SalesAssetsGroup.objects.all()
                    else:
                        groups = SalesAssetsGroup.objects.filter(pk=target)

                    # loop through groups
                    for group in groups:
                        # add to list
                        arr.append(group.obj())

                    response['status_code'] = 200
                    response['message'] = arr

                elif module == 'CarOrigin':
                    key = data.get('key', '*')
                    if key == '*':
                        ogs = CarOrigin.objects.all().order_by('country')
                    else:
                        ogs = CarOrigin.objects.filter(pk=key).order_by('country')
                    arr = []
                    for og in ogs:
                        arr.append({
                            'country': og.country,
                            'pk': og.pk
                        })

                    response['status_code'] = 200
                    response['message'] = arr

                elif module == 'CarManufacturer':
                    key = data.get('key', '*')
                    if key == '*':
                        manfs = CarManufacturer.objects.all().order_by('name')
                    else:
                        manfs = CarManufacturer.objects.filter(pk=key).order_by('name')
                    arr = []
                    for man in manfs:
                        arr.append({
                            'origin': man.origin.country,
                            'pk': man.pk,
                            'name': man.name
                        })

                    response['status_code'] = 200
                    response['message'] = arr

                elif module == 'Car':
                    key = data.get('key', '*')
                    if key == '*':
                        cars = Car.objects.all().order_by('name')
                    else:
                        cars = Car.objects.filter(pk=key).order_by('name')
                    arr = []
                    for car in cars:
                        arr.append(car.myself())

                    response['status_code'] = 200
                    response['message'] = arr


                elif module == 'CarSupplier':
                    key = data.get('key', '*')
                    if key == '*':
                        manfs = CarSupplier.objects.all().order_by('name')
                    else:
                        manfs = CarSupplier.objects.filter(pk=key).order_by('name')
                    arr = []
                    for man in manfs:
                        arr.append({
                            'origin': man.origin.country,
                            'pk': man.pk,
                            'name': man.name,
                            'email': man.email,
                            'phone': man.phone,
                            'website': man.website,
                            'created_on': man.created_on,
                            'created_by': man.created_by.first_name
                        })

                    response['status_code'] = 200
                    response['message'] = arr

                elif module == 'CarModel':
                    key = data.get('key', '*')
                    if key == '*':
                        mods = CarModel.objects.all().order_by('model_name')
                    else:
                        mods = CarModel.objects.filter(pk=key).order_by('model_name')

                    for mod in mods:
                        arr.append(mod.myself())
                    response['message'] = arr
                    response['status_code'] = 200

                elif module == 'ModelsByCar':
                    arr = []
                    car_key = data.get('key')
                    car = Car.objects.get(pk=car_key)
                    car_models = CarModel.objects.filter(car=car)
                    for car_model in car_models:
                        arr.append(car_model.myself())
                    response['message'] = arr
                    response['status_code'] = 200

                elif module == 'CarSpecification':
                    car_mod = data.get('car_model')
                    part = data.get('part', '*')
                    if part == '*':
                        css = CarSpecification.objects.filter(car_model_id=car_mod).order_by('specification_name')
                    else:
                        css = CarSpecification.objects.filter(car_model_id=car_mod, part=part).order_by(
                            'specification_name')

                    for cs in css:
                        arr.append(cs.obj())
                    response['message'] = arr
                    response['status_code'] = 200

                elif module == 'PrintProformaInvoice':
                    proforma_key = data.get('key')
                    proforma = ProformaInvoice.objects.get(pk=proforma_key)
                    from fpdf import FPDF

                    class PDF(FPDF):
                        def header(self):
                            # Set up a border around the page
                            self.set_line_width(0.5)
                            self.rect(5.0, 5.0, 200.0, 287.0)

                        def footer(self):
                            # Go to 1.5 cm from bottom
                            self.set_y(-15)
                            # Select Arial italic 8
                            self.set_font('Arial', 'I', 8)
                            # Page number
                            self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

                    # Create instance of FPDF class
                    pdf = PDF(unit='mm')
                    page_with = 95 * 2

                    # Add a page
                    pdf.add_page()
                    # Add an image
                    pdf.image('static/general/img/sneda_motors_logo.png', x=10, y=6, w=30)
                    pdf.set_font('Arial', '', 6)
                    pdf.cell(0, 3, '+233 244 313 960 | +233 205 046 431', 0, 1, 'R')
                    pdf.cell(0, 3, 'P.O Box GP 3471, Accra, Ghana', 0, 1, 'R')
                    pdf.cell(0, 3, 'Plot No. 82A, Spintex Road', 0, 1, 'R')
                    pdf.cell(0, 3, 'www.snedamotors.com', 0, 1, 'R')
                    # Set font
                    pdf.set_font('Arial', 'B', 12)

                    # Add a cell# Set fill color to black
                    pdf.ln(10)
                    text_bg_color = (230, 230, 230)
                    pdf.set_fill_color(*text_bg_color)
                    pdf.cell(page_with / 2, 10, 'PROFORMA INVOICE', 0, 0, 'L', fill=True)
                    pdf.cell(page_with / 2, 10, f'Entry N0. : {proforma.pk}', 0, 1, 'R', fill=True)
                    pdf.ln(5)
                    text_bg_color = (255, 0, 0)

                    # header
                    #customer
                    pdf.set_font('Arial', 'B', 8)
                    pdf.cell(20, 5, f"Company", 1, 0, 'L')
                    pdf.set_font('Arial', '', 8)
                    pdf.cell(60, 5, f"{proforma.customer.company}", 1, 0, 'L')
                    pdf.set_font('Arial', 'B', 8)
                    pdf.cell(20, 5, "Attn. By", 1, 0, 'L')
                    pdf.set_font('Arial', '', 8)
                    pdf.cell(40, 5, f"{proforma.created_by.get_full_name()}", 1, 0, 'C')
                    pdf.set_font('Arial', 'B', 8)
                    pdf.cell(20, 5, f"QTY", 1, 0, 'L')
                    pdf.set_font('Arial', '', 8)
                    pdf.cell(30, 5, f"{format_currency(proforma.quantity)}", 1, 1, 'R')
                    #rep
                    pdf.set_font('Arial', 'B', 8)
                    pdf.cell(20, 5, "Personal", 1, 0, 'L')
                    pdf.set_font('Arial', '', 8)
                    pdf.cell(60, 5, f"{proforma.customer.name}", 1, 0, 'L')
                    pdf.set_font('Arial', 'B', 8)
                    pdf.cell(20, 5, "Payment", 1, 0, 'L')
                    pdf.set_font('Arial', '', 8)
                    pdf.cell(40, 5, "CASH", 1, 0, 'C')
                    pdf.set_font('Arial', 'B', 8)
                    pdf.cell(20, 5, f"Price", 1, 0, 'L')
                    pdf.set_font('Arial', '', 8)
                    pdf.cell(30, 5, f"{proforma.currency} {format_currency(proforma.price)}", 1, 1, 'R')

                    #currency
                    pdf.set_font('Arial', 'B', 8)
                    pdf.cell(20, 5, "Telephone", 1, 0, 'L')
                    pdf.set_font('Arial', '', 8)
                    pdf.cell(60, 5, f"{proforma.customer.mobile}", 1, 0, 'L')
                    pdf.set_font('Arial', 'B', 8)
                    pdf.cell(20, 5, "Currency", 1, 0, 'L')
                    pdf.set_font('Arial', '', 8)
                    pdf.cell(40, 5, f"{proforma.currency}", 1, 0, 'C')
                    pdf.set_font('Arial', 'B', 8)
                    pdf.cell(20, 5, f"Net", 1, 0, 'L')
                    pdf.set_font('Arial', '', 8)
                    pdf.cell(30, 5, f"{proforma.currency} {format_currency(proforma.net_amount)}", 1, 1, 'R')

                    # contact
                    pdf.set_font('Arial', 'B', 8)
                    pdf.cell(20, 5, "Email", 1, 0, 'L')
                    pdf.set_font('Arial', '', 8)
                    pdf.cell(60, 5, f"{proforma.customer.email}", 1, 0, 'L')
                    pdf.set_font('Arial', 'B', 8)
                    pdf.cell(20, 5, "Address", 1, 0, 'L')
                    pdf.set_font('Arial', '', 8)
                    pdf.cell(40, 5, f"{proforma.customer.address}", 1, 0, 'C')
                    pdf.set_font('Arial', 'B', 8)
                    pdf.cell(20, 5, f"Tax", 1, 0, 'L')
                    pdf.set_font('Arial', '', 8)
                    pdf.cell(30, 5, f"{proforma.currency} {format_currency(proforma.tax_amount)}", 1, 1, 'R')

                    ## car
                    # manufacturer
                    pdf.set_font('Arial', 'B', 8)
                    pdf.cell(20, 5, "Asset", 1, 0, 'L')
                    pdf.set_font('Arial', '', 8)
                    pdf.cell(120, 5,
                             f"{proforma.car_model.car.manufacturer.name} / {proforma.car_model.model_name} / {proforma.car_model.year} / {proforma.chassis}",
                             1, 0, 'L')
                    pdf.set_font('Arial', 'B', 8)
                    pdf.cell(20, 5, f"Payable", 1, 0, 'L')
                    pdf.set_font('Arial', '', 8)
                    pdf.cell(30, 5, f"{proforma.currency} {format_currency(proforma.taxable_amount)}", 1, 1, 'R')

                    # prices

                    pdf.ln(10)
                    #technocal
                    techs = ProformaInvoiceSpec.objects.filter(part='tech', proforma=proforma)
                    pdf.set_font('Arial', 'B', 8)
                    pdf.cell(0, 5, "Technical Features", 0, 1)
                    for tech in techs:
                        pdf.set_font('Arial', 'B', 7)
                        pdf.cell(40, 4, f"{tech.specification_name}", 0, 0)
                        pdf.set_font('Arial', "", 7)
                        pdf.cell(149, 4, f"{tech.specification_value}", 0, 1)

                    pdf.ln(5)
                    # interir
                    intts = ProformaInvoiceSpec.objects.filter(part='int', proforma=proforma)
                    pdf.set_font('Arial', 'B', 8)
                    pdf.cell(0, 5, "Interior", 0, 1)
                    for intt in intts:
                        pdf.set_font('Arial', 'B', 7)
                        pdf.cell(40, 4, f"{intt.specification_name}", 0, )
                        pdf.set_font('Arial', "", 7)
                        pdf.cell(150, 4, f"{intt.specification_value}", 0, 1)

                    pdf.ln(5)
                    # exteriors
                    exts = ProformaInvoiceSpec.objects.filter(part='ext', proforma=proforma)
                    pdf.set_font('Arial', 'B', 8)
                    pdf.cell(0, 5, "Exterior", 0, 1)
                    for ext in exts:
                        pdf.set_font('Arial', 'B', 7)
                        pdf.cell(40, 4, f"{ext.specification_name}", 0, 0)
                        pdf.set_font('Arial', "", 7)
                        pdf.cell(140, 4, f"{ext.specification_value}", 0, 1)

                    pdf.ln(5)
                    pdf.set_font('Arial', "", 8)
                    pdf.set_text_color(0, 0, 255)
                    pdf.multi_cell(0, 5, "For your purchase, full payment is required at the time of order "
                                         "confirmation. We accept Ghanaian Cedis (GHS) at the prevailing market "
                                         "exchange rate on the date of payment. You can find the current exchange rate "
                                         "online or through your bank. This quotation is valid for a period of 8 to 12 "
                                         "weeks from the date of issuance. Prices may be subject to change after this "
                                         "period. We offer a fast and efficient delivery service. You can expect to "
                                         "receive your purchase within 5 working days after your payment is confirmed.")
                    pdf.set_text_color(0, 0, 0)  # Black (0, 0, 0)

                    pdf.ln(10)
                    pdf.cell(30, 10, f'{proforma.created_by.get_full_name()}', 1, 0, 'C')
                    pdf.cell(130, 10, '', 0, 0)
                    pdf.cell(30, 10, f'{proforma.approver()}', 1, 1, 'C')
                    pdf.cell(30, 10, 'PREPARED BY', 0, 0, 'C')
                    pdf.cell(130, 10, '', 0, 0)
                    pdf.cell(30, 10, 'CONFIRMED BY', 0, 1, 'C')
                    # Output the PDF to 10 file
                    file_name = f"static/uploads/attachments/{proforma.customer.name}_{proforma.car_model.model_name}.pdf"
                    pdf.output(file_name)
                    response['message'] = file_name
                    response['status_code'] = 200

            except Exception as e:
                response['status'] = 'error'
                response['status_code'] = 505
                # Get the line number where the exception occurred
                line_number = traceback.extract_tb(e.__traceback__)[-1].lineno

                # Set the message with the line number
                response['message'] = f"Error occurred at line {line_number}: {str(e)}"

        elif method == 'PUT':
            try:
                response['status_code'] = 200
                response['message'] = "Creation Successful"
                if module == 'stock':
                    stage = data.get('stage')
                    if stage == 'hd':
                        loc = data.get('loc')
                        remark = data.get('remark')

                        # check if there is open hd
                        if StockCountHD.objects.filter(status=1).exists():
                            response['message'] = "There is an open stock take"
                            response['status_code'] = 505
                            response['status'] = 'WARNING'

                        else:
                            # open stock
                            StockCountHD(loc=loc, remark=remark).save()
                            response['status_code'] = 200
                            response['status'] = 'success'
                    elif stage == 'tran':
                        item_ref = data.get('item_ref')
                        qty = data.get('qty')
                        myName = data.get('user')

                        server = f"{DB_SERVER},{DB_PORT}"
                        database = DB_NAME
                        username = DB_USER
                        password = DB_PASSWORD
                        driver = '{ODBC Driver 17 for SQL Server}'  # Change this to the driver you're using
                        connection_string = f"DRIVER={driver};SERVER={server};DATABASE={database};UID={username};PWD={password}"
                        connection = pyodbc.connect(connection_string)
                        cursor = connection.cursor()

                        query = f"select barcode,item_ref,item_des1,sell_price from product_master where item_ref = '{item_ref}'"
                        cursor.execute(query)

                        row = cursor.fetchone()
                        if row:
                            barcode = row[0].strip()
                            name = row[2].strip()
                            if row[3] is None:
                                sell_price = 0.00
                            else:
                                sell_price = row[3]

                            val = Decimal(qty) * Decimal(sell_price)
                            hd = StockCountHD.objects.get(status=1)
                            StockCountTrans(stock_count_hd=hd, item_ref=item_ref, barcode=barcode, name=name,
                                            sell_price=sell_price, quantity=qty, owner=myName, value=val).save()

                            response['status_code'] = 200
                            response['status'] = 'success'
                            response['message'] = f"ITEM {name} Added"
                            cursor.close()

                        else:
                            response['status_code'] = 404
                            response['status'] = 'error'
                            response['message'] = f"could not find item with ref {item_ref}"
                    elif stage == 'comment':
                        comment_pk = data.get('comment_pk')
                        comment = data.get('comment')
                        issue = data.get('issue')

                        if StockCountTrans.objects.filter(pk=comment_pk).count() == 1:
                            # update
                            coun = StockCountTrans.objects.get(pk=comment_pk)
                            coun.comment = comment
                            coun.issue = issue
                            coun.save()
                            response['message'] = "Comment Updated"
                        else:
                            response['message'] = f"No count with pk as {comment_pk}"
                    elif stage == 'save_frozen':

                        header = data.get('header')
                        trans = data.get('trans')

                        if len(trans) > 0:

                            try:
                                loc_id = header.get('loc')
                                frozen_ref = header.get('frozen_ref')
                                remarks = header.get('remarks')
                                us_pk = header.get('owner')

                                owner = get_object_or_404(User, pk=us_pk)
                                stock_freeze_hd = StockFreezeHd.objects.create(loc_id=loc_id, ref=frozen_ref,
                                                                               remarks=remarks, owner=owner)

                                try:
                                    for tran in trans:
                                        ref = tran['ref']
                                        barcode = tran['barcode']
                                        qty = tran['qty']
                                        name = tran['name']
                                        price = tran['price']

                                        StockFreezeTrans.objects.create(entry_id=stock_freeze_hd.pk, item_ref=ref,
                                                                        barcode=barcode, qty=qty, name=name,
                                                                        price=price)

                                    response['message'] = "SAVED"
                                    response['status_code'] = 200
                                except Exception as e:
                                    stock_freeze_hd.delete()
                                    line_number = traceback.extract_tb(e.__traceback__)[-1].lineno
                                    response['message'] = f"Error occurred at line {line_number}: {str(e)}"

                            except Exception as e:
                                line_number = traceback.extract_tb(e.__traceback__)[-1].lineno
                                response['message'] = f"Error occurred at line {line_number}: {str(e)}"


                        else:

                            response['message'] = f" {len(trans)} Transactions cannot be empty"
                            response['status'] = "EMPTY"
                            response['status_code'] = 100
                    elif stage == 'save_cont':
                        header = data.get('header')
                        trans = data.get('trans')
                        count_pk = header.get('count_pk')
                        task = header.get('task')

                        frozen_ref = header.get('ref')
                        comment = header.get('comment')
                        owner_pk = header.get('owner_pk')
                        owner = User.objects.get(pk=owner_pk)

                        if StockFreezeHd.objects.filter(pk=frozen_ref).count() == 1:
                            try:

                                frozen = StockFreezeHd.objects.get(pk=frozen_ref)

                                if StockCountHD.objects.filter(frozen=frozen, comment=comment, owner=owner).exists():
                                    # delete
                                    StockCountHD.objects.get(frozen=frozen, comment=comment, owner=owner).delete()
                                StockCountHD(frozen=frozen, comment=comment, owner=owner).save()
                                count_hd = StockCountHD.objects.get(frozen=frozen, comment=comment, owner=owner)

                                # save trans
                                for tran in trans:
                                    print(tran)
                                    ref = tran.get('ref')
                                    barcode = tran.get('barcode')
                                    name = tran.get('name')
                                    frozen_qty = tran.get('frozen')
                                    counted = tran.get('counted')
                                    diff = tran.get('diff')
                                    row_comment = tran.get('row_comment')
                                    row_iss = tran.get('row_iss')

                                    value = tran.get('price')

                                    print("#DIFF * VAL#", barcode)
                                    print('## ', diff, value)
                                    diff_val = Decimal(diff) * Decimal(value)

                                    print("#FROZEN * VAL#", barcode)
                                    print('## ', frozen_qty, value)
                                    froze_val = Decimal(frozen_qty) * Decimal(value)

                                    print("#DIFF * VAL#", barcode)
                                    print('## ', counted, value)
                                    counted_val = Decimal(counted) * Decimal(value)

                                    # save tran
                                    StockCountTrans.objects.create(stock_count_hd=count_hd, item_ref=ref,
                                                                   barcode=barcode,
                                                                   name=name, froze_qty=frozen_qty, counted_qty=counted,
                                                                   diff_qty=diff, comment=row_comment, issue=row_iss,
                                                                   froze_val=froze_val, diff_val=diff_val,
                                                                   counted_val=counted_val)

                                frozen.status = 2
                                frozen.save()
                                response['status_code'] = 200
                                response['status'] = 'success'
                                response['message'] = "Stock Count Saved"



                            except Exception as e:
                                response['status_code'] = 500
                                response['status'] = 'error'
                                line_number = traceback.extract_tb(e.__traceback__)[-1].lineno
                                response['message'] = f"Error occurred at line {line_number}:  - {e}"



                        else:
                            response['statys_code'] = 404
                            response['message'] = f"CANNOT FIND FROZEN with entry {frozen_ref}"

                elif module == 'sales_deal':
                    sales_rep = data.get('sales_rep')
                    sales_cust = data.get('sales_customer')

                    pur_rep_name = data.get('pur_rep_name')
                    pur_rep_email = data.get('pur_rep_email')
                    pur_rep_phone = data.get('pur_rep_phone')

                    asset = data.get('asset')
                    model = data.get('model')

                    stock_type = data.get('stock_type')
                    requirement = data.get('requirement')

                    # validate
                    try:
                        customer = SalesCustomers.objects.get(pk=sales_cust)
                        owner = User.objects.get(pk=sales_rep)

                        SalesDeals(customer=customer, owner=owner, pur_rep_name=pur_rep_name,
                                   pur_rep_phone=pur_rep_phone,
                                   pur_rep_email=pur_rep_email, asset=asset, model=model,
                                   stock_type=stock_type, requirement=requirement).save()

                        # save in logs
                        desc = (f"Initiated discussion with {pur_rep_name} from {customer.company} regarding car "
                                f"purchase. Asset presented is"
                                f"{asset} with model of {model}. He has a requirement which is {requirement}")
                        Logs(owner=owner, description=desc, phone=pur_rep_phone, flag='success', subject="Car Purchase",
                             customer=pur_rep_name).save()

                        response['status_code'] = 200
                        response['message'] = "DEAL CREATED"
                    except Exception as e:
                        response['status_code'] = 505
                        response['message'] = str(e)

                elif module == 'deal_transaction':
                    dealkey = data.get('deal')
                    title = data.get('title')
                    detail = data.get('detail')
                    ownerkey = data.get('owner')

                    try:
                        deal = SalesDeals.objects.get(pk=dealkey)
                        owner = User.objects.get(pk=ownerkey)
                        DealTransactions(deal=deal, title=title, details=detail, owner=owner).save()

                        response['status_code'] = 200
                        response['message'] = "TRANSACTION ADDED"
                    except Exception as e:
                        response['status_code'] = 500
                        response['message'] = str(e)

                elif module == 'sales_customer':
                    try:
                        type_of_client = data.get('type_of_client')
                        sector_of_company = data.get('sector_of_company')
                        company = data.get('company')
                        region = data.get('region')
                        address = data.get('address')
                        city = data.get('city')
                        phone = data.get('phone')
                        email = data.get('email')
                        fax = data.get('fax')

                        position = data.get('position')
                        first_name = data.get('first_name')
                        last_name = data.get('last_name')
                        note = data.get('note')

                        url = data.get('url')
                        owner_pk = data.get('owner')
                        # save
                        owner = User.objects.get(pk=owner_pk)
                        reg = GeoCity.objects.get(pk=region)
                        cit = GeoCitySub.objects.get(pk=city)

                        SalesCustomers(sector_of_company=sector_of_company, type_of_client=type_of_client,
                                       company=company,
                                       region=reg, city=cit,
                                       mobile=phone, email=email, fax=fax, address=address, first_name=first_name,
                                       last_name=last_name, position=position,
                                       note=note, name=f"{first_name} {last_name}", url=url, owner=owner).save()
                        response['status'] = 200
                        response['message'] = "CUSTOMER SAVED"
                    except Exception as e:
                        response['status'] = 505
                        response['message'] = str(e)

                elif module == 'asset_group':
                    try:
                        g_name = data.get('g_name')
                        o_pk = data.get('mypk')

                        owner = User.objects.get(pk=o_pk)
                        SalesAssetsGroup(name=g_name, owner=owner).save()
                        response['status_code'] = 200
                        response['message'] = "Asset Group Saved"
                    except Exception as e:
                        response['status_code'] = 500
                        response['message'] = str(e)

                elif module == 'CarOrigin':

                    try:
                        country = data.get('country')
                        mypk = data.get('mypk')

                        owner = User.objects.get(pk=mypk)
                        CarOrigin(created_by=owner, country=country).save()
                        response['status_code'] = 200
                        response['message'] = f"Origin created successfully"
                    except Exception as e:
                        response['status_code'] = 505
                        response['message'] = str(e)

                elif module == 'CarManufacturer':
                    name = data.get('name')
                    o_key = data.get('o_key')
                    origin = CarOrigin.objects.get(pk=o_key)
                    mypk = data.get('mypk')
                    owner = User.objects.get(pk=mypk)
                    CarManufacturer(name=name, origin=origin, created_by=owner).save()

                elif module == 'CarModel':
                    model_name = data.get('model_name')
                    car_key = data.get('car')
                    mypk = data.get('mypk')
                    year = data.get('year')
                    price = data.get('price')
                    car = Car.objects.get(pk=car_key)
                    created_by = User.objects.get(pk=mypk)

                    CarModel(model_name=model_name, car=car, created_by=created_by, year=year, price=price).save()


                elif module == 'CarSupplier':
                    o_key = data.get('o_key')
                    origin = CarOrigin.objects.get(pk=o_key)
                    mypk = data.get('mypk')
                    created_by = User.objects.get(pk=mypk)
                    name = data.get('name')
                    email = data.get('email')
                    phone = data.get('phone')
                    website = data.get('website')

                    CarSupplier(origin=origin, created_by=created_by, name=name, email=email, phone=phone,
                                website=website).save()

                elif module == 'CarSpecification':
                    mypk = data.get('mypk')
                    part = data.get('part')
                    created_by = User.objects.get(pk=mypk)
                    mod = data.get('model')
                    car_model = CarModel.objects.get(pk=mod)
                    name = data.get('name')
                    value = data.get('value')
                    CarSpecification(car_model=car_model, specification_name=name, specification_value=value,
                                     created_by=created_by, part=part).save()

                elif module == 'Car':
                    name = data.get('name')
                    supp_key = data.get('supplier')
                    supplier = CarSupplier.objects.get(pk=supp_key)
                    manf_key = data.get('manufacturer')
                    manufacturer = CarManufacturer.objects.get(pk=manf_key)
                    mypk = data.get('mypk')
                    created_by = User.objects.get(pk=mypk)

                    Car(name=name, supplier=supplier, created_by=created_by, manufacturer=manufacturer).save()

                elif module == 'ProformaInvoice':
                    model_pk = data.get('model')
                    car_model = CarModel.objects.get(pk=model_pk)
                    customer_pk = data.get('customer')
                    customer = SalesCustomers.objects.get(pk=customer_pk)
                    mypk = data.get('mypk')
                    created_by = User.objects.get(pk=mypk)

                    currency = data.get('currency')

                    taxable = data.get('taxable')
                    chassis = data.get('chassis')
                    price = car_model.price

                    quantity = data.get('quantity')
                    taxable_amount = Decimal(quantity) * Decimal(price)
                    if taxable == 'YES':
                        tax_amount = Decimal(taxable_amount) * Decimal(0.219)
                    else:
                        tax_amount = 0
                    net_amount = taxable_amount - tax_amount

                    if ProformaInvoice.objects.filter(car_model=car_model, customer=customer).count() == 0:

                        p_uni = make_md5_hash(f"{customer_pk}{car_model.model_name}")
                        # save proforma
                        ProformaInvoice(
                            uni=p_uni,
                            chassis=chassis,
                            car_model=car_model,
                            customer=customer,
                            created_by=created_by,
                            currency=currency,
                            price=price,
                            quantity=quantity,
                            taxable_amount=taxable_amount,
                            tax_amount=tax_amount,
                            net_amount=net_amount
                        ).save()

                        # get proforma
                        proforma = ProformaInvoice.objects.get(
                            uni=p_uni
                        )

                        # get car model specifications
                        model_specifications = CarSpecification.objects.filter(car_model=car_model)
                        for spec in model_specifications:
                            part = spec.part
                            specification_name = spec.specification_name
                            specification_value = spec.specification_value

                            ProformaInvoiceSpec(
                                proforma=proforma,
                                created_by=created_by,
                                part=part,
                                specification_name=specification_name,
                                specification_value=specification_value
                            ).save()

                        # response
                        response['status_code'] = 200
                        response['message'] = "Profoma Created"
                    else:
                        response['status_code'] = 505
                        response['message'] = "There is a pending PO for same customer and car that is opened"

                else:
                    response['status_code'] = 404
                    response['message'] = "Invalid Module"
            except Exception as e:
                response['status_code'] = 505
                response['message'] = str(e)
        elif method == 'PATCH':
            try:
                if module == 'stock':
                    stage = data.get('stage')
                    task = data.get('task')
                    if stage == 'hd':
                        if task == 'close':
                            opened = StockCountHD.objects.filter(status=1).last()
                            opened.status = 2
                            opened.save()
                            response = {'status': 200, 'status_code': 'success', 'message': "OPENED COUNT closed"}

                    if stage == 'save_cont':
                        header = data.get('header')
                        count_pk = header.get('count_pk')

                        # print(data)
                        if StockCountHD.objects.filter(pk=count_pk).count() == 1:
                            c_hd = StockCountHD.objects.get(pk=count_pk)
                            c_hd.comment = header.get('comment')

                            trans = data.get('trans')

                            # delete all trans
                            # delete this tran and add again

                            for tran in trans:

                                print(tran)
                                ref = tran.get('ref')
                                barcode = tran.get('barcode').strip()
                                name = tran.get('name')
                                frozen = tran.get('frozen')
                                counted = str(tran.get('counted'))
                                diff = tran.get('diff')
                                row_comment = tran.get('row_comment')
                                row_iss = tran.get('row_iss')
                                StockCountTrans.objects.filter(stock_count_hd=c_hd, item_ref=ref).delete()
                                cursor = db()
                                q = f"SELECT sell_price FROM product_master where barcode = '{barcode}'"

                                if cursor.execute(q).fetchone() is None:
                                    value = 0.00
                                else:
                                    value = cursor.execute(q).fetchone()[0]

                                diff_val = Decimal(value) * Decimal(diff)
                                froze_val = Decimal(value) * Decimal(frozen)
                                counted_val = Decimal(value) * Decimal(counted)

                                cursor.close()

                                # save tran
                                StockCountTrans.objects.create(stock_count_hd=c_hd, item_ref=ref, barcode=barcode,
                                                               name=name, froze_qty=frozen, counted_qty=counted,
                                                               diff_qty=diff, comment=row_comment, issue=row_iss,
                                                               diff_val=diff_val, froze_val=froze_val,
                                                               counted_val=counted_val, sell_price=value)
                            c_hd.save()
                            response['message'] = "UPDATE SUCCESSFUL"
                        else:
                            var = response['status_code'] = 404
                            response['message'] = f"CANNOT FIND DOCUMENT with key {count_pk}"

                elif module == 'approve':
                    doc = data.get('doc')
                    key = data.get('key')

                    if doc == 'FR':
                        hd = StockFreezeHd.objects.filter(pk=key)
                        count = hd.count()
                    elif doc == 'STK':
                        # approved counted stock
                        hd = StockCountHD.objects.filter(pk=key)
                        count = hd.count()
                    else:
                        count = 0

                    if count == 1:
                        try:
                            head = hd.last()
                            head.approve = 1
                            head.save()

                            response['status_code'] = 200
                            response['status'] = 'success'
                            response['message'] = f"APPROVED {doc} = {key}"
                        except Exception as e:
                            response['status_code'] = 505
                            response['status'] = 'error'
                            response['message'] = str(e)

                    else:
                        response['status_code'] = 404
                        response['status'] = 'error'
                        response['message'] = f"Count not find matching document ({doc} - {key})"

                elif module == 'close':
                    doc = data.get('doc')

                    if doc == 'deal':
                        owner = data.get('closer')
                        finale = data.get('finale')
                        closer = User.objects.get(pk=owner)
                        key = data.get('key')
                        deal = SalesDeals.objects.get(pk=key)
                        message = data.get('message')

                        # update transactions
                        DealTransactions(title='CLOSED', details=message, deal=deal, owner=closer).save()

                        deal.status = 1
                        deal.finale = finale
                        deal.save()

                        response['message'] = 'DEAL CLOSED'

                elif module == 'ModelPriceChange':
                    model_key = data.get('model_key')
                    mod = CarModel.objects.get(pk=model_key)
                    model_name = mod.model_name
                    mod.price = data.get('new_price')
                    mod.save()
                    response['status_code'] = 200
                    response['message'] = f"Price Changed For {model_name}"

                elif module == 'CarSpecifications':
                    pk = data.get('pk')
                    key = data.get('key')
                    value = data.get('value')
                    part = data.get('part')

                    # get spec
                    spec = CarSpecification.objects.get(pk=pk)

                    # set new values
                    spec.part = part
                    spec.specification_name = key
                    spec.specification_value = value

                    # save spec
                    spec.save()

                    # response
                    response['status_code'] = 200
                    response['message'] = f"Spec Updated Successfully"
                elif module == 'UpdateProformaInvoiceSpec':
                    pk = data.get('key')
                    mypk = data.get('mypk')
                    created_by = User.objects.get(pk=mypk)
                    proforma = ProformaInvoice.objects.get(pk=pk)
                    # delete specs
                    ProformaInvoiceSpec.objects.filter(proforma=proforma).delete()

                    # insert new
                    car_model = proforma.car_model
                    model_specifications = CarSpecification.objects.filter(car_model=car_model)
                    for spec in model_specifications:
                        spec_name = spec.specification_name
                        spec_value = spec.specification_value
                        spec_part = spec.part
                        # insert into new specs
                        ProformaInvoiceSpec(
                            proforma=proforma,
                            part=spec_part,
                            specification_name=spec_name,
                            specification_value=spec_value,
                            created_by=created_by
                        ).save()
                    response['status_code'] = 200
                    response['message'] = "Proforma Specifications Updated"
                elif module == 'request_proforma_approval':
                    pk = data.get('key')
                    mypk = data.get('mypk')
                    proforma = ProformaInvoice.objects.get(pk=pk)
                    pf_keep = proforma
                    request_by = User.objects.get(pk=mypk)
                    mail_sender = MailSenders.objects.get(purpose='sales_proforma')
                    if proforma.approval_request:
                        raise Exception("Approval Sent Already, Please follow up with Approves")
                    # flag send

                    # send notification
                    subject = "Proforma Approval"
                    msg = f"There is an approval request for a proforma document from {request_by.get_full_name()}. Details is as below"
                    msg += f"<p><strong>Customer: </strong>{pf_keep.customer.name}</p>"
                    msg += f"<p><strong>Asset: </strong>{pf_keep.car_model.car.manufacturer.name} / {pf_keep.car_model.model_name} / {pf_keep.car_model.year} / {pf_keep.chassis}</p>"
                    msg += f"<p><strong>Price: </strong>{pf_keep.currency} {format_currency(pf_keep.price)}</p>"
                    msg += f"<p><strong>Quantity: </strong>{format_currency(pf_keep.quantity)}</p>"
                    msg += f"<p><strong>Taxable Amount: </strong>{pf_keep.currency} {format_currency(pf_keep.net_amount)}</p>"
                    msg += f"<p><strong>Tax Amount: </strong>{pf_keep.currency} {format_currency(pf_keep.tax_amount)}</p>"
                    msg += f"<p><strong>Amount Payable: </strong>{pf_keep.currency} {format_currency(pf_keep.taxable_amount)}</p>"

                    msg += f"<a href='#'>Approve</a>"

                    # que email

                    proforma.approval_request = True
                    proforma.save()
                    # insert transactions
                    ProformaTransactions(
                        proforma=proforma,
                        task='Approval Request',
                        created_by=request_by
                    ).save()

                    MailQueues(
                        sender=mail_sender,
                        recipient=CMMS_PROF_APPROVER,
                        subject=subject,
                        body=msg
                    ).save()

                    response['status_code'] = 200
                    response['message'] = "Approval Request Sent!!"
                elif module == 'approve_proforma':
                    mypk = data.get('mypk')
                    approved_by = User.objects.get(pk=mypk)
                    prof_pk = data.get('key')
                    proforma = ProformaInvoice.objects.get(pk=prof_pk)
                    proforma.is_approved = True
                    proforma.approved_by = approved_by
                    proforma.save()

                    ProformaTransactions(
                        proforma=proforma,
                        task='Approval Request',
                        created_by=approved_by
                    ).save()
                    response['status_code'] = 200
                    response['message'] = "Document Approved"
                elif module == 'send_proforma':
                    prof_key = data.get('key')
                    mypk = data.get('mypk')
                    proforma = ProformaInvoice.objects.get(pk=prof_key)
                    sent_by = User.objects.get(pk=mypk)

                    # create email
                    mail_sender = MailSenders.objects.get(purpose='sales_proforma')
                    msg = ""
                    if proforma.is_sent == False:
                        mob = MailQueues.objects.get_or_create(
                            sender=mail_sender,
                            recipient=proforma.customer.email,
                            subject=f'Proforma Invoice for {proforma.car_model.model_name}',
                            body=msg
                        )

                        ## add attachment
                        MailAttachments(
                            mail=mob,
                            attachment=f'static/uploads/attachments/{proforma.customer.name}_{proforma.car_model.model_name}.pdf'
                        ).save()

                        proforma.is_sent = True
                        proforma.sent_by = sent_by
                        proforma.save()

            except Exception as e:
                response['status_code'] = 500
                response['message'] = str(e)
        elif method == 'DELETE':
            module = body.get('module') or 'before'
            doc = data.get('doc')
            key = data.get('key')
            if module == 'before':
                if doc == 'FR':
                    hd = StockFreezeHd.objects.filter(pk=key)
                    count = hd.count()
                elif doc == 'STK':
                    # approved counted stock
                    hd = StockCountHD.objects.filter(pk=key)
                    count = hd.count()
                else:
                    count = 0

                if count == 1:
                    try:
                        head = hd.last()
                        head.delete()

                        response['status_code'] = 200
                        response['status'] = 'success'
                        response['message'] = f"DELETED {doc} = {key}"
                    except Exception as e:
                        response['status_code'] = 505
                        response['status'] = 'error'
                        response['message'] = str(e)

                else:
                    response['status_code'] = 404
                    response['status'] = 'error'
                    response['message'] = f"Count not find matching document ({doc} - {key})"

            elif module == 'sales_customer':
                customer = SalesCustomers.objects.get(pk=data.get('key'))
                name = customer.name
                customer.delete()
                response['message'] = f"{name} deleted"

            elif module == 'CarSpecifications':
                pk = data.get('pk')
                CarSpecification.objects.get(pk=pk).delete()

                response['status_code'] = 200
                response['message'] = "Spec Deleted"

    except json.JSONDecodeError as e:
        response["status_code"] = 400
        response["status"] = "Bad Request"
        response["message"] = f"Error decoding JSON: {str(e)}"

    return JsonResponse(response)
