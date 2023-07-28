import json
import traceback
from datetime import date

from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from fpdf import FPDF

from ocean.settings import DB_SERVER, DB_NAME, DB_USER, DB_PORT, DB_PASSWORD
from django.contrib.auth import get_user_model
import pyodbc
from cmms.models import *
from decimal import Decimal
from django.contrib import messages

from cmms.extra import db


# Create your views here.
def base(request):
    return None


@login_required(login_url='/login/')
def carjobs(request):
    page = {
        'nav': True,
        'title': "Car Jobs"
    }

    context = {
        'page': page,
        'nav': True,
        'searchButton': 'carJob'
    }
    messages.info(request, "You can now view all data but limited to 100 records")
    return render(request, 'cmms/car-jobs.html', context=context)


@login_required()
def tools(request):
    context = {
        'nav': True
    }
    return render(request, 'cmms/tools.html', context=context)


@login_required()
def stock(request):
    opened = False
    if StockCountHD.objects.filter(status=1).count() == 1:
        opened = True
    return render(request, 'cmms/stock.html',
                  context={'nav': True, 'stocks': StockCountHD.objects.all(), 'open': opened})


@login_required()
def new_frozen(request):
    context = {'nav': True}
    return render(request, 'cmms/new_stock.html', context=context)


def new_count(request, frozen):
    context = {'nav': True}
    return render(request, 'cmms/count.html', context=context)


@login_required()
def forezen(request):
    froze = StockFreezeHd.objects.all()
    if froze.count() < 1:
        messages.warning(request, "PLEASE FREEZE NEW STOCK")
        return redirect('/cmms/stock/frozen/new/')
    else:
        f_pk = froze.last().pk
    context = {'nav': True, 'f_pk': f_pk}
    return render(request, 'cmms/frozen.html', context=context)


@login_required()
def new_stock_count(request):
    # if StockCountHD.objects.filter(status=1).count() != 1:
    #     messages.error(request, 'NO OPEN STOCK COUNT.')
    #     return redirect('/cmms/stock/')
    context = {
        'nav': True,
        'page': {
            'title': "STOCK COUNT"
        }
    }
    return render(request, 'cmms/count.html', context=context)


@login_required()
def view_stock_count(request):
    if StockCountHD.objects.filter(status=1).count() < 1:
        messages.error(request, 'NO OPEN STOCK COUNT.')
        return redirect('/cmms/stock/new/')
    context = {
        'nav': True,
        'page': {
            'title': "STOCK COUNT VIEW"
        },
        'c_pk': StockCountHD.objects.filter(status=1).last().pk
    }
    return render(request, 'cmms/count_view.html', context=context)


@login_required()
def edit_stock_count(request, pk):
    if StockCountHD.objects.filter(status=1).count() < 1:
        messages.error(request, 'NO OPEN STOCK COUNT.')
        return redirect('/cmms/stock/new/')
    context = {
        'nav': True,
        'page': {
            'title': "STOCK COUNT VIEW"
        },
        'c_pk': StockCountHD.objects.filter(status=1).last().pk
    }
    return render(request, 'cmms/count_edit.html', context=context)


@login_required()
def compare(request, pk, as_of, group):
    # get counts
    hd = StockCountHD.objects.get(pk=pk)
    wrong_entries = StockCountTrans.objects.filter(stock_count_hd=hd, issue__in='wr_entry').count()
    sys_error = StockCountTrans.objects.filter(stock_count_hd=hd, issue__in='sys_error').count()
    lost = StockCountTrans.objects.filter(stock_count_hd=hd, issue__in='lost').count()
    unknown = StockCountTrans.objects.filter(stock_count_hd=hd, issue__in='unknown').count()
    return render(request, 'cmms/compare.html', context={
        'nav': True,
        'as_of': as_of,
        'pk': pk,
        'group': group,
        'wrong_entries': wrong_entries,
        'sys_error': sys_error,
        'lost': lost,
        'unknown': unknown,

    })


@csrf_exempt
def api(request):
    global header
    method = request.method
    response = {"status_code": "", "status": "", "message": ""}

    try:
        body = json.loads(request.body)
        module = body.get('module')
        data = body.get('data')

        if method == 'VIEW':
            try:
                if module == 'product':
                    scope = data.get('range')
                    item_uni = data.get('item_uni')
                    db_col = data.get('db_col') or 'barcode'

                    if scope == 'single':
                        q = f"SELECT barcode,item_ref,item_des1 FROM product_master WHERE {db_col} = '{item_uni}'"
                    else:
                        q = "SELECT barcode,item_ref,item_des1  FROM product_master"

                    server = f"{DB_SERVER},{DB_PORT}"
                    database = DB_NAME
                    username = DB_USER
                    password = DB_PASSWORD
                    driver = '{ODBC Driver 17 for SQL Server}'  # Change this to the driver you're using
                    connection_string = f"DRIVER={driver};SERVER={server};DATABASE={database};UID={username};PWD={password}"
                    connection = pyodbc.connect(connection_string)
                    cursor = connection.cursor()

                    cursor.execute(q)
                    count = 0
                    arr = []
                    for part in cursor.fetchall():
                        count += 1
                        obj = {'barcode': part[0].strip(), 'item_ref': part[1].strip(), 'name': part[2].strip()}
                        arr.append(obj)

                    resp = {'count': count, 'trans': arr}
                    response['status_code'] = 200
                    response['status'] = 'success'
                    response['message'] = resp
                    cursor.close()
                    connection.close()

                elif module == 'groups':
                    cursor = db()
                    query = "select Group_code,group_des from group_master order by group_des"
                    cursor.execute(query)
                    rows = cursor.fetchall()
                    arr = []
                    grp_counts = 0
                    for row in rows:
                        grp_counts += 1
                        group_code = row[0].strip()
                        group_name = row[1].strip()
                        arr.append({'code': group_code, 'name': group_name})

                    response['message'] = {
                        'counts': grp_counts, 'groups': arr
                    }
                    response['status_code'] = 200
                    response['status'] = 'success'


                elif module == 'stock':
                    stage = data.get('stage')
                    if stage == 'active':
                        # check for active
                        if StockCountHD.objects.filter(status=1).exists():
                            active = StockCountHD.objects.get(status=1)
                            trans = StockCountTrans.objects.filter(stock_count_hd=active).order_by('-pk')
                            arr = []
                            for tran in trans:
                                obj = {
                                    'item_ref': tran.item_ref,
                                    'barcode': tran.barcode,
                                    'name': tran.name,
                                    'quantity': tran.quantity,
                                    'price': tran.sell_price,
                                    'value': tran.quantity * tran.sell_price,
                                    'owner': tran.owner,
                                    'comment': tran.comment
                                }
                                arr.append(obj)
                            response['message'] = {
                                'count': trans.count(), 'trans': arr
                            }
                            response['status_code'] = 200
                            response['status'] = 'success'
                        else:
                            response['message'] = 'NO DATA'
                            response['status_code'] = 404
                            response['status'] = 'error'

                    elif stage == 'preview':
                        style = data.get('compare')
                        if style == 'final_compare':
                            doc = data.get('doc')
                            tot_phy = Decimal(0.00)
                            tot_sys = Decimal(0.00)
                            tot_qdif = Decimal(0.00)
                            tot_vdif = Decimal(0.00)
                            if doc == 'excel':
                                import openpyxl
                                workbook = openpyxl.Workbook()
                                sheet = workbook.active
                                # make sheet head
                                sheet[f'A1'] = "ITEM REFERENCE"
                                sheet[f'B1'] = "BARCODED"
                                sheet[f'C1'] = "DESCRIPTION"
                                sheet[f'D1'] = "PHYSICAL QUANTITY"
                                sheet[f'E1'] = "SYSTEM QUANTITY"
                                sheet[f'F1'] = "QUANTITY DIFFERENCE"
                                sheet[f'G1'] = "SELLING PRICE"
                                sheet[f'H1'] = "VALUE DIFFERENCE"
                                sheet[f'I1'] = "COMMENT"

                            # check if there is open stock
                            pk = data.get('pk')
                            open_st = StockCountHD.objects.filter(pk=pk).count()
                            is_open = False
                            if open_st == 1:

                                stock_hd = StockCountHD.objects.get(pk=pk)
                                trans = StockCountTrans.objects.filter(stock_count_hd=stock_hd)
                                frozen = stock_hd.frozen
                                header = {
                                    'frozen': {
                                        'pk': frozen.pk,
                                        'loc': frozen.loc_id,
                                        'remarks': frozen.remarks,
                                        'status': frozen.status,
                                        'owner': frozen.owner.username,
                                        'created': f"{frozen.created_date} {frozen.created_time}",
                                        'entry': frozen.ent(),

                                    },
                                    'count': {
                                        'pk': stock_hd.pk,
                                        'entry': stock_hd.entry_no(),
                                        'created': stock_hd.created_at,
                                        'comment': stock_hd.comment,
                                        'status': stock_hd.status,
                                        'owner': stock_hd.owner.username,
                                        'next': stock_hd.next(),
                                        'prev': stock_hd.prev()
                                    }
                                }
                                arr = []
                                not_arr = []
                                entries = {
                                    'wr_entry': 0,
                                    'lost': 0,
                                    'sys_error': 0,
                                    'unknown': 0,
                                    'not_counted': 0
                                }
                                values = {
                                    'wr_entry': 0,
                                    'lost': 0,
                                    'sys_error': 0,
                                    'unknown': 0,
                                    'not_counted': 0
                                }
                                row_c = 2
                                for tran in trans:
                                    comment = "NO COMMENT"
                                    if len(tran.comment) > 0:
                                        comment = tran.comment
                                    arr.append({
                                        'item_ref': tran.item_ref,
                                        'barcode': tran.barcode,
                                        'name': tran.name,
                                        'froze_qty': tran.froze_qty,
                                        'counted_qty': tran.counted_qty,
                                        'diff_qty': tran.diff_qty,
                                        'comment': comment,
                                        'issue': tran.issue
                                    })

                                if doc == 'preview':
                                    header['entries'] = entries
                                    header['values'] = values
                                    response['message'] = {
                                        'header': header, 'trans': arr
                                    }
                                elif doc == 'excel':
                                    sheet[f'A2'] = "SUMMARY"
                                    sheet['D2'] = tot_phy
                                    sheet['E2'] = tot_sys
                                    sheet['F2'] = tot_qdif
                                    sheet['G2'] = "-"
                                    sheet['H2'] = tot_vdif
                                    from datetime import datetime
                                    current_datetime = datetime.now()
                                    formatted_datetime = current_datetime.strftime("%Y%m%d%H%M%S")
                                    file = f"static/general/tmp/{stock_hd.loc}_{g_name.strip().replace(' ', '_')}_AS_OF_{as_of}_{formatted_datetime}.xlsx"
                                    workbook.save(file)
                                    response['message'] = file

                                response['status_code'] = 200
                                response['status'] = 'success'
                            else:
                                response['message'] = "NO OPEN STOCK"
                                response['status_code'] = 404
                                response['status'] = 'limit'

                        elif style == 'count_sheet':
                            print(data)
                            doc = data.get('doc')
                            key = data.get('key')
                            print(doc)

                            if doc == 'fr':
                                #frozen
                                if StockFreezeHd.objects.filter(pk=key).count() == 1:

                                    f_hd = StockFreezeHd.objects.get(pk=key)
                                    trans = f_hd.trans_only()

                                    pdf = FPDF('P', 'mm', 'A4')
                                    pdf.add_page()
                                    pdf.set_font('Arial', 'BU', 10)
                                    pdf.cell(180, 5, "STOCK COUNT SHEET", 0, 1, 'C')
                                    pdf.ln(10)

                                    pdf.set_font('Arial', 'B', 10)
                                    pdf.cell(25, 5, "LOCATION : ", 0, 0, 'L')
                                    pdf.set_font('Arial', '', 7)
                                    pdf.cell(80, 5, f" {f_hd.loc_id}", 0, 1, 'L')

                                    pdf.set_font('Arial', 'B', 10)
                                    pdf.cell(25, 5, "REMARKS : ", 0, 0, 'L')
                                    pdf.set_font('Arial', '', 7)
                                    pdf.cell(80, 5, f" {f_hd.remarks}", 0, 1, 'L')
                                    pdf.ln(10)
                                    # header
                                    pdf.set_font('Arial', 'B', 10)
                                    pdf.cell(10, 5, f"LN", 1, 0, 'L')
                                    pdf.cell(10, 5, f"LN", 1, 0, 'L')
                                    pdf.cell(25, 5, f"ITEM REF", 1, 0, 'L')
                                    pdf.cell(30, 5, f"BARCODE", 1, 0, 'L')
                                    pdf.cell(100, 5, f"NAME", 1, 0, 'L')
                                    # pdf.cell(20, 5, f"FROZEN", 1, 0, 'L')
                                    pdf.cell(20, 5, f"COUNTED", 1, 1, 'L')

                                    pdf.set_font('Arial', '', 7)

                                    line = 0

                                    for tran in trans:
                                        line += 1
                                        pdf.cell(10, 5, f"{line}", 1, 0, 'L')
                                        pdf.cell(25, 5, f"{tran.item_ref}", 1, 0, 'L')
                                        pdf.cell(30, 5, f"{tran.barcode}", 1, 0, 'L')
                                        pdf.cell(100, 5, f"{tran.name}", 1, 0, 'L')
                                        # pdf.cell(20, 5, f"{tran.qty}", 1, 0, 'L')
                                        pdf.cell(20, 5, f"", 1, 1, 'L')

                                    file = f'static/general/tmp/{f_hd.loc_id}.pdf'
                                    pdf.output(file, 'F')

                                    response = {
                                        'status_code': 200, 'message': file, 'status': 'success'
                                    }
                                else:
                                    response = {
                                        'status_code': 404, 'message': "ENTRY NOT FOUND", 'status': 'not_found'
                                    }
                            elif doc == 'stc':
                                if StockCountHD.objects.filter(pk=key).count() == 1:
                                    hd = StockCountHD.objects.get(pk=key)
                                    f_hd = hd.frozen
                                    trans = hd.trans()
                                    trans_only = trans['trans']

                                    pdf = FPDF('P', 'mm', 'A4')
                                    pdf.add_page()
                                    pdf.set_font('Arial', 'BU', 10)
                                    pdf.cell(180, 5, "COUNT REPORT", 0, 1, 'C')
                                    pdf.ln(10)

                                    pdf.set_font('Arial', 'B', 10)
                                    pdf.cell(25, 5, "LOCATION : ", 0, 0, 'L')
                                    pdf.set_font('Arial', '', 7)
                                    pdf.cell(80, 5, f" {f_hd.loc_id}", 0, 1, 'L')

                                    pdf.set_font('Arial', 'B', 10)
                                    pdf.cell(25, 5, "REMARKS : ", 0, 0, 'L')
                                    pdf.set_font('Arial', '', 7)
                                    pdf.cell(80, 5, f" {f_hd.remarks}", 0, 1, 'L')
                                    pdf.ln(10)
                                    # header
                                    pdf.set_font('Arial', 'B', 8)
                                    pdf.cell(10, 5, f"LN", 1, 0, 'L')
                                    pdf.cell(20, 5, f"BARCODE", 1, 0, 'L')
                                    pdf.cell(55, 5, f"NAME", 1, 0, 'L')
                                    pdf.cell(15, 5, f"SELL", 1, 0, 'L')
                                    pdf.cell(15, 5, f"FRZN", 1, 0, 'L')
                                    pdf.cell(15, 5, f"FR VL", 1, 0, 'L')
                                    pdf.cell(15, 5, f"CT", 1, 0, 'L')
                                    pdf.cell(15, 5, f"CT VL", 1, 0, 'L')
                                    pdf.cell(15, 5, f"DIFF", 1, 0, 'L')
                                    pdf.cell(15, 5, f"DF VL", 1, 1, 'L')



                                    line = 0
                                    summary = trans['summary']
                                    pdf.cell(100, 5, f"SUMMARY", 1, 0, 'L')
                                    pdf.cell(15, 5, f"{summary['total_frozen']}", 1, 0, 'L')
                                    pdf.cell(15, 5, f"{summary['value_frozen']}", 1, 0, 'L')
                                    pdf.cell(15, 5, f"{summary['total_counted']}", 1, 0, 'L')
                                    pdf.cell(15, 5, f"{summary['value_counted']}", 1, 0, 'L')
                                    pdf.cell(15, 5, f"{summary['qty_difference']}", 1, 0, 'L')
                                    pdf.cell(15, 5, f"{summary['value_difference']}", 1, 1, 'L')

                                    pdf.set_font('Arial', '', 5)
                                    for tran in trans_only:
                                        line += 1
                                        pdf.cell(10, 5, f"{line}", 1, 0, 'L')
                                        pdf.cell(20, 5, f"{tran.barcode}", 1, 0, 'L')
                                        pdf.cell(55, 5, f"{tran.name}", 1, 0, 'L')
                                        pdf.cell(15, 5, f"{tran.sell_price}", 1, 0, 'L')
                                        pdf.cell(15, 5, f"{tran.froze_qty}", 1, 0, 'L')
                                        pdf.cell(15, 5, f"{tran.froze_qty * tran.sell_price}", 1, 0, 'L')
                                        pdf.cell(15, 5, f"{tran.counted_qty}", 1, 0, 'L')
                                        pdf.cell(15, 5, f"{tran.counted_qty * tran.sell_price}", 1, 0, 'L')
                                        pdf.cell(15, 5, f"{tran.diff_qty}", 1, 0, 'L')
                                        pdf.cell(15, 5, f"{tran.diff_qty * tran.sell_price}", 1, 1, 'L')

                                        # summary



                                    pdf.set_font('Arial', 'B', 8)
                                    pdf.cell(100, 5, f"SUMMARY", 1, 0, 'L')
                                    pdf.cell(15, 5, f"{summary['total_frozen']}", 1, 0, 'L')
                                    pdf.cell(15, 5, f"{summary['value_frozen']}", 1, 0, 'L')
                                    pdf.cell(15, 5, f"{summary['total_counted']}", 1, 0, 'L')
                                    pdf.cell(15, 5, f"{summary['value_counted']}", 1, 0, 'L')
                                    pdf.cell(15, 5, f"{summary['qty_difference']}", 1, 0, 'L')
                                    pdf.cell(15, 5, f"{summary['value_difference']}", 1, 1, 'L')

                                    # pdf.ln(10)
                                    # pdf.set_font('Arial', 'B', 8)
                                    # pdf.cell(30, 5, f"FROZEN QTY", 1, 0, 'L')
                                    # pdf.set_font('Arial', '', 5)
                                    # pdf.cell(30, 5, f"{summary['total_frozen']}", 1, 0, 'L')

                                    # pdf.set_font('Arial', 'B', 8)
                                    # pdf.cell(30, 5, f"COUNTED QTY", 1, 0, 'L')
                                    # pdf.set_font('Arial', '', 5)
                                    # pdf.cell(30, 5, f"{summary['total_counted']}", 1, 0, 'L')
                                    #
                                    # pdf.set_font('Arial', 'B', 8)
                                    # pdf.cell(30, 5, f"DIFF QTY", 1, 0, 'L')
                                    # pdf.set_font('Arial', '', 5)
                                    # pdf.cell(25, 5, f"{summary['qty_difference']}", 1, 1, 'L')
                                    #
                                    # pdf.set_font('Arial', 'B', 8)
                                    # pdf.cell(30, 5, f"FROZEN VAL", 1, 0, 'L')
                                    # pdf.set_font('Arial', '', 5)
                                    # pdf.cell(30, 5, f"{summary['value_frozen']}", 1, 0, 'L')
                                    #
                                    # pdf.set_font('Arial', 'B', 8)
                                    # pdf.cell(30, 5, f"COUNTED VAL", 1, 0, 'L')
                                    # pdf.set_font('Arial', '', 5)
                                    # pdf.cell(30, 5, f"{summary['value_counted']}", 1, 0, 'L')
                                    #
                                    # pdf.set_font('Arial', 'B', 8)
                                    # pdf.cell(30, 5, f"DIFF VAL", 1, 0, 'L')
                                    # pdf.set_font('Arial', '', 5)
                                    # pdf.cell(25, 5, f"{summary['value_difference']}", 1, 1, 'L')


                                    file = f'static/general/tmp/{f_hd.loc_id}.pdf'
                                    pdf.output(file, 'F')

                                    response = {
                                        'status_code': 200, 'message': file, 'status': 'success'
                                    }


                                else:
                                    response = {
                                        'status_code': 404, 'message': "ENTRY NOT FOUND", 'status': 'not_found'
                                    }
                            else:
                                response['message'] = f"UNKNOWN DOCUMENT {doc}"


                    elif stage == 'frozen':
                        pk = data.get('pk')

                        # check if hd exist
                        hd_x = StockFreezeHd.objects.filter(pk=pk)

                        if hd_x.count() == 1:
                            response['status_code'] = 200

                            hd = StockFreezeHd.objects.get(pk=pk)
                            trans = hd.trans()

                            header = {
                                'pk': hd.pk,
                                'loc': hd.loc_id,
                                'remarks': hd.remarks,
                                'ref': hd.ref,
                                'time': f"{hd.created_date} {hd.created_time}",
                                'status': hd.status,
                                'owner': hd.owner.username,
                                'next': hd.next(),
                                'prev': hd.prev(),
                                'approve':hd.approve
                            }

                            tr = []

                            for t in trans['trans']:
                                ref = t.item_ref
                                barcode = t.barcode
                                name = t.name
                                qty = t.qty

                                tr.append({
                                    'item_ref': ref,
                                    'barcode': barcode,
                                    'name': name,
                                    'qty': qty
                                })

                            response['message'] = {
                                'header': header,
                                'count': trans['count'],
                                'trans': tr
                            }

                        else:
                            response['status_code'] = 404
                            response['message'] = f"Expected 1 but got {hd_x.count()} counts"

                    elif stage == 'frozen_hd':

                        frozn = StockFreezeHd.objects.filter(status=1).order_by('-pk')
                        counts = frozn.count()
                        arr = []
                        for fr in frozn:
                            arr.append({
                                'pk': fr.pk,
                                'entry': f"FR{fr.loc_id}{fr.pk}",
                                'remarks': fr.remarks,
                                'location': fr.loc_id,
                                'approve': fr.approve
                            })

                        response['message'] = arr
                        response['status_code'] = 200



            except Exception as e:
                response['status'] = 'error'
                response['status_code'] = 505
                # Get the line number where the exception occurred
                line_number = traceback.extract_tb(e.__traceback__)[-1].lineno

                # Set the message with the line number
                response['message'] = f"Error occurred at line {line_number}: {str(e)}"

        elif method == 'PUT':
            if module == 'stock':
                stage = data.get('stage')
                if stage == 'hd':
                    loc = data.get('loc')
                    remark = data.get('remark')

                    # check if there is open hd
                    if StockCountHD.objects.filter(status=1).exists():
                        response['message'] = "There is an open stock take"
                        response['status_code'] = 505
                        response['status'] = 'WARNING'

                    else:
                        # open stock
                        StockCountHD(loc=loc, remark=remark).save()
                        response['status_code'] = 200
                        response['status'] = 'success'
                elif stage == 'tran':
                    item_ref = data.get('item_ref')
                    qty = data.get('qty')
                    myName = data.get('user')

                    username = myName
                    server = f"{DB_SERVER},{DB_PORT}"
                    database = DB_NAME
                    username = DB_USER
                    password = DB_PASSWORD
                    driver = '{ODBC Driver 17 for SQL Server}'  # Change this to the driver you're using
                    connection_string = f"DRIVER={driver};SERVER={server};DATABASE={database};UID={username};PWD={password}"
                    connection = pyodbc.connect(connection_string)
                    cursor = connection.cursor()

                    query = f"select barcode,item_ref,item_des1,sell_price from product_master where item_ref = '{item_ref}'"
                    cursor.execute(query)

                    row = cursor.fetchone()
                    if row:
                        barcode = row[0].strip()
                        name = row[2].strip()
                        if row[3] is None:
                            sell_price = 0.00
                        else:
                            sell_price = row[3]

                        val = Decimal(qty) * Decimal(sell_price)
                        hd = StockCountHD.objects.get(status=1)
                        StockCountTrans(stock_count_hd=hd, item_ref=item_ref, barcode=barcode, name=name,
                                        sell_price=sell_price, quantity=qty, owner=myName, value=val).save()

                        response['status_code'] = 200
                        response['status'] = 'success'
                        response['message'] = f"ITEM {name} Added"
                        cursor.close()

                    else:
                        response['status_code'] = 404
                        response['status'] = 'error'
                        response['message'] = f"could not find item with ref {item_ref}"
                elif stage == 'comment':
                    comment_pk = data.get('comment_pk')
                    comment = data.get('comment')
                    issue = data.get('issue')

                    if StockCountTrans.objects.filter(pk=comment_pk).count() == 1:
                        # update
                        coun = StockCountTrans.objects.get(pk=comment_pk)
                        coun.comment = comment
                        coun.issue = issue
                        coun.save()
                        response['message'] = "Comment Updated"
                    else:
                        response['message'] = f"No count with pk as {comment_pk}"
                elif stage == 'save_frozen':

                    header = data.get('header')
                    trans = data.get('trans')

                    if len(trans) > 0:

                        try:
                            loc_id = header.get('loc')
                            frozen_ref = header.get('frozen_ref')
                            remarks = header.get('remarks')
                            us_pk = header.get('owner')

                            owner = get_object_or_404(User, pk=us_pk)
                            stock_freeze_hd = StockFreezeHd.objects.create(loc_id=loc_id, ref=frozen_ref,
                                                                           remarks=remarks, owner=owner)

                            try:
                                for tran in trans:
                                    ref = tran['ref']
                                    barcode = tran['barcode']
                                    qty = tran['qty']
                                    name = tran['name']

                                    StockFreezeTrans.objects.create(entry_id=stock_freeze_hd.pk, item_ref=ref,
                                                                    barcode=barcode, qty=qty, name=name)

                                response['message'] = "SAVED"
                                response['status_code'] = 200
                            except Exception as e:
                                stock_freeze_hd.delete()
                                response['message'] = str(e)

                        except Exception as e:
                            response['message'] = str(e)


                    else:

                        response['message'] = f" {len(trans)} Transactions cannot be empty"
                        response['status'] = "EMPTY"
                        response['status_code'] = 100
                elif stage == 'save_cont':
                    header = data.get('header')
                    trans = data.get('trans')
                    count_pk = header.get('count_pk')
                    task = header.get('task')

                    frozen_ref = header.get('ref')
                    comment = header.get('comment')
                    owner_pk = header.get('owner_pk')
                    owner = User.objects.get(pk=owner_pk)

                    if StockFreezeHd.objects.filter(pk=frozen_ref).count() == 1:
                        try:

                            frozen = StockFreezeHd.objects.get(pk=frozen_ref)
                            frozen.status = 2

                            if StockCountHD.objects.filter(frozen=frozen, comment=comment, owner=owner).exists():
                                # delete
                                StockCountHD.objects.get(frozen=frozen, comment=comment, owner=owner).delete()
                            StockCountHD(frozen=frozen, comment=comment, owner=owner).save()
                            count_hd = StockCountHD.objects.get(frozen=frozen, comment=comment, owner=owner)

                            # save trans
                            for tran in trans:
                                print(tran)
                                ref = tran.get('ref')
                                barcode = tran.get('barcode')
                                name = tran.get('name')
                                frozen = tran.get('frozen')
                                counted = tran.get('counted')
                                diff = tran.get('diff')
                                row_comment = tran.get('row_comment')
                                row_iss = tran.get('row_iss')

                                # get this item value
                                cursor= db()

                                q = f"SELECT sell_price FROM product_master where barcode = '{barcode}'"
                                value = cursor.execute(q).fetchone()[0]
                                if value is None:
                                    value = 0.00

                                print({'VAL': value, 'diff': Decimal(str(diff)), 'frozen': Decimal(str(frozen)),
                                       'counted': Decimal(counted)})

                                diff_val = Decimal(value) * Decimal(diff)
                                froze_val = Decimal(value) * Decimal(frozen)
                                counted_val = Decimal(value) * Decimal(counted)

                                cursor.close()



                                # save tran
                                StockCountTrans.objects.create(stock_count_hd=count_hd, item_ref=ref, barcode=barcode,
                                                               name=name, froze_qty=frozen, counted_qty=counted,
                                                               diff_qty=diff, comment=row_comment, issue=row_iss,froze_val=froze_val,diff_val=diff_val,counted_val=counted_val)

                            frozen.save()
                            response['status_code'] = 200
                            response['status'] = 'success'
                            response['message'] = "Stock Count Saved"



                        except Exception as e:
                            response['status_code'] = 500
                            response['status'] = 'error'
                            response['message'] = str(e)



                    else:
                        response['statys_code'] = 404
                        response['message'] = f"CANNOT FIND FROZEN with entry {frozen_ref}"



        elif method == 'PATCH':
            if module == 'stock':
                stage = data.get('stage')
                task = data.get('task')
                if stage == 'hd':
                    if task == 'close':
                        opened = StockCountHD.objects.filter(status=1).last()
                        opened.status = 2
                        opened.save()
                        response = {'status': 200, 'status_code': 'success', 'message': "OPENED COUNT closed"}

                if stage == 'save_cont':
                    header = data.get('header')
                    count_pk = header.get('count_pk')

                    # print(data)
                    if StockCountHD.objects.filter(pk=count_pk).count() == 1:
                        c_hd = StockCountHD.objects.get(pk=count_pk)
                        c_hd.comment = header.get('comment')

                        trans = data.get('trans')


                        # delete all trans
                        # delete this tran and add again
                        StockCountTrans.objects.filter(stock_count_hd=c_hd).delete()


                        for tran in trans:
                            ref = tran.get('ref')
                            barcode = tran.get('barcode').strip()
                            name = tran.get('name')
                            frozen = tran.get('frozen')
                            counted = str(tran.get('counted'))
                            diff = tran.get('diff')
                            row_comment = tran.get('row_comment')
                            row_iss = tran.get('row_iss')

                            cursor = db()
                            q = f"SELECT sell_price FROM product_master where barcode = '{barcode}'"
                            value = cursor.execute(q).fetchone()[0]
                            if value is None:
                                value = 0.00

                            diff_val = Decimal(value) * Decimal(diff)
                            froze_val = Decimal(value) * Decimal(frozen)
                            counted_val = Decimal(value) * Decimal(counted)

                            cursor.close()



                            # save tran
                            StockCountTrans.objects.create(stock_count_hd=c_hd, item_ref=ref, barcode=barcode,
                                                           name=name, froze_qty=frozen, counted_qty=counted,
                                                           diff_qty=diff, comment=row_comment, issue=row_iss,diff_val=diff_val,froze_val=froze_val,counted_val=counted_val,sell_price=value)
                        c_hd.save()
                        response['message'] = "UPDATE SUCCESSFUL"
                    else:
                        var = response['status_code'] = 404
                        response['message'] = f"CANNOT FIND DOCUMENT with key {count_pk}"

            elif module == 'approve':
                doc = data.get('doc')
                key = data.get('key')

                if doc == 'FR':
                    hd = StockFreezeHd.objects.filter(pk=key)
                    count = hd.count()
                else:
                    count = 0

                if count == 1:
                    try:
                        head = hd.last()
                        head.approve = 1
                        head.save()

                        response['status_code'] = 200
                        response['status'] = 'success'
                        response['message'] = "APPROVED"
                    except Exception as e:
                        response['status_code'] = 505
                        response['status'] = 'error'
                        response['message'] = str(e)

                else:
                    response['status_code'] = 404
                    response['status'] = 'error'
                    response['message'] = f"Count not find matching document ({doc} - {key})"




    except json.JSONDecodeError as e:
        response["status_code"] = 400
        response["status"] = "Bad Request"
        response["message"] = f"Error decoding JSON: {str(e)}"

    return JsonResponse(response)
