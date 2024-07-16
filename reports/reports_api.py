import json
import os
import sys

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from fpdf import FPDF

from admin_panel.anton import format_currency
from admin_panel.models import Emails, MailQueues, MailAttachments, TransferHD, TransferTran
from cmms.models import ProformaInvoice
from inventory.models import GrnHd
from meeting.models import MeetingHD
from reports.models import ReportForms, ReportLegend, LegendSubs, DepartmentReportMailQue
from taskmanager.models import Tasks


@csrf_exempt
def interface(request):
    response = {
        'status_code': 0,
        'message': ""
    }

    success_response = {
        'status_code': 200,
        'message': "Procedure Completed Successfully"
    }

    # get method
    method = request.method

    try:

        body = json.loads(request.body)
        module = body.get('module')
        data = body.get('data')
        doc = data.get('doc')
        key = data.get('key')
        output = data.get('output')
        header = {}
        transactions = []

        if doc == 'TSK' and output == 'PDF':
            import re
            clean = re.compile('<.*?>')
            head = Tasks.objects.get(uni=key)
            pdf = FPDF('P', 'mm', 'A4')
            pdf.add_page()

            pdf.set_font('Arial', 'B', 16)
            pdf.cell(200, 5, head.title, 0, 1)
            pdf.ln(5)
            pdf.set_font('Arial', '', 10)
            pdf.multi_cell(0, 5, f"{re.sub(clean, '', head.description)}", 0, 'L')

            pdf.ln(5)
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(200, 5, 'TRANSACTIONS', 0, 1)
            pdf.ln(2)

            t_count = 1
            for tran in head.transaction():
                pdf.set_font('Arial', 'B', 10)
                pdf.cell(200, 5, f'{t_count}. {tran.title}', 0, 1)
                pdf.set_font('Arial', '', 10)
                pdf.cell(200, 5, f"By {tran.owner.username} On {tran.created_date} {tran.created_time}", 0, 1)
                pdf.ln(1)
                pdf.multi_cell(0, 5, f"{re.sub(clean, '', tran.description)}", 0, 'L')
                pdf.ln(2)
                t_count += 1

            fil_name = f"static/general/tmp/test.pdf"
            pdf.output(fil_name, 'F')

            success_response['message'] = fil_name
            response = success_response

        elif doc == 'SAVE_FORM':
            ReportForms(key=key, code=output, description=key).save()

            success_response['message'] = "Form Saved"
            response = success_response

        elif doc == 'SAVE_REPORT_LEGEND':
            ReportLegend(name=key, description=output).save()

            success_response['message'] = "Legend Saved"
            response = success_response

        elif doc == 'SAVE_REPORT_LEGEND_SUB':
            description = data.get('description')
            action = data.get('action')
            legend = data.get('legend')
            name = data.get('name')

            lg = ReportLegend.objects.get(pk=legend)

            LegendSubs(legend=lg, description=description, action=action, name=name).save()

            success_response['message'] = "Legend Saved"
            response = success_response

        elif doc == 'send_dept_mail':

            pending_mail_list = DepartmentReportMailQue.objects.filter(status=0)
            resp = []
            deptt = []
            if pending_mail_list.count() > 0:
                for mails in pending_mail_list:
                    head_email = mails.department.email_of_head
                    files = mails.files
                    clean_files = files.rstrip(',')
                    files_array = clean_files.split(',')

                    import smtplib
                    from email.mime.multipart import MIMEMultipart
                    from email.mime.text import MIMEText
                    from email.mime.application import MIMEApplication
                    # Email configuration
                    smtp_server = "smtp.gmail.com"
                    smtp_port = 587
                    sender_email = "donotreply@protonghana.com"
                    sender_password = "arbogirmfctrnnpt"  # Use an app password for security
                    subject = f"{mails.department.name} PRODUCTIVITY REPORT "
                    html_content = "Productivity report base on task manager. This report will includes all tasks"

                    # Create a MIME multipart message
                    msg = MIMEMultipart()
                    msg["From"] = sender_email
                    msg["To"] = head_email
                    msg["Subject"] = subject

                    # Attach the HTML content to the message
                    msg.attach(MIMEText(html_content, "html"))
                    files_array = clean_files.split(',')
                    ff = []
                    for att_file in files_array:
                        # Attach the file
                        ff.append(att_file)
                        attachment_filename = f"static/general/task-reports/{att_file}"
                        attachment_path = attachment_filename
                        try:
                            with open(attachment_path, 'rb') as attachment:
                                part = MIMEApplication(attachment.read())
                                part.add_header('Content-Disposition', 'attachment', filename=att_file.strip())
                                msg.attach(part)
                                print(attachment_filename)
                        except FileNotFoundError:
                            print(f"File not found: {attachment_filename}")
                        except Exception as e:
                            print(f"Error attaching file {attachment_filename}: {str(e)}")

                            # Connect to the Gmail SMTP server and send the email
                    try:
                        server = smtplib.SMTP(smtp_server, smtp_port)
                        server.starttls()
                        server.login(sender_email, sender_password)
                        server.sendmail(sender_email, head_email, msg.as_string())
                        deptt.append({'department': mails.department.name, 'files': ff})
                        mails.status = 1
                        mails.save()

                    except Exception as e:
                        # response['status_code'] = 505
                        # response['message'] = str(e)
                        deptt.append({'department': mails.department.name, 'files': e})
                        message = f"COULD NOT SEND EMAIL {str(e)}"
                        print("Error sending email:", e)
                    finally:
                        server.quit()

                success_response['message'] = deptt
                response = success_response
            else:
                success_response['message'] = "No Emails To Send For Department"
                response = success_response

        elif doc == 'mail_sync':
            pending_mails = Emails.objects.filter(status=0)
            c = 0
            for email in pending_mails:
                recipient = email.sent_to
                subject = email.subject
                sent_from = email.sent_from
                body = email.body
                this_email = Emails.objects.filter(pk=email.pk)
                email_type = email.email_type
                email_ref = email.ref
                attacs = email.attachments
                cc = email.send_cc()

                import smtplib
                from email.mime.multipart import MIMEMultipart
                from email.mime.text import MIMEText
                from email.mime.application import MIMEApplication

                # Email configuration
                smtp_server = "smtp.gmail.com"
                smtp_port = 587
                sender_email = "donotreply@protonghana.com"
                sender_password = "arbogirmfctrnnpt"  # Use an app password for security
                html_content = body

                # Create a MIME multipart message
                msg = MIMEMultipart()
                msg["From"] = sender_email
                msg["To"] = recipient
                msg["Subject"] = subject
                msg['Cc'] = email.cc

                # Attach the HTML content to the message
                msg.attach(MIMEText(html_content, "html"))
                files_array = attacs.split(',')
                ff = []

                for att_file in files_array:
                    # Attach the file
                    ff.append(att_file)
                    attachment_filename = f"static/general/crm-logs-reports/{att_file}"
                    attachment_path = attachment_filename
                    try:
                        with open(attachment_path, 'rb') as attachment:
                            part = MIMEApplication(attachment.read())
                            part.add_header('Content-Disposition', 'attachment', filename=att_file.strip())
                            msg.attach(part)
                            print(att_file)
                    except FileNotFoundError:
                        print(f"File not found: {attachment_filename}")
                    except Exception as e:
                        print(f"Error attaching file {attachment_filename}: {str(e)}")

                        # Connect to the Gmail SMTP server and send the email
                try:
                    server = smtplib.SMTP(smtp_server, smtp_port)
                    server.starttls()
                    server.login(sender_email, sender_password)
                    server.sendmail(sender_email, recipient, msg.as_string())

                    email.status = 1
                    email.save()

                except Exception as e:
                    # response['status_code'] = 505
                    # response['message'] = str(e)

                    message = f"COULD NOT SEND EMAIL {str(e)}"
                    print("Error sending email:", e)
                finally:
                    server.quit()

                c += 1
            success_response['message'] = f"{c} EMAILS SENT"
            response = success_response

        elif doc == 'mail_sync_v2':
            import smtplib
            from email.mime.multipart import MIMEMultipart
            from email.mime.text import MIMEText
            from email.mime.application import MIMEApplication

            mails = MailQueues.objects.filter(is_sent=False)

            for mail in mails:
                sender = mail.sender
                recipient = mail.recipient
                subject = mail.subject
                body = mail.body
                cc = mail.cc

                smtp_server = sender.host
                smtp_port = sender.port
                sender_email = sender.address
                sender_password = sender.password
                html_content = body

                msg = MIMEMultipart()
                msg['From'] = sender_email
                msg['To'] = recipient
                msg['Subject'] = subject
                msg['Cc'] = cc

                to_address = [recipient] + cc.split(',') if cc else [recipient]

                # attachments
                msg.attach(MIMEText(html_content, 'html'))
                attachments = MailAttachments.objects.filter(mail=mail)

                for attachment in attachments:
                    attachment_filename = attachment.attachment.path
                    file_name = os.path.basename(attachment_filename)
                    print(attachment_filename)
                    try:
                        with open(attachment_filename, 'rb') as attached:
                            part = MIMEApplication(attached.read())  # Read the content of the file
                            part.add_header('Content-Disposition', f'attachment; filename="{file_name}"')
                            msg.attach(part)
                    except FileNotFoundError:
                        print(f"File not found: {attachment_filename}")
                    except Exception as e:
                        print(f"Error attaching file {attachment_filename}: {str(e)}")

                # send email

                try:
                    server = smtplib.SMTP(smtp_server, smtp_port)
                    server.starttls()
                    server.login(sender_email, sender_password)
                    server.sendmail(sender_email, to_address, msg.as_string())

                    mail.is_sent = 1
                    mail.sent_response = "Email Sent"
                    mail.save()
                    server.quit()
                    response['message'] = "Email Sent"
                except Exception as e:
                    # response['status_code'] = 505
                    # response['message'] = str(e)
                    mail.sent_response = f"Could Not Send {e}"
                    mail.save()

                    message = f"COULD NOT SEND EMAIL {str(e)}"
                    response['message'] = message

                success_response['message'] = f"{mails.count()} Sent"
                response = success_response

        elif doc == 'print':
            document = data.get('document')
            if document == 'grn':
                entry_key = data.get('pk')
                header = GrnHd.objects.filter(pk=entry_key)
                if header.count() == 1:
                    hd = header.last()
                    pdf = FPDF()
                    pdf.add_page('P')

                    # supplier
                    pdf.set_font('Arial', 'B', 10)
                    pdf.cell(190, 10, 'GOODS RECIEVING NOTE', 0, 1, 'C')
                    pdf.set_font('Arial', 'B', 8)
                    pdf.cell(20, 5, "Supplier :  ", 0, 0)
                    pdf.set_font('Arial', '', 8)
                    pdf.cell(100, 5, hd.supplier.company, 0, 1)

                    # Tax Amount
                    pdf.set_font('Arial', 'B', 8)
                    pdf.cell(20, 5, "Date :  ", 0, 0)
                    pdf.set_font('Arial', '', 8)
                    pdf.cell(100, 5, f"{hd.created_on}", 0, 1)

                    # Tax Amount
                    pdf.set_font('Arial', 'B', 8)
                    pdf.cell(20, 5, "Entry No :  ", 0, 0)
                    pdf.set_font('Arial', '', 8)
                    pdf.cell(100, 5, f"{hd.pk}", 0, 1)

                    # location
                    pdf.set_font('Arial', 'B', 8)
                    pdf.cell(20, 5, "Location :  ", 0, 0)
                    pdf.set_font('Arial', '', 8)
                    pdf.cell(100, 5, hd.loc.descr, 0, 1)

                    cost = hd.cost()
                    total_amount = cost['taxable_amt']

                    # Net Amount
                    pdf.set_font('Arial', 'B', 8)
                    pdf.cell(20, 5, "Amount :  ", 0, 0)
                    pdf.set_font('Arial', '', 8)
                    pdf.cell(100, 5, f"{format_currency(total_amount)}", 0, 1)

                    # table header
                    pdf.set_font('Arial', 'B', 8)
                    pdf.cell(45, 5, "BARCODE", 1, 0)
                    pdf.cell(70, 5, "NAME", 1, 0)
                    pdf.cell(20, 5, "PACKING", 1, 0)
                    pdf.cell(20, 5, "PRICE", 1, 0)
                    pdf.cell(20, 5, "QTY", 1, 0)
                    pdf.cell(20, 5, "TOTAL", 1, 1)

                    pdf.set_font('Arial', '', 8)
                    for tran in hd.trans():
                        pdf.cell(45, 5, tran.product.barcode[:10], 1, 0)
                        pdf.cell(70, 5, tran.product.descr[:60], 1, 0)
                        pdf.cell(20, 5, f"{tran.packing.packing_un.code} ({tran.packing.pack_qty})", 1, 0)
                        pdf.cell(20, 5, f"{format_currency(tran.un_cost)}", 1, 0)
                        pdf.cell(20, 5, f"{tran.qty}", 1, 0)
                        pdf.cell(20, 5, f"{format_currency(tran.tot_cost)}", 1, 1)

                    file_name = f"static/general/servicing/grn.pdf"
                    pdf.output(file_name)
                    success_response['message'] = file_name
                    response = success_response

                else:
                    raise Exception(f"Cannot find document {entry_key}")

            elif document == 'TR':

                entry_key = data.get('key')
                sender = data.get("sender")
                hd = TransferHD.objects.get(entry_no=entry_key)
                trans = TransferTran.objects.filter(parent=hd)

                pdf = FPDF()
                pdf.add_page('P')
                pdf.set_margins(4,4,4)
                # Header
                pdf.set_font('Arial', 'B', 10)
                pdf.cell(190, 10, 'TRANSFER NOTE', 0, 1, 'C')
                pdf.set_font('Arial', 'B', 8)
                pdf.cell(20, 5, "Entry No:  ", 0, 0)
                pdf.set_font('Arial', '', 8)
                pdf.cell(100, 5, f"{hd.entry_no}", 0, 1)

                # Tax Amount
                pdf.set_font('Arial', 'B', 8)
                pdf.cell(20, 5, "Entry Date :  ", 0, 0)
                pdf.set_font('Arial', '', 8)
                pdf.cell(100, 5, f"{hd.created_on}", 0, 1)

                # Tax Amount
                pdf.set_font('Arial', 'B', 8)
                pdf.cell(20, 5, "From :  ", 0, 0)
                pdf.set_font('Arial', '', 8)
                pdf.cell(100, 5, f"{hd.loc_fr.descr}", 0, 1)

                # location
                pdf.set_font('Arial', 'B', 8)
                pdf.cell(20, 5, "To :  ", 0, 0)
                pdf.set_font('Arial', '', 8)
                pdf.cell(100, 5, f"{hd.loc_to.descr}", 0, 1)

                pdf.set_font('Arial', 'B', 8)
                pdf.cell(20, 5, "Cost :  ", 0, 0)
                pdf.set_font('Arial', '', 8)
                pdf.cell(100, 5, f"{hd.total_cost()}", 0, 1)

                # remarks
                pdf.set_font('Arial', 'B', 8)
                pdf.cell(20, 5, "REMARKS :  ", 0, 1)
                pdf.set_font('Arial', '', 8)
                pdf.multi_cell(0,5,f"{hd.remarks}",0,"L")
                pdf.ln(10)


                # make transactions
                pdf.set_font('Arial', 'B', 8)
                pdf.cell(15, 5, "LINE", 1, 0)
                pdf.cell(25, 5, "BARCODE", 1, 0)
                pdf.cell(45, 5, "NAME", 1, 0)
                pdf.cell(15, 5, "PACK", 1, 0)
                pdf.cell(25, 5, "PACK QTY", 1, 0)
                pdf.cell(25, 5, "TRAN QTY", 1, 0)
                pdf.cell(25, 5, "UNIT COST", 1, 0)
                pdf.cell(25, 5, "TOTAL COST", 1, 1)
                pdf.set_font('Arial', '', 8)

                for tr in trans:
                    pdf.cell(15, 5, f"{tr.line}", 1, 0)
                    pdf.cell(25, 5, f"{tr.product.barcode}", 1, 0)
                    pdf.cell(45, 5, f"{tr.product.descr[:30]}", 1, 0)
                    pdf.cell(15, 5, f"{tr.packing}", 1, 0)
                    pdf.cell(25, 5, f"{format_currency(tr.pack_qty)}", 1, 0)
                    pdf.cell(25, 5, f"{format_currency(tr.tran_qty)}", 1, 0)
                    pdf.cell(25, 5, f"{format_currency(tr.unit_cost)}", 1, 0)
                    pdf.cell(25, 5, f"{format_currency(tr.cost)}", 1, 1)

                pdf.ln(10)
                pdf.cell(50, 10, f"{hd.created_by.username}", 1, 0, 'C')
                pdf.cell(18, 10, "")
                pdf.cell(50, 10, f"{sender}", 1, 0, 'C')
                pdf.cell(18, 10, "")
                pdf.cell(50, 10, f"", 1, 1, 'C')

                pdf.cell(50, 10, f"Created By", 0, 0, 'C')
                pdf.cell(18, 10, "")
                pdf.cell(50, 10, f"Delivered By", 0, 0, 'C')
                pdf.cell(18, 10, "")
                pdf.cell(50, 10, f"Received By", 0, 1, 'C')


                file_name = f"static/general/tmp/transfer_{hd.entry_no}.pdf"
                pdf.output(file_name)
                success_response['message'] = file_name
                response = success_response

            else:
                raise Exception("Unknown Printing Document Type")

        elif doc == 'sales_proforma':
            start_date = data.get('start_date')
            end_date = data.get('end_date')
            status = data.get('status')
            export_type = data.get('export_type','json')
            if export_type == 'excel':
                import openpyxl
                book = openpyxl.Workbook()
                sheet = book.active
                sheet.title = "Proforma Report"

                sheet['A1'] = "Proforma Report"
                sheet.merge_cells("A1:J1")
                sheet['A2'] = "Industry"
                sheet['B2'] = "Organization"
                sheet['C2'] = "Personnel"
                sheet['D2'] = "Asset"
                sheet['E2'] = 'QTY'
                sheet['F2'] = "Currency"
                sheet['D2'] = "Unit Price"
                sheet['E2'] = "Total Amount"
                sheet['F2'] = "Created By"
                sheet['G2'] = "Created Date"
                sheet['H2'] = "Approved By"
                sheet['I2'] = "approved Date"
                sheet["J2"] = "Status"
                sheet_next = 3
            arr = []
            proformas = ProformaInvoice.objects.all()
            for proforma in proformas:
                # print(proforma.customer.name)
                if export_type == 'json':
                    arr.append(proforma.obj())

                if export_type == 'excel':
                    sheet[f'A{sheet_next}'] = proforma.customer.sector_of_company
                    sheet[f'B{sheet_next}'] = proforma.customer.company
                    sheet[f'C{sheet_next}'] = proforma.customer.name
                    sheet[f'D{sheet_next}'] = proforma.my_ass()
                    sheet[f'E{sheet_next}'] = proforma.quantity
                    sheet[f'F{sheet_next}'] = proforma.currency
                    sheet[f'D{sheet_next}'] = "Unit Price"
                    sheet[f'E{sheet_next}'] = "Total Amount"
                    sheet[f'F{sheet_next}'] = "Created By"
                    sheet[f'G{sheet_next}'] = "Created Date"
                    sheet[f'H{sheet_next}'] = "Approved By"
                    sheet[f'I{sheet_next}'] = "approved Date"
                    sheet[f"J{sheet_next}"] = "Status"

                    sheet_next += 1

            if export_type == 'json':
                success_response['message'] = arr

            if export_type == 'excel':
                file_name = 'static/general/tmp/proforma.xlsx'
                book.save(file_name)
                success_response['message'] = file_name

            response = success_response

    except Exception as e:
        error_type, error_instance, traceback = sys.exc_info()
        tb_path = traceback.tb_frame.f_code.co_filename
        line_number = traceback.tb_lineno
        response["status_code"] = 500
        response[
            "message"] = f"An error of type {error_type} occurred on line {line_number} in file {tb_path}. Details: {e}"

    return JsonResponse(response)
