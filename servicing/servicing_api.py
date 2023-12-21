import json
import sys

from django.contrib.auth.models import User
from django.db.models import Q
from django.http import JsonResponse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from fpdf import FPDF
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment

from admin_panel.anton import make_md5_hash, current_date, new_sms
from admin_panel.models import Emails, TicketHd, Sms, SmsApi, UserAddOns, TicketTrans, MailSenders, MailQueues, \
    MailAttachments
from admin_panel.sms_hold import *
from appscenter.models import App
from meeting.models import MeetingHD

from reports.models import ReportForms, ReportLegend, LegendSubs, DepartmentReportMailQue
from servicing.models import Services, SubServices, ServiceCard, ServiceMaterials, ServiceTechnicians, CheckList
from taskmanager.models import Tasks, TaskTransactions


@csrf_exempt
def interface(request):
    response = {
        'status_code': 0,
        'message': ""
    }

    success_response = {
        'status_code': 200,
        'message': "Procedure Completed Successfully"
    }

    # get method
    method = request.method

    try:

        body = json.loads(request.body)
        module = body.get('module')
        data = body.get('data')

        # write data
        if method == 'PUT':
            # save new service
            if module == 'services':
                service_code = data['service_code']
                service_description = data['service_description']
                service_name = data['service_name']
                service_owner = data['service_owner']

                Services(code=service_code, name=service_name, description=service_description,
                         owner_id=service_owner).save()

                t_service = Services.objects.get(code=service_code)

                # add sub
                SubServices(service=t_service, name=service_name, description=f"Default service for {service_name}",
                            owner_id=service_owner).save()

                success_response['message'] = "Serviced Added"


            elif module == 'service_sub':
                service_code = data.get('service_code')
                sub_name = data.get('sub_name')
                description = data.get('description')
                owner_id = data.get('owner_id')

                service = Services.objects.get(code=service_code)
                owner = User.objects.get(id=owner_id)
                SubServices(service=service, name=sub_name, description=description).save()

            # save job card
            elif module == 'jobcard':
                head = data.get('head')
                materials = data.get('materials')

                # get header data
                cardno = f"JC{ServiceCard.objects.all().count() + 1}"
                client_pk = head.get('client')
                entry_date = head.get('ticket_date')
                client = User.objects.get(pk=client_pk)
                phone = head.get('phone').replace('+233', '0')
                owner = User.objects.get(pk=head.get('owner'))
                remarks = head.get('remarks')
                annal = head.get('annal')
                ckl = head.get('checklist') or 'none'
                checklist = ckl.split(',')
                service = Services.objects.get(pk=head.get('service'))
                service_sub = SubServices.objects.get(pk=head.get('service_sub'))
                technician = ServiceTechnicians.objects.get(
                    technician=User.objects.get(pk=head.get('technician')),
                    service=service
                )
                tech_adon = UserAddOns.objects.get(user=technician.technician)
                importance = head.get('importance')
                tick = head.get('ticket')
                ticket_title = head.get('ticket_title')
                ticked_description = head.get('description')
                app = App.objects.get(pk=head.get('app'))

                if tick == '0':
                    # generate
                    TicketHd(owner=client, title=f"{ticket_title}", descr=f"{ticked_description}", status=1,
                             created_on=entry_date).save()
                    ticket = TicketHd.objects.filter(owner=client, title=f"{ticket_title}",
                                                     descr=f"{ticked_description}", status=1).last()
                else:
                    ticket2 = TicketHd.objects.get(pk=tick)
                    ticket2.status = 1
                    ticket2.title = ticket_title
                    ticket2.descr = ticked_description
                    ticket2.save()
                    ticket = TicketHd.objects.get(pk=tick)

                    # add to task managers
                uni = make_md5_hash(ticket_title)
                Tasks(title=ticket_title, uni=uni, description=ticked_description,
                      owner=technician.technician).save()

                task = Tasks.objects.get(uni=uni)
                # save service card
                just_service_card = ServiceCard(client=client, task=task, owner=owner,
                            remarks=f"{service.name}/{service_sub.name}/{ticket_title}",
                            service=service, service_sub=service_sub,
                            technician=technician, ticket=ticket, importance=importance, cardno=cardno, app=app,analysis=annal).save()

                # just_service_card = ServiceCard.objects.all().last()

                for check in checklist:
                    CheckList(cardno=just_service_card,name=check).save()

                # deal with materials
                for material in list(materials):
                    line = material['line']
                    item = material['item']
                    description = material['description']
                    price = material['price']
                    quantity = material['quantity']
                    total_price = material['total_price']

                    # save material
                    ServiceMaterials(line=line, item=item, description=description, price=price, quantity=quantity,
                                     total_price=total_price, service_card=just_service_card).save()

                tech_number = tech_adon.phone.replace('+233', '0')
                # queue SMS
                sms_api = SmsApi.objects.get(is_default=1)
                Sms(api=SmsApi.objects.get(is_default=1),
                    message=f"A ticket has been opened for your query \n\nTICKET : "
                            f"{cardno} \n\nSERVICE "
                            f"TYPE : {service.name} \n\nTITLE : "
                            f"{ticket_title} \n\n"
                            f"ASSIGNED TO : {tech_adon.user.get_full_name()} - {tech_number}\n\n"
                            f"You can track service using the "
                            f"link below "
                            f"\n\nhttp://ocean.snedaghana.loc/servicing/jobcard/tracking/{uni}/",
                    to=phone).save()

                # technician notification

                tech_message = (f"NEW TICKET!!\n\n"
                                f"You have been assigned a ticket and below are the details. \n\n"
                                f"TITLE : {ticket_title} \n\n"
                                f"DESCRIPTION : Sent via email \n\n"
                                f"REQUEST BY : {client.get_full_name()} - {phone} \n\n"
                                f"Please attend to the issue and call the owner if need be")

                email_body = (f"<strong>REPORTED BY</strong> : {client.get_full_name()}<br>"
                              f"<strong>TITLE</strong> : {ticket_title}<hr>"
                              f"{ticked_description}"
                              f"<hr>"
                              f"<i>NOTE: Please contact client with the details below for further discussion of issue </><br>"
                              f"<strong>Mobile </strong> : <a href='tel:{phone}'>{phone}</a> <br>"
                              f"<strong>Email : </strong> <a href='mailto:{client.email}'>{client.email}</a>")

                Sms(api=sms_api, message=tech_message, to=tech_number).save()
                Emails(sent_from='issues@snedaghana.loc', sent_to=technician.technician.email,
                       subject=f"NEW TICKET!! {ticket_title}", body=email_body).save()
                success_response['message'] = "Job Opened"

            elif module == 'send_ticket_to_client':
                cardno = data.get('cardno')
                service = ServiceCard.objects.get(cardno=cardno)
                message = data.get('message')

                # get client details
                client = service.client
                client_details = UserAddOns.objects.get(user=client)

                sms = (
                    f"TICKET CLOSING\n\nTICKET No: {cardno}\n\nService Type: {service.service.name}\n\nMessage: {message}\n"
                    f"Final: use the link below to close or give further response to the ticket "
                    f"\n\nhttp://ocean.snedaghana.loc/servicing/jobcard/tracking/{cardno}/")

                print(DEFAULT_SMS_API)
                if SMS:
                    Sms(api_id=SMS_KEY, message=sms, to=client_details.phone).save()
                    service.client_approval = 1
                    ticket = service.ticket
                    service.save()
                    if ServiceCard.objects.filter(ticket_id=ticket).count() == 1:
                        service = ServiceCard.objects.get(ticket_id=ticket)
                        task = service.task
                        TaskTransactions(task=task, title=f"SENT TO CLOSING",
                                         description=f"Sent To {client.get_full_name()} for approval and closing",
                                         owner_id=request.user.pk).save()
                    TicketTrans(ticket_id=service.ticket_id,
                                tran=f"Sent To {client.get_full_name()} for approval withm message {message}",
                                title="CLIENT APPROVAL", user_id=request.user.pk).save()
                    success_response['message'] = f"Sent To {client.get_full_name()} for approval and closing"
                else:
                    success_response['message'] = "Could Not Send SMS"

            elif module == "service_technician":
                service_code = data.get('service_code')
                technician = data.get('technician')
                service = Services.objects.get(code=service_code)

                ServiceTechnicians(service=service, technician=User.objects.get(pk=technician)).save()



        # read data
        elif method == 'VIEW':
            if module == 'services':
                pk = data.get('pk') or '*'
                if pk == '*':
                    services = Services.objects.filter(status__gt=0)
                else:
                    services = Services.objects.filter(pk=pk, status__gt=0)

                if services.count() > 0:
                    arr = []
                    for service in services:
                        subs = SubServices.objects.filter(service=service)
                        technicians = ServiceTechnicians.objects.filter(service=service)

                        ss = []
                        tt = []
                        for sub in subs:
                            ss.append({
                                'pk': sub.pk,
                                'name': sub.name,
                                'desc': sub.description
                            })
                        for technician in technicians:
                            tt.append({
                                'pk': technician.technician.pk,
                                'name': technician.technician.get_full_name()
                            })
                        arr.append({
                            'pk': service.pk,
                            'code': service.code,
                            'description': service.description,
                            'subs': ss,
                            'technicians': tt

                        })
                        print(arr)
                    success_response['message'] = arr
                else:
                    success_response['status_code'] = 404
                    success_response['message'] = f"NO SERVICES AVAILABLE {pk}"

            elif module == 'servicecard':
                cardno = data.get('cardno')

                if ServiceCard.objects.filter(cardno=cardno).exists():
                    card = ServiceCard.objects.get(cardno=cardno)
                    technician = card.technician
                    cardpk = card.pk
                    print(cardpk)
                    materials = []
                    ticket_trans = []
                    if card.task is None:
                        uni = cardno
                    else:
                        uni = card.task.uni
                    for material in card.materials():
                        materials.append({
                            'line': material.line,
                            'item': material.item,
                            'description': material.description,
                            'price': material.price,
                            'quantity': material.quantity,
                            'total_price': material.total_price,

                        })

                    for ticktran in card.ticket.transactions():
                        ticket_trans.append({
                            'title': ticktran.title,
                            'descr': ticktran.tran,
                            'owner': {
                                'pk': ticktran.user.pk,
                                'username': ticktran.user.username,
                                'email': ticktran.user.email,
                                'fullname': f"{ticktran.user.first_name} {ticktran.user.last_name}",
                                'phone': UserAddOns.objects.get(user=ticktran.user).phone
                            },
                            'date': ticktran.created_on
                        })

                    header = {
                        'pk': card.pk,
                        'client': {
                            'pk': card.client.pk,
                            'username': card.client.username,
                            'email': card.client.email,
                            'fullname': f"{card.client.first_name} {card.client.last_name}",
                            'phone': UserAddOns.objects.get(user=card.client).phone
                        },
                        'annal':card.analysis,
                        'app': card.app_details(),
                        'owner': {
                            'pk': card.owner.pk,
                            'username': card.owner.username,
                            'email': card.owner.email,
                            'fullname': f"{card.owner.first_name} {card.owner.last_name}",
                            'phone': UserAddOns.objects.get(user=card.owner).phone
                        },
                        'remarks': card.remarks,
                        'service': {
                            'main': {
                                'code': card.service.code,
                                'name': card.service.name,
                                'description': card.service.description
                            },
                            'sub': {
                                'name': card.service_sub.name,
                                'description': card.service_sub.description
                            },
                        },
                        'technician': {
                            'pk': technician.pk,
                            'username': technician.technician.username,
                            'email': technician.technician.email,
                            'fullname': f"{technician.technician.get_full_name()}",
                            'phone': UserAddOns.objects.get(user=technician.technician).phone
                        },
                        'ticket': {
                            'hd': {
                                'pk': card.ticket.pk,
                                'title': card.ticket.title,
                                'description': card.ticket.descr,
                                'status': card.ticket.status,

                            },
                            'trans': ticket_trans
                        },
                        'importance': card.importance,
                        'status': card.status,
                        'timedata': {
                            'datecreated': card.created_date,
                            'timecreated': card.created_time,
                            'dateupdated': card.updated_date,
                            'timeupdated': card.updated_time
                        },
                        'materials': materials,
                        'next': {
                            'count': 0,
                            'code': 0
                        },
                        'prev': {
                            'count': 0,
                            'code': 0
                        },
                        'task': {
                            'uni': uni
                        }
                    }

                    if ServiceCard.objects.filter(id__gt=cardpk).exists():
                        header['next']['count'] = ServiceCard.objects.filter(id__gt=cardpk).count()
                        next_c = ServiceCard.objects.filter(id__gt=cardpk).first()
                        header['next']['nextcode'] = next_c.cardno

                    if ServiceCard.objects.filter(id__lt=cardpk).exists():
                        header['prev']['count'] = ServiceCard.objects.filter(id__lt=cardpk).count()
                        p_c = ServiceCard.objects.filter(id__lt=cardpk).last()

                        header['prev']['code'] = p_c.cardno

                    success_response['message'] = header

                else:
                    success_response['status_code'] = 404
                    success_response['message'] = "NO JOB CARD FOUND"

            elif module == 'findjob':
                jobstring = data.get('jobstring')
                cards = ServiceCard.objects.filter(Q(remarks__icontains=jobstring) | Q(cardno__icontains=jobstring) |
                                                   Q(ticket__title__icontains=jobstring) |
                                                   Q(technician__username__icontains=jobstring))

                if cards.count() > 0:
                    cardx = []
                    for card in cards:
                        cardx.append({
                            'cardno': card.cardno,
                            'title': card.ticket.title,
                            'description': card.ticket.descr,
                            'owner': card.client.get_full_name(),
                            'status': card.status,
                            'date': card.created_date
                        })
                    success_response['message'] = cardx
                else:
                    success_response['status_code'] = 404
                    success_response['message'] = f"{jobstring} does not match any record"

            elif module == 'service_report':
                target_user = data.get('user')
                target_status = data.get('status')
                target_from = data.get('from')
                target_to = data.get('to')
                doc = data.get('doc') or 'preview'
                cards = []

                # if target_user == '*' and target_status != '*':
                #     query = ServiceCard.objects.filter(status=target_status,
                #                                        created_date__range=(target_from, target_to))
                # elif target_user == '*' and target_status == '*':
                #     query = ServiceCard.objects.filter(created_date__range=(target_from, target_to))
                #
                # elif target_user != '*' and target_status == '*':
                #     query = ServiceCard.objects.filter(created_date__range=(target_from, target_to),
                #                                        client_id=target_user)
                # else:
                #     query = ServiceCard.objects.filter(status=target_status,
                #                                        created_date__range=(target_from, target_to),
                #                                        client_id=target_user)

                query = ServiceCard.objects.filter(
                    Q(client_id=target_user) if target_user != '*' else Q(),
                    Q(status=target_status) if target_status != '*' else Q(),
                    Q(created_date__range=(target_from, target_to))
                )
                UNATTENDED = 0
                OPEN = 0
                CLIENT_APPROVAL = 0
                CLOSED = 0

                if doc == 'excel':
                    import openpyxl
                    book = openpyxl.Workbook()
                    sheet = book.active

                    sheet.title = f"{target_from} To {target_to}"
                    sheet.merge_cells("A1:G1")
                    font_style = Font(size=20, bold=True, color='FFFFFF', )
                    fill_color = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
                    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    header = sheet['A1']
                    header.value = f"IT SUPPORT REPORT FROM {target_from} TO {target_to}"
                    header.font = font_style
                    header.fill = fill_color
                    header.alignment = align
                    sheet.row_dimensions[1].height = 34

                    sheet['A2'] = "DATE".upper()
                    sheet['B2'] = "TITLE".upper()
                    sheet['C2'] = "Service".upper()
                    sheet['D2'] = "Description".upper()
                    sheet['E2'] = 'Status'.upper()
                    sheet['F2'] = "Technician".upper()
                    sheet['G2'] = "Owner".upper()
                    sheet['H2'] = "Last Transaction"

                    sheet_count = 3

                for service in query:
                    if doc == 'preview':
                        cards.append({
                            'cardno': service.cardno,
                            'title': service.ticket.title,
                            'description': service.ticket.descr,
                            'service': f"{service.service.name}/{service.service_sub.name}",
                            'technician': service.technician.technician.get_full_name(),
                            'date': service.created_date,
                            'owner': service.owner.get_full_name()
                        })
                    elif doc == 'excel':
                        sheet[f"A{sheet_count}"] = service.created_date
                        sheet[f"B{sheet_count}"] = service.ticket.title
                        sheet[f"C{sheet_count}"] = f"{service.service.name}/{service.service_sub.name}"
                        sheet[f"D{sheet_count}"] = service.ticket.descr[:100]
                        sheet[f"E{sheet_count}"] = service.text_status()
                        sheet[f"F{sheet_count}"] = service.technician.technician.get_full_name()
                        sheet[f"G{sheet_count}"] = service.ticket.owner.get_full_name()
                        sheet[f"H{sheet_count}"] = service.last_transaction()[:100]
                        if service.text_status() == 'attending_to':
                            OPEN += 1
                        elif service.text_status() == 'sent_to_client':
                            CLIENT_APPROVAL += 1

                        elif service.text_status() == 'closed':
                            CLOSED += 1

                        elif service.text_status() == 'unattended':
                            UNATTENDED += 1

                        sheet_count += 1

                if doc == 'excel':
                    chat_sheet = book.create_sheet("CHART")
                    data = [
                        ["Category", "Value"],
                        ["OPEN", OPEN], ["CLIENT APPROVAL", CLIENT_APPROVAL], ["CLOSED", CLOSED],
                        ["UNATTENDED", UNATTENDED]
                    ]

                    for row in data:
                        chat_sheet.append(row)

                    chat = PieChart()
                    chat.title = "EXECUTION"
                    data2 = Reference(chat_sheet, min_col=2, min_row=1, max_col=2, max_row=len(data))
                    lebels = Reference(chat_sheet, min_col=1, min_row=2, max_row=len(data))
                    chat.add_data(data2, titles_from_data=True)
                    chat.set_categories(labels=lebels)
                    chat_sheet.add_chart(chat, 'D1')
                    file_name = f"static/general/tmp/SERVICING_REPORT.xlsx"
                    book.save(file_name)

                    cards = file_name

                success_response['message'] = cards

            elif module == 'material_request':
                cardno = data.get('cardno')
                send = data.get('send') or 'NO'
                card = ServiceCard.objects.get(cardno=cardno)
                materials = card.materials()
                cost = 0

                if materials.count() > 0:
                    cost = card.materials_cost()

                from fpdf import FPDF
                pdf = FPDF()
                pdf.add_page('P')

                pdf.set_font('Arial', 'B', 10)
                pdf.cell(190, 10, 'MATERIAL REQUEST', 0, 1, 'C')
                # job card
                pdf.set_font('Arial', 'B', 8)
                pdf.cell(20, 5, "Total Cost :  ", 0, 0)
                pdf.set_font('Arial', '', 8)
                pdf.cell(100, 5, f"{cost}", 0, 1)

                pdf.set_font('Arial', 'B', 8)
                pdf.cell(20, 5, "JOB CARD ", 0, 0)
                pdf.set_font('Arial', '', 8)
                pdf.cell(100, 5, card.cardno, 0, 1)

                # Type
                pdf.set_font('Arial', 'B', 8)
                pdf.cell(20, 5, "Type ", 0, 0)
                pdf.set_font('Arial', '', 8)
                pdf.cell(100, 5, f"{card.service.name}", 0, 1)

                # Title
                pdf.set_font('Arial', 'B', 8)
                pdf.cell(20, 5, "Title ", 0, 0)
                pdf.set_font('Arial', '', 8)
                pdf.cell(100, 5, card.ticket.title, 0, 1)
                # owner
                pdf.set_font('Arial', 'B', 8)
                pdf.cell(20, 5, "Reported By ", 0, 0)
                pdf.set_font('Arial', '', 8)
                pdf.cell(100, 5, f"{card.owner.get_full_name()}", 0, 1)

                # table header
                pdf.set_font('Arial', 'B', 8)
                pdf.cell(38, 5, "ITEM", 1, 0)
                pdf.cell(38, 5, "DESCRIPTION", 1, 0)
                pdf.cell(38, 5, "PRICE", 1, 0)
                pdf.cell(38, 5, "QTY", 1, 0)
                pdf.cell(38, 5, "TOTAL", 1, 1)

                pdf.set_font('Arial', '', 8)
                for material in card.materials():
                    pdf.cell(38, 5, material.item, 1, 0)
                    pdf.cell(38, 5, material.description, 1, 0)
                    pdf.cell(38, 5, f"{material.price}", 1, 0)
                    pdf.cell(38, 5, f"{material.quantity}", 1, 0)
                    pdf.cell(38, 5, f"{material.total_price}", 1, 1)

                file_name = f"static/general/servicing/{card.cardno}.pdf"
                pdf.output(file_name)
                if send == 'YES' and materials.count() > 0:
                    mail_subject = "Material Request for Technical Issue Resolution"
                    mail_message = f"""
                        Dear Jawara,
                        Attached is a material request document outlining the items needed to address a pressing technical issue we are currently facing.
                        The timely approval and procurement of these items are crucial for resolving the technical matter efficiently. Your prompt attention to this request is highly appreciated.
                        Please feel free to reach out if you require any further details or clarification.
                        <h3>Issue</h3>
                        <p><strong>Title</strong>{card.ticket.title}</p>
                        <p><strong>Description</strong>{card.ticket.descr}</p>
                        
                    """
                    if MailSenders.objects.filter(address='solomon@snedaghana.com').exists():
                        sender = MailSenders.objects.get(address='solomon@snedaghana.com')
                    else:
                        sender = MailSenders.objects.get(is_default=True)
                    mail = MailQueues(sender=sender, subject=mail_subject, body=mail_message,
                                      recipient='jawaratu@snedaghana.com',
                                      cc='ajay@snedaghana.com,bharat@snedaghana.com')
                    mail.save()
                    MailAttachments(mail=mail, attachment=file_name).save()
                success_response['message'] = file_name


        # update data
        elif method == 'PATCH':
            if module == 'jobcardstatus':
                cardno = data.get('cardno')
                status = data.get('status')
                service = ServiceCard.objects.get(cardno=cardno)
                ticket = service.ticket
                owner = service.owner
                owner_details = UserAddOns.objects.get(user=owner)

                service.status = status
                service.save()
                Sms(to=owner_details.phone, message=f"Ticket has been closed\nTICKET NO: {cardno}",
                    api=SmsApi.objects.get(is_default=1)).save()
                mesg = "DOCUMENT UPDATED"
                if status == 0:
                    mesg = "DOCUMENT DELETED"
                    ticket.status = status
                    ticket.save()
                elif status == 2:
                    mesg = "DOCUMENT CLOSED"
                    # close ticket
                    ticket.status = status
                    ticket.save()

                success_response['message'] = mesg

            elif module == 'client_ticket_approval':
                cardno = data.get('cardno')
                service = ServiceCard.objects.get(cardno=cardno)
                ticket = service.ticket
                message = data.get('message')
                status = data.get('status')
                client = service.client

                if status == 2:
                    # approved

                    msg = f"TICKET: {cardno}\nSTATUS: Closed by Client\nMessage: {message}"
                    ticket.status = 2
                    ticket.save()
                    service.status = 2
                    service.client_approval = 2


                else:
                    # not approved
                    service.status = 1
                    msg = f"TICKET: {cardno}\nSTATUS: Rejected by Client\nMessage: {message}"
                    service.status = 1
                    service.client_approval = 0

                try:
                    Sms(api_id=SMS_KEY, message=msg, to='0546310011').save()
                    TicketTrans(title="Client Feedback", tran=msg, ticket_id=service.ticket_id,
                                user_id=request.user.pk).save()
                    success_response['message'] = "Feedback Sent"
                    service.save()

                    if ServiceCard.objects.filter(ticket_id=ticket.pk).count() == 1:
                        service = ServiceCard.objects.get(ticket_id=ticket.pk)
                        task = service.task
                        TaskTransactions(task=task, title=f"Client Feedback",
                                         description=msg,
                                         owner_id=request.user.pk).save()
                        if status == 2:
                            task.status = 2
                            task.save()

                except Exception as e:
                    success_response['message'] = f"Feedback Not Sent {e}"

            elif module == 'change_service_app':
                card_pk = data.get('card_pk', None)
                app = data.get('app_id', None)

                card = ServiceCard.objects.get(pk=card_pk)
                jobcard = card.cardno
                app = App.objects.get(pk=app)
                card.app = app
                card.save()
                success_response['message'] = jobcard
                response = success_response

            elif module == 'close_sent':
                sent_to_client = ServiceCard.objects.filter(client_approval=1)
                obj = []
                for sent in sent_to_client:
                    client = sent.client
                    client_ad_on = UserAddOns.objects.get(user=client)
                    phone = client_ad_on.phone
                    title = sent.ticket.title
                    # get sent record
                    sent_tran = TicketTrans.objects.filter(ticket=sent.ticket,title='CLIENT APPROVAL')
                    # Get the current date and time in UTC
                    current_datetime = timezone.now()

                    # Calculate the time difference

                    if sent_tran.exists():
                        tran = sent_tran.last()
                        date_sent = tran.created_on
                        time_difference = current_datetime - date_sent
                        if time_difference.total_seconds() > 24 * 3600:
                            message = f"""
                            AUTO-CONFIRM\n\nTITLE:{title}\n\nIssue has been closed because it is past 24hours since 
                            it was completed and sent for your approval\nhttp://ocean.snedaghana.loc/servicing/jobcard/tracking/{sent.cardno}\n\n"""
                            new_sms(phone,message)
                            obj.append({
                                'title':title,
                                'sent_to':phone
                            })
                            sent.status = 2
                            sent.client_approval = 2
                            sent.save()
                success_response['message'] = obj



        # delete data
        elif method == 'DELETE':
            if module == 'services':
                service = data['service']
                serv = Services.objects.get(pk=service)
                name = serv.name
                serv.status = -1
                serv.save()

                success_response['message'] = f"{name} has been deleted"

            elif module == 'jobcard':
                cardno = data.get('cardno')
                service = ServiceCard.objects.get(cardno=cardno)
                owner = service.owner
                owner_details = UserAddOns.objects.get(user=owner)

                service.status = 0
                service.save()
                # Sms(to=owner_details.phone,message=f"Ticket has been closed\nTICKET NO: {cardno}",api=SmsApi.objects.get(is_default=1)).save()

                success_response['message'] = "Ticket Closed"



        # else
        else:
            success_response['status_code'] = 505
            success_response['message'] = f"Unknown  Request Method {method}"

        response = success_response


    except Exception as e:
        error_type, error_instance, traceback = sys.exc_info()
        tb_path = traceback.tb_frame.f_code.co_filename
        line_number = traceback.tb_lineno
        response["status_code"] = 500
        response[
            "message"] = f"An error of type {error_type} occurred on line {line_number} in file {tb_path}. Details: {e}"

    return JsonResponse(response)
